"""
Extract tables from PDF files using Azure Document Intelligence (layout model)
and save them as Excel files. Handles multi-page tables with repeated headers.

Dependencies: azure-ai-documentintelligence, azure-core, openpyxl, rich, python-dotenv
"""

import os
import sys
import re
import argparse
import time
import tkinter as tk
from tkinter import filedialog
from pathlib import Path
from dataclasses import dataclass
from dotenv import load_dotenv

load_dotenv(Path(__file__).resolve().parent / "backend" / ".env")

from azure.core.credentials import AzureKeyCredential
from azure.ai.documentintelligence import DocumentIntelligenceClient
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from rich.console import Console
from rich.table import Table
from rich.progress import Progress, SpinnerColumn, TextColumn, BarColumn, TaskProgressColumn, TimeElapsedColumn
from rich.panel import Panel
from rich.text import Text

console = Console()

# Number of header rows at the top of each table page.
HEADER_ROWS = 2

# Azure Document Intelligence Layout pricing: $10 per 1000 pages
COST_PER_PAGE = 10.0 / 1000


INSTITUTION_KEYWORDS = [
    'School', 'College', 'University', 'Academy', 'Institute',
    'Vidyalaya', 'Vidyalayam', 'Patasala', 'Hall', 'Mandapam',
    'Kalyanamandapam', 'Mahal', 'Choultry', 'Hospital', 'Dispensary',
    'Library', 'Madrasa', 'Madarasa', 'Seminary',
]

DESCRIPTOR_KEYWORDS = [
    'BUILDING', 'BLDG', 'PORTION', 'WING', 'BLOCK', 'FLOOR', 'ROOM',
    'SHED', 'SIDE', 'SECTION', 'ANNEX', 'EXTENSION', 'ADDL', 'ADDITIONAL',
    'NEW', 'OLD', 'MAIN', 'NORTH', 'SOUTH', 'EAST', 'WEST', 'LEFT', 'RIGHT',
    'UPPER', 'LOWER', 'GROUND', 'FIRST', 'SECOND', 'THIRD', 'FRONT', 'REAR',
    'BACK', 'CENTRAL', 'MIDDLE',
]

BOOTH_COLUMN_KEYWORDS = ['polling station', 'booth', 'building name', 'address', 'location']


def extract_booth_name(text):
    """Extract the core institution/building name from a polling station address."""
    if not text or not isinstance(text, str):
        return ""

    # Step 1: Take text before first comma
    text = text.split(",", 1)[0].strip()

    # Step 2: Remove pincode patterns (6-digit numbers, optionally preceded by dash/space)
    text = re.sub(r'[\s\-]*\d{6}\b', '', text).strip()

    # Step 3: Remove trailing parenthetical descriptors
    match = re.search(r'\(([^)]+)\)\s*$', text)
    if match:
        paren_content = match.group(1).upper()
        if any(kw in paren_content for kw in DESCRIPTOR_KEYWORDS):
            text = text[:match.start()].strip()

    # Step 4A: Standalone institution keyword — return text up to and including it
    for keyword in INSTITUTION_KEYWORDS:
        pattern = re.compile(r'^(.*?\b' + re.escape(keyword) + r'\b)', re.IGNORECASE)
        m = pattern.search(text)
        if m:
            return m.group(1).strip()

    # Step 4B: Dotted abbreviation at start (e.g. P.U.M.S, P.U.E.S, P.U.M.School)
    m = re.match(r'^((?:[A-Z]\.){2,}[A-Za-z]*)', text)
    if m:
        return m.group(1).strip()

    # Fallback: return cleaned text as-is
    return text.strip()


def add_booth_name_column(header_rows, data_rows, source_col_idx):
    """Insert a 'Booth Name' column immediately after the source column."""
    new_headers = []
    for i, row in enumerate(header_rows):
        new_row = list(row)
        label = "Booth Name" if i == 0 else ""
        new_row.insert(source_col_idx + 1, label)
        new_headers.append(new_row)

    new_data = []
    for row in data_rows:
        new_row = list(row)
        source_text = row[source_col_idx] if source_col_idx < len(row) else ""
        booth_name = extract_booth_name(source_text)
        new_row.insert(source_col_idx + 1, booth_name)
        new_data.append(new_row)

    return new_headers, new_data


def pick_booth_column(header_rows):
    """Display column headers and let user pick the source column interactively."""
    from rich.prompt import IntPrompt

    if not header_rows:
        return None

    # Use only the first header row (actual column names),
    # skip serial number rows and any other sub-headers
    first_row = header_rows[0]
    col_count = len(first_row)
    labels = []
    for c in range(col_count):
        val = str(first_row[c]).strip()
        labels.append(val if val else f"(Column {c + 1})")

    console.print()
    console.print("[bold]Select the source column for booth name extraction:[/bold]")
    table = Table(show_header=True, header_style="bold cyan")
    table.add_column("#", justify="right", width=4)
    table.add_column("Column Name")
    for i, label in enumerate(labels):
        table.add_row(str(i + 1), label)
    console.print(table)

    choice = IntPrompt.ask(
        "Enter column number",
        default=1,
        console=console,
    )
    idx = choice - 1
    if idx < 0 or idx >= col_count:
        console.print("[red]Invalid choice. Skipping booth name extraction.[/red]")
        return None
    console.print(f"[green]Selected:[/green] {labels[idx]}")
    return idx


@dataclass
class FileResult:
    filename: str
    pages: int = 0
    tables: int = 0
    rows: int = 0
    time_s: float = 0.0
    success: bool = False
    error: str = ""


def get_client():
    endpoint = os.environ.get("AZURE_DI_ENDPOINT")
    key = os.environ.get("AZURE_DI_KEY")
    if not endpoint or not key:
        console.print(
            "[bold red]Error:[/] Set AZURE_DI_ENDPOINT and AZURE_DI_KEY environment variables."
        )
        sys.exit(1)
    return DocumentIntelligenceClient(
        endpoint=endpoint, credential=AzureKeyCredential(key)
    )


def analyze_pdf(client, pdf_path):
    """Send a local PDF to Azure Document Intelligence and return the result."""
    with open(pdf_path, "rb") as f:
        poller = client.begin_analyze_document(
            "prebuilt-layout",
            body=f,
            content_type="application/octet-stream",
        )
    return poller.result()


def tables_to_rows(result):
    """Convert Azure DI tables into a single list of rows, deduplicating headers.

    First 2 rows of each table are headers. Only kept from the first table,
    skipped on subsequent tables to avoid duplicates.
    """
    if not result.tables:
        return [], []

    header_rows = []
    data_rows = []

    for table_idx, table in enumerate(result.tables):
        col_count = table.column_count
        grid = {}
        for cell in table.cells:
            grid[(cell.row_index, cell.column_index)] = cell.content

        max_row = table.row_count
        for r in range(max_row):
            row = [grid.get((r, c), "") for c in range(col_count)]
            if r < HEADER_ROWS:
                if table_idx == 0:
                    header_rows.append(row)
            else:
                data_rows.append(row)

    return header_rows, data_rows


def write_excel(header_rows, data_rows, output_path):
    """Write header + data rows to an Excel file with bold headers and auto-width."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Table"

    bold_font = Font(bold=True)

    for row in header_rows:
        ws.append(row)
    for r in range(1, len(header_rows) + 1):
        for cell in ws[r]:
            cell.font = bold_font

    for row in data_rows:
        ws.append(row)

    for col_idx in range(1, ws.max_column + 1):
        max_len = 0
        col_letter = get_column_letter(col_idx)
        for cell in ws[col_letter]:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_len + 2, 60)

    wb.save(output_path)


def process_pdf(client, pdf_path, output_path, progress, file_task_id, booth_col_idx=None):
    """Analyze a single PDF and save the extracted table as an Excel file."""
    filename = os.path.basename(pdf_path)
    start = time.perf_counter()
    result_info = FileResult(filename=filename)

    # Step 1: Upload & Analyze
    progress.update(file_task_id, description=f"[cyan]Analyzing [bold]{filename}[/bold]...")
    try:
        result = analyze_pdf(client, pdf_path)
    except Exception as e:
        result_info.error = str(e)
        result_info.time_s = time.perf_counter() - start
        return result_info

    page_count = len(result.pages) if result.pages else 0
    table_count = len(result.tables) if result.tables else 0
    result_info.pages = page_count
    result_info.tables = table_count

    if table_count == 0:
        result_info.error = "No tables found"
        result_info.time_s = time.perf_counter() - start
        return result_info

    # Step 2: Extract tables
    progress.update(file_task_id, description=f"[cyan]Extracting tables from [bold]{filename}[/bold]... ({page_count} pages)")
    header_rows, data_rows = tables_to_rows(result)

    # Step 2.5: Add booth name column if requested
    if booth_col_idx is not None:
        progress.update(file_task_id, description=f"[cyan]Extracting booth names from [bold]{filename}[/bold]...")
        header_rows, data_rows = add_booth_name_column(header_rows, data_rows, booth_col_idx)

    result_info.rows = len(header_rows) + len(data_rows)

    # Step 3: Write Excel
    progress.update(file_task_id, description=f"[cyan]Writing Excel for [bold]{filename}[/bold]...")
    try:
        write_excel(header_rows, data_rows, output_path)
    except Exception as e:
        result_info.error = f"Excel write failed: {e}"
        result_info.time_s = time.perf_counter() - start
        return result_info

    result_info.success = True
    result_info.time_s = time.perf_counter() - start
    return result_info


def run_batch(client, pdf_files, output_dir, booth_col_idx=None):
    """Process a list of PDF files with rich progress UI."""
    results = []
    total_start = time.perf_counter()

    console.print()
    with Progress(
        SpinnerColumn(),
        TextColumn("[progress.description]{task.description}"),
        BarColumn(),
        TaskProgressColumn(),
        TimeElapsedColumn(),
        console=console,
    ) as progress:
        batch_task = progress.add_task(
            f"[bold green]Processing PDFs", total=len(pdf_files)
        )
        file_task = progress.add_task("", total=None)  # spinner for current file

        for pdf_file in pdf_files:
            xlsx_path = output_dir / (pdf_file.stem + ".xlsx")
            info = process_pdf(client, str(pdf_file), str(xlsx_path), progress, file_task, booth_col_idx)
            results.append(info)

            if info.success:
                progress.update(file_task, description=f"[green]Done: {info.filename}")
            elif info.error:
                progress.update(file_task, description=f"[red]Failed: {info.filename}")
            progress.advance(batch_task)

        progress.update(file_task, visible=False)

    total_time = time.perf_counter() - total_start

    # --- Summary table ---
    table = Table(title="Results", show_lines=True)
    table.add_column("File", style="bold")
    table.add_column("Pages", justify="right")
    table.add_column("Tables", justify="right")
    table.add_column("Rows", justify="right")
    table.add_column("Time", justify="right")
    table.add_column("Status", justify="center")

    succeeded = 0
    failed = 0
    total_pages = 0
    total_rows = 0

    for r in results:
        total_pages += r.pages
        total_rows += r.rows
        time_str = f"{r.time_s:.1f}s"
        if r.success:
            succeeded += 1
            table.add_row(r.filename, str(r.pages), str(r.tables), str(r.rows), time_str, "[green]OK[/green]")
        else:
            failed += 1
            err = r.error[:40] + "..." if len(r.error) > 40 else r.error
            table.add_row(r.filename, str(r.pages), str(r.tables), str(r.rows), time_str, f"[red]{err}[/red]")

    console.print()
    console.print(table)

    # --- Stats line ---
    est_cost = total_pages * COST_PER_PAGE
    avg_per_page = (total_time / total_pages) if total_pages > 0 else 0

    stats = Text()
    stats.append(f"Files: ", style="bold")
    stats.append(f"{succeeded} succeeded", style="green")
    if failed:
        stats.append(f" | {failed} failed", style="red")
    stats.append(f"  |  Pages: {total_pages}", style="bold")
    stats.append(f"  |  Rows: {total_rows}", style="bold")
    stats.append(f"  |  Est. cost: ${est_cost:.2f}", style="bold yellow")
    stats.append(f"  |  Time: {total_time:.1f}s", style="bold")
    if total_pages > 0:
        stats.append(f" ({avg_per_page:.1f}s/page)", style="dim")

    console.print()
    console.print(Panel(stats, title="Summary", border_style="blue"))
    console.print()



def pick_pdf_files():
    """Open a Windows file picker dialog to select one or more PDF files."""
    root = tk.Tk()
    root.withdraw()
    root.attributes("-topmost", True)
    file_paths = filedialog.askopenfilenames(
        title="Select PDF file(s)",
        filetypes=[("PDF files", "*.pdf")],
    )
    root.destroy()
    if not file_paths:
        console.print("[yellow]No files selected.[/yellow]")
        sys.exit(0)
    return [Path(p) for p in file_paths]


def pick_output_folder():
    """Open a Windows folder picker dialog to choose where to save Excel files."""
    root = tk.Tk()
    root.withdraw()
    root.attributes("-topmost", True)
    folder = filedialog.askdirectory(title="Select output folder for Excel files")
    root.destroy()
    if not folder:
        console.print("[yellow]No output folder selected.[/yellow]")
        sys.exit(0)
    return Path(folder)


def main():
    parser = argparse.ArgumentParser(
        description="Extract tables from PDFs using Azure Document Intelligence."
    )
    parser.add_argument(
        "--input", required=False,
        help="Path to a PDF file or a folder of PDFs. "
             "If omitted, a file picker dialog will open."
    )
    parser.add_argument(
        "--output", required=False,
        help="Output .xlsx file path (single file) or folder (batch mode). "
             "If omitted, a folder picker dialog will open."
    )
    args = parser.parse_args()

    console.print(Panel("[bold]PDF Table Extractor[/bold]\nAzure Document Intelligence + Excel", border_style="blue"))

    # --- Interactive mode selector ---
    from rich.prompt import IntPrompt
    console.print()
    console.print("[bold]What would you like to do?[/bold]")
    mode_table = Table(show_header=True, header_style="bold cyan")
    mode_table.add_column("#", justify="right", width=4)
    mode_table.add_column("Mode")
    mode_table.add_column("Description")
    mode_table.add_row("1", "Booth Extraction", "Extract tables + add booth/institution name column")
    mode_table.add_row("2", "Standard Extraction", "Extract tables to Excel as-is")
    console.print(mode_table)

    mode_choice = IntPrompt.ask("Select mode", default=1, console=console)
    extract_booth = mode_choice == 1

    console.print()
    client = get_client()

    # --- Determine PDF files and output directory ---
    if args.input is None:
        pdf_files = pick_pdf_files()
        output_dir = Path(args.output) if args.output else pick_output_folder()
    else:
        input_path = Path(args.input)
        if input_path.is_file():
            if not input_path.suffix.lower() == ".pdf":
                console.print(f"[red]Error:[/red] {input_path} is not a PDF file.")
                sys.exit(1)
            pdf_files = [input_path]
        elif input_path.is_dir():
            pdf_files = sorted(input_path.glob("*.pdf"))
            if not pdf_files:
                console.print(f"[red]No PDF files found in {input_path}[/red]")
                sys.exit(1)
        else:
            console.print(f"[red]Error:[/red] {input_path} does not exist.")
            sys.exit(1)

        if args.output:
            output_dir = Path(args.output)
        elif input_path.is_file():
            output_dir = input_path.parent
        else:
            output_dir = input_path

    output_dir.mkdir(parents=True, exist_ok=True)

    console.print(f"[bold]{len(pdf_files)}[/bold] PDF file(s) selected")
    console.print(f"Output folder: [dim]{output_dir}[/dim]")

    # --- Booth name extraction: peek at first PDF to let user pick column ---
    booth_col_idx = None
    if extract_booth:
        console.print()
        console.print("[bold yellow]Booth name extraction enabled.[/bold yellow]")
        console.print("Peeking at first PDF to detect columns...")
        try:
            peek_result = analyze_pdf(client, str(pdf_files[0]))
            peek_headers, _ = tables_to_rows(peek_result)
            if peek_headers:
                booth_col_idx = pick_booth_column(peek_headers)
            else:
                console.print("[red]No headers found in first PDF. Skipping booth extraction.[/red]")
        except Exception as e:
            console.print(f"[red]Failed to peek at first PDF: {e}[/red]")
            console.print("[yellow]Continuing without booth extraction.[/yellow]")

    run_batch(client, pdf_files, output_dir, booth_col_idx)


if __name__ == "__main__":
    main()
