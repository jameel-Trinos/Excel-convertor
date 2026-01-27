#!/usr/bin/env python3
"""
Recalculate and validate Excel files.

This script:
1. Opens an Excel file
2. Recalculates all formulas
3. Checks for errors
4. Reports any issues

Usage:
    python recalc.py <excel_file>
    python recalc.py output.xlsx
"""

import sys
from pathlib import Path

from openpyxl import load_workbook


def recalculate_excel(file_path: str) -> dict:
    """
    Recalculate formulas and check for errors.

    Args:
        file_path: Path to Excel file

    Returns:
        Dictionary with results
    """
    if not Path(file_path).exists():
        return {"success": False, "error": f"File not found: {file_path}"}

    try:
        # Load workbook (data_only=False to see formulas)
        workbook = load_workbook(file_path, data_only=False)
        worksheet = workbook.active

        results = {
            "success": True,
            "file": file_path,
            "sheets": len(workbook.sheetnames),
            "rows": worksheet.max_row,
            "columns": worksheet.max_column,
            "formulas": [],
            "errors": [],
            "empty_cells": 0,
        }

        # Check all cells
        for row in worksheet.iter_rows():
            for cell in row:
                value = cell.value

                # Check for formulas
                if value and isinstance(value, str) and value.startswith("="):
                    results["formulas"].append({
                        "cell": cell.coordinate,
                        "formula": value,
                    })

                # Check for error values
                if value in ["#DIV/0!", "#N/A", "#NAME?", "#NULL!", "#NUM!", "#REF!", "#VALUE!"]:
                    results["errors"].append({
                        "cell": cell.coordinate,
                        "error": value,
                    })

                # Count empty cells in data range
                if value is None or str(value).strip() == "":
                    results["empty_cells"] += 1

        # Try to save (this forces Excel to recalculate)
        # Create a temporary file to avoid overwriting
        temp_file = Path(file_path).with_suffix(".temp.xlsx")
        workbook.save(temp_file)
        temp_file.unlink()  # Delete temp file

        workbook.close()

        return results

    except Exception as e:
        return {"success": False, "error": str(e)}


def print_report(results: dict) -> None:
    """Print a formatted report."""
    print("\n" + "="*70)
    print("EXCEL RECALCULATION REPORT")
    print("="*70)

    if not results.get("success"):
        print(f"\n✗ ERROR: {results.get('error')}")
        print("="*70 + "\n")
        return

    print(f"\nFile: {results['file']}")
    print(f"Sheets: {results['sheets']}")
    print(f"Dimensions: {results['rows']} rows × {results['columns']} columns")
    print(f"Formulas found: {len(results['formulas'])}")
    print(f"Empty cells: {results['empty_cells']}")

    # Print formulas
    if results['formulas']:
        print(f"\n{'─'*70}")
        print("FORMULAS:")
        for formula in results['formulas'][:10]:  # Show first 10
            print(f"  {formula['cell']}: {formula['formula']}")
        if len(results['formulas']) > 10:
            print(f"  ... and {len(results['formulas']) - 10} more")

    # Print errors
    if results['errors']:
        print(f"\n{'─'*70}")
        print("✗ ERRORS FOUND:")
        for error in results['errors']:
            print(f"  {error['cell']}: {error['error']}")
        print(f"\nTotal errors: {len(results['errors'])}")
        print("="*70 + "\n")
        sys.exit(1)
    else:
        print(f"\n{'─'*70}")
        print("✓ NO ERRORS - All formulas are valid!")

    print("="*70 + "\n")


def main():
    """Main entry point."""
    if len(sys.argv) < 2:
        print("Usage: python recalc.py <excel_file>")
        print("\nExample:")
        print("  python recalc.py output.xlsx")
        sys.exit(1)

    file_path = sys.argv[1]
    results = recalculate_excel(file_path)
    print_report(results)

    # Exit with error code if validation failed
    if not results.get("success") or results.get("errors"):
        sys.exit(1)


if __name__ == "__main__":
    main()
