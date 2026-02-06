"""AI-powered Excel structure analyzer for intelligent translation."""

import json
import logging
import os
from typing import Any, Dict, List, Optional, Tuple

import openpyxl
from anthropic import Anthropic

logger = logging.getLogger(__name__)


class ExcelStructureAnalyzer:
    """
    Use Claude AI to analyze Excel structure before translation.

    This ensures:
    - Proper identification of headers vs data
    - Preservation of table structure
    - Correct column mappings
    - Merged cell handling
    """

    def __init__(self, api_key: Optional[str] = None):
        """
        Initialize Excel structure analyzer with Claude AI.

        Args:
            api_key: Anthropic API key (reads from ANTHROPIC_API_KEY env if not provided)
        """
        self.api_key = api_key or os.getenv("ANTHROPIC_API_KEY")

        if self.api_key:
            self.client = Anthropic(api_key=self.api_key)
            self.model = "claude-sonnet-4-20250514"
            self.enabled = True
            logger.info("ExcelStructureAnalyzer initialized with Claude AI")
        else:
            self.client = None
            self.enabled = False
            logger.warning("No ANTHROPIC_API_KEY found, structure analysis disabled")

    async def analyze_structure(
        self,
        workbook_path: str,
        sheet_name: Optional[str] = None
    ) -> Dict[str, Any]:
        """
        Analyze Excel structure using Claude AI.

        Args:
            workbook_path: Path to Excel file
            sheet_name: Specific sheet to analyze (analyzes first sheet if None)

        Returns:
            Structure analysis results
        """
        if not self.enabled:
            logger.warning("Structure analysis disabled (no API key)")
            return self._fallback_analysis(workbook_path, sheet_name)

        try:
            # Load workbook
            wb = openpyxl.load_workbook(workbook_path)
            ws = wb[sheet_name] if sheet_name else wb.active

            # Extract structure information
            structure_data = self._extract_structure_data(ws)

            # Build prompt for Claude
            prompt = self._build_analysis_prompt(structure_data)

            # Call Claude AI
            response = self.client.messages.create(
                model=self.model,
                max_tokens=4096,
                temperature=0,
                messages=[{
                    "role": "user",
                    "content": prompt
                }]
            )

            # Parse response
            analysis = self._parse_claude_response(response.content[0].text)

            wb.close()

            logger.info(f"Structure analysis complete: {analysis.get('num_header_rows', 0)} header rows, "
                       f"{analysis.get('num_data_rows', 0)} data rows")

            return analysis

        except Exception as e:
            logger.error(f"Structure analysis failed: {e}")
            return self._fallback_analysis(workbook_path, sheet_name)

    def _extract_structure_data(self, worksheet) -> Dict[str, Any]:
        """Extract key structure information from worksheet."""
        # Get dimensions
        max_row = worksheet.max_row
        max_col = worksheet.max_column

        # Extract first 10 rows (usually contains headers)
        sample_rows = []
        for row_idx in range(1, min(11, max_row + 1)):
            row_data = []
            for col_idx in range(1, max_col + 1):
                cell = worksheet.cell(row=row_idx, column=col_idx)
                row_data.append({
                    "value": str(cell.value) if cell.value else "",
                    "is_merged": self._is_cell_in_merged_range(worksheet, cell),
                    "has_style": cell.has_style,
                    "font_bold": cell.font.bold if cell.font else False,
                    "fill_color": str(cell.fill.fgColor.rgb) if cell.fill and hasattr(cell.fill.fgColor, 'rgb') else None,
                })
            sample_rows.append(row_data)

        # Extract last 5 rows for pattern detection
        sample_end_rows = []
        for row_idx in range(max(max_row - 4, 11), max_row + 1):
            row_data = []
            for col_idx in range(1, max_col + 1):
                cell = worksheet.cell(row=row_idx, column=col_idx)
                row_data.append(str(cell.value) if cell.value else "")
            sample_end_rows.append(row_data)

        # Get merged cell ranges
        merged_ranges = [str(mr) for mr in worksheet.merged_cells.ranges]

        return {
            "total_rows": max_row,
            "total_columns": max_col,
            "sample_start_rows": sample_rows,
            "sample_end_rows": sample_end_rows,
            "merged_ranges": merged_ranges,
            "sheet_name": worksheet.title,
        }

    def _is_cell_in_merged_range(self, worksheet, cell) -> bool:
        """Check if cell is part of a merged range."""
        for merged_range in worksheet.merged_cells.ranges:
            if cell.coordinate in merged_range:
                return True
        return False

    def _build_analysis_prompt(self, structure_data: Dict[str, Any]) -> str:
        """Build prompt for Claude AI analysis."""
        prompt = f"""Analyze this Excel table structure to ensure proper translation:

**Sheet Information:**
- Sheet Name: {structure_data['sheet_name']}
- Total Rows: {structure_data['total_rows']}
- Total Columns: {structure_data['total_columns']}
- Merged Cell Ranges: {len(structure_data['merged_ranges'])}

**First 10 Rows (with cell metadata):**
```json
{json.dumps(structure_data['sample_start_rows'], indent=2)}
```

**Last 5 Rows:**
```json
{json.dumps(structure_data['sample_end_rows'], indent=2)}
```

**Merged Ranges:**
{json.dumps(structure_data['merged_ranges'], indent=2)}

**Your Task:**
Analyze this Excel structure and identify:

1. **Header Rows**: Which rows contain column headers? (Often rows 1-3, may be bold, colored, or merged)
2. **Title Rows**: Which rows contain the document title? (Often row 1, merged across all columns)
3. **Data Start Row**: Which row does the actual data start from?
4. **Column Structure**: What are the proper column headers? List them in order.
5. **Special Cells**: Are there any summary rows, total rows, or formula rows?
6. **Structure Type**: Is this a simple table, multi-header table, or complex nested structure?

**Important Notes:**
- Title rows are usually merged across all columns
- Header rows often have bold font or background color
- Data rows contain the actual content
- Some tables have multi-level headers (merged cells spanning multiple rows)

Return your analysis in JSON format:
{{
  "structure_type": "simple_table|multi_header|complex",
  "title_rows": [1, 2],  // Row numbers containing titles
  "header_rows": [3, 4],  // Row numbers containing column headers
  "data_start_row": 5,  // First row with actual data
  "column_headers": ["Header1", "Header2", ...],
  "special_rows": {{
    "totals": [100],  // Row numbers with totals/summaries
    "formulas": [101]  // Rows with formulas
  }},
  "merged_header_info": {{
    "description": "Rows 1-2 are merged title, row 3 has column headers"
  }},
  "translation_strategy": {{
    "translate_titles": true,
    "translate_headers": true,
    "translate_data": true,
    "preserve_formulas": true
  }},
  "warnings": ["List any structural issues or concerns"]
}}

Be precise and accurate. This analysis will guide the translation to preserve structure."""

        return prompt

    def _parse_claude_response(self, response_text: str) -> Dict[str, Any]:
        """Parse Claude's JSON response."""
        try:
            # Extract JSON from response (Claude sometimes wraps in ```json blocks)
            if "```json" in response_text:
                json_start = response_text.find("```json") + 7
                json_end = response_text.find("```", json_start)
                response_text = response_text[json_start:json_end].strip()
            elif "```" in response_text:
                json_start = response_text.find("```") + 3
                json_end = response_text.find("```", json_start)
                response_text = response_text[json_start:json_end].strip()

            analysis = json.loads(response_text)

            # Add derived fields
            analysis['num_header_rows'] = len(analysis.get('header_rows', []))
            analysis['num_title_rows'] = len(analysis.get('title_rows', []))
            analysis['num_data_rows'] = analysis.get('data_start_row', 1)

            return analysis

        except json.JSONDecodeError as e:
            logger.error(f"Failed to parse Claude response: {e}")
            logger.debug(f"Response text: {response_text}")
            return self._create_default_analysis()

    def _fallback_analysis(self, workbook_path: str, sheet_name: Optional[str]) -> Dict[str, Any]:
        """Fallback analysis without AI (basic heuristics)."""
        try:
            wb = openpyxl.load_workbook(workbook_path)
            ws = wb[sheet_name] if sheet_name else wb.active

            # Simple heuristic: First row is header, rest is data
            first_row_values = []
            for col_idx in range(1, ws.max_column + 1):
                cell = ws.cell(row=1, column=col_idx)
                first_row_values.append(str(cell.value) if cell.value else "")

            wb.close()

            return {
                "structure_type": "simple_table",
                "title_rows": [],
                "header_rows": [1],
                "data_start_row": 2,
                "column_headers": first_row_values,
                "special_rows": {"totals": [], "formulas": []},
                "translation_strategy": {
                    "translate_titles": True,
                    "translate_headers": True,
                    "translate_data": True,
                    "preserve_formulas": True
                },
                "warnings": ["Using fallback analysis (no AI)"],
                "num_header_rows": 1,
                "num_title_rows": 0,
                "num_data_rows": ws.max_row - 1
            }

        except Exception as e:
            logger.error(f"Fallback analysis failed: {e}")
            return self._create_default_analysis()

    def _create_default_analysis(self) -> Dict[str, Any]:
        """Create a safe default analysis."""
        return {
            "structure_type": "unknown",
            "title_rows": [],
            "header_rows": [1],
            "data_start_row": 2,
            "column_headers": [],
            "special_rows": {"totals": [], "formulas": []},
            "translation_strategy": {
                "translate_titles": True,
                "translate_headers": True,
                "translate_data": True,
                "preserve_formulas": True
            },
            "warnings": ["Using default analysis"],
            "num_header_rows": 1,
            "num_title_rows": 0,
            "num_data_rows": 0
        }

    async def validate_translated_structure(
        self,
        original_path: str,
        translated_path: str,
        original_analysis: Dict[str, Any],
        sheet_name: Optional[str] = None
    ) -> Dict[str, Any]:
        """
        Validate that translated Excel maintains structure.

        Args:
            original_path: Path to original Excel
            translated_path: Path to translated Excel
            original_analysis: Structure analysis from original
            sheet_name: Sheet to validate

        Returns:
            Validation results
        """
        if not self.enabled:
            return {"valid": True, "warnings": ["Validation disabled (no AI)"]}

        try:
            # Load both workbooks
            wb_orig = openpyxl.load_workbook(original_path)
            wb_trans = openpyxl.load_workbook(translated_path)

            ws_orig = wb_orig[sheet_name] if sheet_name else wb_orig.active
            ws_trans = wb_trans[sheet_name] if sheet_name else wb_trans.active

            # Compare structures
            issues = []

            # Check dimensions
            if ws_orig.max_row != ws_trans.max_row:
                issues.append(f"Row count mismatch: {ws_orig.max_row} vs {ws_trans.max_row}")

            if ws_orig.max_column != ws_trans.max_column:
                issues.append(f"Column count mismatch: {ws_orig.max_column} vs {ws_trans.max_column}")

            # Check merged cells
            orig_merged = set(str(mr) for mr in ws_orig.merged_cells.ranges)
            trans_merged = set(str(mr) for mr in ws_trans.merged_cells.ranges)

            if orig_merged != trans_merged:
                missing = orig_merged - trans_merged
                extra = trans_merged - orig_merged
                if missing:
                    issues.append(f"Missing merged cells: {missing}")
                if extra:
                    issues.append(f"Extra merged cells: {extra}")

            # Check header structure
            header_rows = original_analysis.get('header_rows', [1])
            for row_idx in header_rows:
                for col_idx in range(1, min(ws_orig.max_column, ws_trans.max_column) + 1):
                    orig_cell = ws_orig.cell(row=row_idx, column=col_idx)
                    trans_cell = ws_trans.cell(row=row_idx, column=col_idx)

                    # Check if cell has content (translated or original)
                    if orig_cell.value and not trans_cell.value:
                        issues.append(f"Empty cell in translation at row {row_idx}, col {col_idx}")

            wb_orig.close()
            wb_trans.close()

            valid = len(issues) == 0

            return {
                "valid": valid,
                "issues": issues,
                "warnings": [] if valid else ["Structure validation found issues"]
            }

        except Exception as e:
            logger.error(f"Validation failed: {e}")
            return {
                "valid": False,
                "issues": [f"Validation error: {str(e)}"],
                "warnings": ["Validation failed"]
            }
