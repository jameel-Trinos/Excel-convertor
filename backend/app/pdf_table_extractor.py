"""
Extract tables from PDF files using Azure Document Intelligence (layout model)
and save them as Excel files. Handles multi-page tables with repeated headers.

Dependencies: azure-ai-documentintelligence, azure-core, openpyxl, rich, python-dotenv
"""

import os
import sys
import argparse
import time
import tkinter as tk
from tkinter import filedialog
from pathlib import Path
from dataclasses import dataclass
from dotenv import load_dotenv

load_dotenv()

from azure.core.credentials import AzureKeyCredential
from azure.ai.documentintelligence import DocumentIntelligenceClient
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from rich.console import Console
from rich.table import Table
from rich.progress import Progress, SpinnerColumn, TextColumn, BarColumn, TaskProgressColumn, TimeElapsedColumn
from rich.panel import Panel
from rich.text import Text

console = Console()

# Number of header rows repeated on each page of the PDF table.
HEADER_ROWS = 2

# Azure Document Intelligence Layout pricing: $10 per 1000 pages
COST_PER_PAGE = 10.0 / 1000


@dataclass
class FileResult:
    filename: str
    pages: int = 0
    tables: int = 0
    rows: int = 0
    time_s: float = 0.0
    success: bool = False
    error: str = ""


def get_client():
    endpoint = os.environ.get("AZURE_DI_ENDPOINT")
    key = os.environ.get("AZURE_DI_KEY")
    if not endpoint or not key:
        console.print(
            "[bold red]Error:[/] Set AZURE_DI_ENDPOINT and AZURE_DI_KEY environment variables."
        )
        sys.exit(1)
    return DocumentIntelligenceClient(
        endpoint=endpoint, credential=AzureKeyCredential(key)
    )


def _send_to_azure(client, pdf_path):
    """Send a PDF to Azure DI and return the raw result."""
    with open(pdf_path, "rb") as f:
        poller = client.begin_analyze_document(
            "prebuilt-layout",
            body=f,
            content_type="application/octet-stream",
        )
    return poller.result()


def _fix_rotated_pdf(pdf_path, page_angles):
    """Create a corrected PDF using rotation angles detected by Azure DI.

    Args:
        pdf_path: Original PDF path.
        page_angles: dict of {page_index: angle} for pages that need fixing.

    Returns:
        temp_file_path — caller must delete after use.
    """
    import tempfile
    from pypdf import PdfReader, PdfWriter

    reader = PdfReader(pdf_path)
    writer = PdfWriter()

    for i, page in enumerate(reader.pages):
        if i in page_angles:
            angle = page_angles[i]
            # Round to nearest 90° and counter-rotate
            correction = -round(angle / 90) * 90
            console.print(f"  [yellow]Page {i+1}: detected {angle}° — applying {correction}° correction[/]")
            page.rotate(correction)
        writer.add_page(page)

    tmp = tempfile.NamedTemporaryFile(suffix=".pdf", delete=False)
    writer.write(tmp)
    tmp.close()
    return tmp.name


def analyze_pdf(client, pdf_path):
    """Send a PDF to Azure DI with automatic rotation correction.

    Pass 1: Send original PDF, check page angles from the result.
    Pass 2: If any page is rotated (|angle| > 5°), correct with pypdf
             and re-send for accurate table extraction.
    """
    # Pass 1 — detect rotation
    result = _send_to_azure(client, pdf_path)

    if not result.pages:
        return result

    # Check which pages are rotated
    rotated = {}
    for page in result.pages:
        angle = getattr(page, "angle", 0) or 0
        if abs(angle) > 5:
            rotated[page.page_number - 1] = angle  # 0-indexed

    if not rotated:
        return result  # No rotation — pass 1 result is good

    # Pass 2 — fix and re-send
    console.print(
        f"[yellow]Rotated pages detected (pages {[i+1 for i in sorted(rotated)]}) "
        f"— correcting and re-analysing...[/]"
    )
    temp_file = _fix_rotated_pdf(pdf_path, rotated)
    try:
        return _send_to_azure(client, temp_file)
    finally:
        os.unlink(temp_file)


def tables_to_rows(result):
    """Convert Azure DI tables into a single list of rows, deduplicating headers."""
    if not result.tables:
        return [], []

    header_rows = []
    data_rows = []

    for table_idx, table in enumerate(result.tables):
        col_count = table.column_count
        grid = {}
        for cell in table.cells:
            grid[(cell.row_index, cell.column_index)] = cell.content

        max_row = table.row_count
        for r in range(max_row):
            row = [grid.get((r, c), "") for c in range(col_count)]
            if r < HEADER_ROWS:
                if table_idx == 0:
                    header_rows.append(row)
            else:
                data_rows.append(row)

    return header_rows, data_rows


def write_excel(header_rows, data_rows, output_path):
    """Write header + data rows to an Excel file with bold headers and auto-width."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Table"

    bold_font = Font(bold=True)

    for row in header_rows:
        ws.append(row)
    for r in range(1, len(header_rows) + 1):
        for cell in ws[r]:
            cell.font = bold_font

    for row in data_rows:
        ws.append(row)

    for col_idx in range(1, ws.max_column + 1):
        max_len = 0
        col_letter = get_column_letter(col_idx)
        for cell in ws[col_letter]:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_len + 2, 60)

    wb.save(output_path)


def process_pdf(client, pdf_path, output_path, progress, file_task_id):
    """Analyze a single PDF and save the extracted table as an Excel file."""
    filename = os.path.basename(pdf_path)
    start = time.perf_counter()
    result_info = FileResult(filename=filename)

    # Step 1: Upload & Analyze
    progress.update(file_task_id, description=f"[cyan]Analyzing [bold]{filename}[/bold]...")
    try:
        result = analyze_pdf(client, pdf_path)
    except Exception as e:
        result_info.error = str(e)
        result_info.time_s = time.perf_counter() - start
        return result_info

    page_count = len(result.pages) if result.pages else 0
    table_count = len(result.tables) if result.tables else 0
    result_info.pages = page_count
    result_info.tables = table_count

    if table_count == 0:
        result_info.error = "No tables found"
        result_info.time_s = time.perf_counter() - start
        return result_info

    # Step 2: Extract tables
    progress.update(file_task_id, description=f"[cyan]Extracting tables from [bold]{filename}[/bold]... ({page_count} pages)")
    header_rows, data_rows = tables_to_rows(result)
    result_info.rows = len(header_rows) + len(data_rows)

    # Step 3: Write Excel
    progress.update(file_task_id, description=f"[cyan]Writing Excel for [bold]{filename}[/bold]...")
    try:
        write_excel(header_rows, data_rows, output_path)
    except Exception as e:
        result_info.error = f"Excel write failed: {e}"
        result_info.time_s = time.perf_counter() - start
        return result_info

    result_info.success = True
    result_info.time_s = time.perf_counter() - start
    return result_info


def run_batch(client, pdf_files, output_dir):
    """Process a list of PDF files with rich progress UI."""
    results = []
    total_start = time.perf_counter()

    console.print()
    with Progress(
        SpinnerColumn(),
        TextColumn("[progress.description]{task.description}"),
        BarColumn(),
        TaskProgressColumn(),
        TimeElapsedColumn(),
        console=console,
    ) as progress:
        batch_task = progress.add_task(
            f"[bold green]Processing PDFs", total=len(pdf_files)
        )
        file_task = progress.add_task("", total=None)  # spinner for current file

        for pdf_file in pdf_files:
            xlsx_path = output_dir / (pdf_file.stem + ".xlsx")
            info = process_pdf(client, str(pdf_file), str(xlsx_path), progress, file_task)
            results.append(info)

            if info.success:
                progress.update(file_task, description=f"[green]Done: {info.filename}")
            elif info.error:
                progress.update(file_task, description=f"[red]Failed: {info.filename}")
            progress.advance(batch_task)

        progress.update(file_task, visible=False)

    total_time = time.perf_counter() - total_start

    # --- Summary table ---
    table = Table(title="Results", show_lines=True)
    table.add_column("File", style="bold")
    table.add_column("Pages", justify="right")
    table.add_column("Tables", justify="right")
    table.add_column("Rows", justify="right")
    table.add_column("Time", justify="right")
    table.add_column("Status", justify="center")

    succeeded = 0
    failed = 0
    total_pages = 0
    total_rows = 0

    for r in results:
        total_pages += r.pages
        total_rows += r.rows
        time_str = f"{r.time_s:.1f}s"
        if r.success:
            succeeded += 1
            table.add_row(r.filename, str(r.pages), str(r.tables), str(r.rows), time_str, "[green]OK[/green]")
        else:
            failed += 1
            err = r.error[:40] + "..." if len(r.error) > 40 else r.error
            table.add_row(r.filename, str(r.pages), str(r.tables), str(r.rows), time_str, f"[red]{err}[/red]")

    console.print()
    console.print(table)

    # --- Stats line ---
    est_cost = total_pages * COST_PER_PAGE
    avg_per_page = (total_time / total_pages) if total_pages > 0 else 0

    stats = Text()
    stats.append(f"Files: ", style="bold")
    stats.append(f"{succeeded} succeeded", style="green")
    if failed:
        stats.append(f" | {failed} failed", style="red")
    stats.append(f"  |  Pages: {total_pages}", style="bold")
    stats.append(f"  |  Rows: {total_rows}", style="bold")
    stats.append(f"  |  Est. cost: ${est_cost:.2f}", style="bold yellow")
    stats.append(f"  |  Time: {total_time:.1f}s", style="bold")
    if total_pages > 0:
        stats.append(f" ({avg_per_page:.1f}s/page)", style="dim")

    console.print()
    console.print(Panel(stats, title="Summary", border_style="blue"))
    console.print()



def pick_pdf_files():
    """Open a Windows file picker dialog to select one or more PDF files."""
    root = tk.Tk()
    root.withdraw()
    root.attributes("-topmost", True)
    file_paths = filedialog.askopenfilenames(
        title="Select PDF file(s)",
        filetypes=[("PDF files", "*.pdf")],
    )
    root.destroy()
    if not file_paths:
        console.print("[yellow]No files selected.[/yellow]")
        sys.exit(0)
    return [Path(p) for p in file_paths]


def pick_output_folder():
    """Open a Windows folder picker dialog to choose where to save Excel files."""
    root = tk.Tk()
    root.withdraw()
    root.attributes("-topmost", True)
    folder = filedialog.askdirectory(title="Select output folder for Excel files")
    root.destroy()
    if not folder:
        console.print("[yellow]No output folder selected.[/yellow]")
        sys.exit(0)
    return Path(folder)


def main():
    parser = argparse.ArgumentParser(
        description="Extract tables from PDFs using Azure Document Intelligence."
    )
    parser.add_argument(
        "--input", required=False,
        help="Path to a PDF file or a folder of PDFs. "
             "If omitted, a file picker dialog will open."
    )
    parser.add_argument(
        "--output", required=False,
        help="Output .xlsx file path (single file) or folder (batch mode). "
             "If omitted, a folder picker dialog will open."
    )
    args = parser.parse_args()

    console.print(Panel("[bold]PDF Table Extractor[/bold]\nAzure Document Intelligence + Excel", border_style="blue"))

    client = get_client()

    # --- Determine PDF files and output directory ---
    if args.input is None:
        pdf_files = pick_pdf_files()
        output_dir = Path(args.output) if args.output else pick_output_folder()
    else:
        input_path = Path(args.input)
        if input_path.is_file():
            if not input_path.suffix.lower() == ".pdf":
                console.print(f"[red]Error:[/red] {input_path} is not a PDF file.")
                sys.exit(1)
            pdf_files = [input_path]
        elif input_path.is_dir():
            pdf_files = sorted(input_path.glob("*.pdf"))
            if not pdf_files:
                console.print(f"[red]No PDF files found in {input_path}[/red]")
                sys.exit(1)
        else:
            console.print(f"[red]Error:[/red] {input_path} does not exist.")
            sys.exit(1)

        if args.output:
            output_dir = Path(args.output)
        elif input_path.is_file():
            output_dir = input_path.parent
        else:
            output_dir = input_path

    output_dir.mkdir(parents=True, exist_ok=True)

    console.print(f"[bold]{len(pdf_files)}[/bold] PDF file(s) selected")
    console.print(f"Output folder: [dim]{output_dir}[/dim]")

    run_batch(client, pdf_files, output_dir)


if __name__ == "__main__":
    main()
