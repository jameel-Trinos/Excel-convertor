"""Excel formatting utilities for professional spreadsheet styling."""

from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from .utils import sanitize_text


class ExcelFormatter:
    """
    Professional Excel formatting matching the specification.

    Formatting includes:
    - Header style: Dark blue (#366092) background, white bold text, centered
    - All cells: Thin black borders on all sides, center-aligned
    - Data cells: Default font (Calibri 11pt), no background
    - TOTAL row: Bold text, SUM formulas for numeric columns
    - Column width: Auto-fit based on content
    """

    # Color constants
    HEADER_BG_COLOR = "366092"  # Dark blue
    HEADER_FONT_COLOR = "FFFFFF"  # White
    TOTAL_BG_COLOR = "D9E2F3"  # Light blue for total row

    def __init__(self):
        """Initialize the formatter with predefined styles."""
        # Header cell fill
        self.header_fill = PatternFill(
            start_color=self.HEADER_BG_COLOR,
            end_color=self.HEADER_BG_COLOR,
            fill_type="solid",
        )

        # Header font
        self.header_font = Font(
            bold=True,
            color=self.HEADER_FONT_COLOR,
            size=11,
            name="Calibri",
        )

        # Data font
        self.data_font = Font(
            bold=False,
            color="000000",
            size=11,
            name="Calibri",
        )

        # Bold font for totals
        self.bold_font = Font(
            bold=True,
            color="000000",
            size=11,
            name="Calibri",
        )

        # Center alignment with wrap text and proper text direction
        self.center_alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True,
            text_rotation=0,  # No rotation
            readingOrder=1,  # Left-to-right (1=LTR, 2=RTL, 0=context)
        )

        # Center alignment without wrap (for data cells) and proper text direction
        self.center_alignment_no_wrap = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=False,
            text_rotation=0,  # No rotation
            readingOrder=1,  # Left-to-right (1=LTR, 2=RTL, 0=context)
        )

        # Thin black border on all sides
        thin_side = Side(style="thin", color="000000")
        self.thin_border = Border(
            left=thin_side,
            right=thin_side,
            top=thin_side,
            bottom=thin_side,
        )

        # Total row fill
        self.total_fill = PatternFill(
            start_color=self.TOTAL_BG_COLOR,
            end_color=self.TOTAL_BG_COLOR,
            fill_type="solid",
        )

    def format_header_row(
        self,
        worksheet: Worksheet,
        row_number: int,
        start_col: int = 1,
        end_col: int = None,
    ) -> None:
        """
        Apply header formatting to a row.

        Args:
            worksheet: The worksheet to format
            row_number: The row number to format (1-indexed)
            start_col: Starting column (1-indexed)
            end_col: Ending column (1-indexed), None for max column
        """
        if end_col is None:
            end_col = worksheet.max_column

        for col in range(start_col, end_col + 1):
            cell = worksheet.cell(row=row_number, column=col)
            cell.fill = self.header_fill
            cell.font = self.header_font
            cell.alignment = self.center_alignment
            cell.border = self.thin_border

    def format_data_cells(
        self,
        worksheet: Worksheet,
        start_row: int,
        end_row: int,
        start_col: int = 1,
        end_col: int = None,
    ) -> None:
        """
        Apply data cell formatting to a range.

        Args:
            worksheet: The worksheet to format
            start_row: Starting row (1-indexed)
            end_row: Ending row (1-indexed)
            start_col: Starting column (1-indexed)
            end_col: Ending column (1-indexed), None for max column
        """
        if end_col is None:
            end_col = worksheet.max_column

        for row in range(start_row, end_row + 1):
            for col in range(start_col, end_col + 1):
                cell = worksheet.cell(row=row, column=col)
                cell.font = self.data_font
                # Data cells don't need wrap_text (prevents unnecessary height)
                cell.alignment = self.center_alignment_no_wrap
                cell.border = self.thin_border

    def format_total_row(
        self,
        worksheet: Worksheet,
        row_number: int,
        start_col: int = 1,
        end_col: int = None,
    ) -> None:
        """
        Apply total row formatting.

        Args:
            worksheet: The worksheet to format
            row_number: The row number to format (1-indexed)
            start_col: Starting column (1-indexed)
            end_col: Ending column (1-indexed), None for max column
        """
        if end_col is None:
            end_col = worksheet.max_column

        for col in range(start_col, end_col + 1):
            cell = worksheet.cell(row=row_number, column=col)
            cell.fill = self.total_fill
            cell.font = self.bold_font
            cell.alignment = self.center_alignment
            cell.border = self.thin_border

    def auto_fit_columns(
        self,
        worksheet: Worksheet,
        min_width: int = 10,
        max_width: int = 50,
    ) -> None:
        """
        Adjust column widths based on content.

        Args:
            worksheet: The worksheet to adjust
            min_width: Minimum column width in characters
            max_width: Maximum column width in characters
        """
        for column in worksheet.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)

            for cell in column:
                try:
                    if cell.value:
                        cell_length = len(str(cell.value))
                        if cell_length > max_length:
                            max_length = cell_length
                except Exception:
                    pass

            # Calculate adjusted width with padding
            adjusted_width = min(max(max_length + 2, min_width), max_width)
            worksheet.column_dimensions[column_letter].width = adjusted_width

    def set_fixed_column_widths(
        self,
        worksheet: Worksheet,
        default_width: int = 16,
        column_widths: dict[int, int] = None,
    ) -> None:
        """
        Set fixed column widths to prevent collapsed appearance.

        This is the RECOMMENDED method for professional Excel files.
        Auto-fit can cause columns to be too narrow or too wide.

        Args:
            worksheet: The worksheet to adjust
            default_width: Default width for all columns (in characters)
            column_widths: Optional dict mapping column numbers to specific widths
        """
        max_col = worksheet.max_column

        for col in range(1, max_col + 1):
            col_letter = get_column_letter(col)

            # Use specific width if provided, otherwise use default
            if column_widths and col in column_widths:
                width = column_widths[col]
            else:
                width = default_width

            worksheet.column_dimensions[col_letter].width = width

    def set_row_heights(
        self,
        worksheet: Worksheet,
        header_row: int,
        data_start_row: int,
        data_end_row: int,
        header_height: int = 70,
        data_height: int = 18,
    ) -> None:
        """
        Set proper row heights to prevent collapsed appearance.

        CRITICAL: This prevents the Excel file from looking squished.

        Args:
            worksheet: The worksheet to adjust
            header_row: Row number of the header
            data_start_row: First data row
            data_end_row: Last data row
            header_height: Height for header row (default 70 for wrapped text)
            data_height: Height for data rows (default 18 for breathing room)
        """
        # Set header row height (needs to be taller for wrapped text)
        worksheet.row_dimensions[header_row].height = header_height

        # Set all data rows to consistent height
        for row in range(data_start_row, data_end_row + 1):
            worksheet.row_dimensions[row].height = data_height

    def apply_number_formatting(
        self,
        worksheet: Worksheet,
        start_row: int,
        end_row: int,
        start_col: int = 1,
        end_col: int = None,
    ) -> None:
        """
        Apply number formatting (thousand separators) to numeric cells.

        Args:
            worksheet: The worksheet to format
            start_row: Starting row (1-indexed)
            end_row: Ending row (1-indexed)
            start_col: Starting column (1-indexed)
            end_col: Ending column (1-indexed), None for max column
        """
        if end_col is None:
            end_col = worksheet.max_column

        for row in range(start_row, end_row + 1):
            for col in range(start_col, end_col + 1):
                cell = worksheet.cell(row=row, column=col)
                value = cell.value

                # Check if value is numeric
                if value is not None:
                    try:
                        # Try to convert to number
                        if isinstance(value, str):
                            # Remove commas and try to convert
                            cleaned = value.replace(",", "").strip()
                            if cleaned.replace(".", "").replace("-", "").isdigit():
                                num_value = float(cleaned)
                                if num_value == int(num_value):
                                    cell.value = int(num_value)
                                else:
                                    cell.value = num_value
                                cell.number_format = "#,##0"
                        elif isinstance(value, (int, float)):
                            cell.number_format = "#,##0"
                    except (ValueError, TypeError):
                        pass

    def add_title_section(
        self,
        worksheet: Worksheet,
        title: str,
        subtitle: str = None,
        start_row: int = 1,
        num_columns: int = None,
    ) -> int:
        """
        Add a title section at the top of the worksheet.

        Args:
            worksheet: The worksheet to modify
            title: Main title text
            subtitle: Optional subtitle text
            start_row: Row to start the title (1-indexed)
            num_columns: Number of columns to merge, None for max column

        Returns:
            The next available row after the title section
        """
        if num_columns is None:
            num_columns = max(worksheet.max_column, 5)

        # Title cell
        title_cell = worksheet.cell(row=start_row, column=1)
        title_cell.value = title
        title_cell.font = Font(bold=True, size=14, name="Calibri")
        title_cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
            text_rotation=0,
            readingOrder=1,
        )

        # Merge cells for title
        worksheet.merge_cells(
            start_row=start_row,
            start_column=1,
            end_row=start_row,
            end_column=num_columns,
        )

        next_row = start_row + 1

        # Subtitle if provided
        if subtitle:
            subtitle_cell = worksheet.cell(row=next_row, column=1)
            subtitle_cell.value = subtitle
            subtitle_cell.font = Font(italic=True, size=11, name="Calibri")
            subtitle_cell.alignment = Alignment(
                horizontal="center",
                vertical="center",
                text_rotation=0,
                readingOrder=1,
            )

            worksheet.merge_cells(
                start_row=next_row,
                start_column=1,
                end_row=next_row,
                end_column=num_columns,
            )
            next_row += 1

        # Add empty row for spacing
        next_row += 1

        return next_row

    def add_multi_row_title_section(
        self,
        worksheet: Worksheet,
        title_rows: list[list[str]],
        start_row: int = 1,
        num_columns: int = None,
    ) -> int:
        """
        Add a multi-row title section with merged cells matching PDF structure.

        Args:
            worksheet: The worksheet to modify
            title_rows: List of title rows from PDF (each row is list of cells)
            start_row: Row to start the titles (1-indexed)
            num_columns: Total number of columns in the sheet

        Returns:
            The next available row after the title section
        """
        if num_columns is None:
            num_columns = max(worksheet.max_column, 5)

        current_row = start_row

        for title_row in title_rows:
            # Find the first non-empty cell
            non_empty = [cell for cell in title_row if cell.strip()]

            if non_empty:
                # Write the first non-empty value (sanitized to remove RTL characters)
                cell = worksheet.cell(row=current_row, column=1)
                cell.value = sanitize_text(non_empty[0])
                cell.font = Font(bold=True, size=12, name="Calibri")
                cell.alignment = Alignment(
                    horizontal="center",
                    vertical="center",
                    text_rotation=0,
                    readingOrder=1,
                )

                # Merge across all columns
                worksheet.merge_cells(
                    start_row=current_row,
                    start_column=1,
                    end_row=current_row,
                    end_column=num_columns,
                )

            current_row += 1

        # Add empty row for spacing
        current_row += 1

        return current_row

    def add_multi_row_headers(
        self,
        worksheet: Worksheet,
        header_rows: list[list[str]],
        start_row: int,
    ) -> int:
        """
        Add multi-row headers with proper formatting.

        Args:
            worksheet: The worksheet to modify
            header_rows: List of header rows (each row is list of cells)
            start_row: Row to start the headers (1-indexed)

        Returns:
            The next available row after the header section
        """
        current_row = start_row

        for header_row in header_rows:
            for col_idx, value in enumerate(header_row, 1):
                cell = worksheet.cell(row=current_row, column=col_idx)
                # Sanitize header value to remove RTL/bidirectional control characters
                cell.value = sanitize_text(value) if value else ""

            # Format this header row
            self.format_header_row(
                worksheet,
                row_number=current_row,
                start_col=1,
                end_col=len(header_row),
            )

            current_row += 1

        return current_row
