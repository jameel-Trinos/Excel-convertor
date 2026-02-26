"""Simplified FastAPI application for PDF to Excel Converter."""

import asyncio
import json
import logging
import os
import re
import tempfile
import uuid
from datetime import datetime
from pathlib import Path
from typing import Dict

import aiofiles
from dotenv import load_dotenv
from fastapi import BackgroundTasks, FastAPI, File, HTTPException, Query, UploadFile
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse
from sse_starlette.sse import EventSourceResponse

from .column_filter import ColumnFilterService
from .constituency_excel_creator import ConstituencyExcelCreator
from .constituency_processor import ConstituencyProcessor
from .excel_creator import ExcelCreator
from .models import (
    AddBoothNameColumnRequest,
    ConversionTask,
    DownloadModifiedRequest,
    ErrorResponse,
    FilterColumnsRequest,
    FilterColumnsResponse,
    FilterExcelRequest,
    FullPreviewData,
    GeocodeApplyRequest,
    GeocodeApplyResponse,
    GeocodeProgressEvent,
    GeocodeRequest,
    GeocodeStartResponse,
    PreviewData,
    ProgressEvent,
    StatusResponse,
    TranslateProgressEvent,
    TranslateRequest,
    TranslateStartResponse,
    TranslateStatusResponse,
    UploadResponse,
    NormalizeColumnRequest,
)
from .geocoding_service import GeocodingService, extract_addresses_from_column
from .pdf_processor import PDFProcessor
from .translation_service import TranslationAPIError
from .utils import cleanup_file, validate_pdf_file, sanitize_text, clean_excel_filename

import openpyxl
from .party_normalizer import PartyNormalizer

# Load environment variables from .env file
load_dotenv()

# Configure logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# Configuration
UPLOAD_DIR = Path(os.getenv("UPLOAD_DIR", "./uploads"))
OUTPUT_DIR = Path(os.getenv("OUTPUT_DIR", "./outputs"))
MAX_FILE_SIZE = int(os.getenv("MAX_FILE_SIZE", 50 * 1024 * 1024))  # 50MB default
ALLOWED_ORIGINS = os.getenv("ALLOWED_ORIGINS", "http://localhost:3000,http://localhost:3001").split(",")

# Ensure directories exist
UPLOAD_DIR.mkdir(parents=True, exist_ok=True)
OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

# Shared task storage (imported from task_store to allow router access)
from .task_store import tasks, geocode_tasks, translation_tasks, voter_convert_jobs, bulk_voter_jobs, update_task_progress

# Booth name extraction helpers

# Keywords in parenthetical booth descriptors (e.g., "(ADDL. BUILDING)")
_PAREN_DESCRIPTOR_KEYWORDS = {
    "BUILDING", "BLDG", "PORTION", "WING", "BLOCK", "FLOOR",
    "ROOM", "SHED", "SIDE", "SECTION", "ANNEX", "ANNEXE", "EXTENSION",
    "EXT", "PART", "UNIT", "ADDL", "ADDITIONAL", "NEW", "OLD", "MAIN",
    "NORTH", "SOUTH", "EAST", "WEST", "LEFT", "RIGHT", "UPPER", "LOWER",
    "GROUND", "FIRST", "SECOND", "THIRD", "FRONT", "REAR", "BACK",
    "CENTRAL", "MIDDLE",
}

# Institution type keywords — truncate after the LAST occurrence found.
# Works for all forms: "Higher Secondary School", "P.U.M.School", "Aided.E.School"
_INSTITUTION_RE = re.compile(
    r'\b('
    r'School|College|University|Academy|Institute|'
    r'Vidyalaya|Vidyalayam|Patasala|'
    r'Hall|Mandapam|Kalyanamandapam|Mahal|Choultry|'
    r'Hospital|Dispensary|Library|'
    r'Madrasa|Madarasa|Seminary'
    r')\b',
    re.IGNORECASE,
)

# Dotted abbreviation pattern at start of string (e.g., P.U.M.S, P.U.E.S, P.U.M.School)
_ABBREVIATION_RE = re.compile(r'^([A-Z]\.){2,}[A-Za-z]*', re.IGNORECASE)


def extract_booth_name(source_str: str) -> str:
    """Extract only the institution/building name from election booth location string.

    Examples:
      "P.U.M.S Thalavadi-638461 South Facing Terraced Building East Side"
        → "P.U.M.S"
      "Govt. Boys Higher Secondary School Athani Road Sathyamangalam 638401East Facing..."
        → "Govt. Boys Higher Secondary School"
      "P.U.M.School Thottagajanur-638461 East Facing terraced Building Right Side"
        → "P.U.M.School"
      "PANCHAYAT UNION PRIMARY SCHOOL (ADDL. BUILDING), THENGUMARAHADA - 638451"
        → "PANCHAYAT UNION PRIMARY SCHOOL"

    Strategy:
      1. Clean text: comma split, remove pincodes, remove descriptor parentheticals
      2. Extract institution name using one of two methods:
         a. Standalone institution keyword found (School, College, Hall, etc.)
            → truncate after it
         b. Dotted abbreviation at start (P.U.M.S, P.U.E.S, etc.)
            → extract just the abbreviation
    """
    text = source_str.strip()
    if not text:
        return ""

    # Step 1: Take text before first comma
    if "," in text:
        text = text.split(",")[0].strip()

    # Step 2: Remove pincode patterns (6-digit numbers)
    text = re.sub(r'[-\u2013]\s*\d{6}', '', text)
    text = re.sub(r'\b\d{6}(?=[A-Za-z\s]|$)', '', text)
    text = re.sub(r'\s+', ' ', text).strip()

    # Step 3: Remove trailing parenthetical if it contains a descriptor keyword
    paren_match = re.search(r'\s*\(([^)]+)\)\s*$', text)
    if paren_match:
        paren_content = paren_match.group(1).upper()
        tokens = re.split(r'[\s.,]+', paren_content)
        if any(tok in _PAREN_DESCRIPTOR_KEYWORDS for tok in tokens):
            text = text[:paren_match.start()].strip()

    # Step 4: Extract institution name
    # Method A: Standalone institution keyword (e.g., "Higher Secondary School")
    inst_match = _INSTITUTION_RE.search(text)
    if inst_match:
        return text[:inst_match.end()].strip()

    # Method B: Dotted abbreviation at start (e.g., "P.U.M.S", "P.U.E.S", "P.U.M.School")
    abbr_match = _ABBREVIATION_RE.match(text)
    if abbr_match:
        return abbr_match.group(0).strip()

    # Fallback: return cleaned text as-is
    return text


# Create FastAPI app
app = FastAPI(
    title="PDF to Excel Converter API",
    description="Convert tabular PDFs to professionally formatted Excel spreadsheets",
    version="1.0.0",
)

# CORS middleware
app.add_middleware(
    CORSMiddleware,
    allow_origins=ALLOWED_ORIGINS,
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# Include routers


async def process_conversion(
    task_id: str,
    file_path: Path,
    use_constituency_processor: bool = False,
    force_ocr: bool = False,
):
    """Simplified background task for PDF to Excel conversion."""
    try:
        tasks[task_id].status = "processing"
        tasks[task_id].message = "Starting conversion..."

        update_task_progress(task_id, 10, "Extracting tables from PDF...")
        logger.info(
            f"Processing task {task_id} (constituency={use_constituency_processor}, force_ocr={force_ocr})"
        )

        # Use PDFProcessor for extraction in both cases
        # force_ocr=True: always use OCR path with higher DPI (election / complex image PDFs)
        processor = PDFProcessor(str(file_path), force_ocr=force_ocr)

        def progress_callback(progress: int, message: str):
            # Scale extraction progress to 10-80%
            scaled_progress = 10 + int(progress * 0.7)
            update_task_progress(task_id, scaled_progress, message)

        extraction_result = await processor.extract_tables(progress_callback=progress_callback)

        if not extraction_result.tables or all(t.is_empty for t in extraction_result.tables):
            error_msg = "No tables found in the constituency PDF" if use_constituency_processor else "No tables found in the PDF"
            raise ValueError(error_msg)

        update_task_progress(task_id, 85, "Creating Excel file...")

        # Extract AC number from page texts (only for election results, not constituency)
        ac_number = None
        if not use_constituency_processor and extraction_result.page_texts:
            from .ac_extractor import extract_ac_number
            ac_number = extract_ac_number(extraction_result.page_texts)
            if ac_number:
                logger.info(f"Extracted AC number: {ac_number}")
            else:
                logger.warning("Could not extract AC number from PDF")

        # Create Excel file - use constituency creator if using constituency processor
        if use_constituency_processor:
            creator = ConstituencyExcelCreator()
        else:
            creator = ExcelCreator()
        
        output_filename = f"{task_id}.xlsx"
        output_path = OUTPUT_DIR / output_filename

        # Pass AC number to ExcelCreator (only for election results)
        if use_constituency_processor:
            await asyncio.to_thread(
                creator.create_from_tables,
                extraction_result.tables,
                str(output_path),
                source_filename=tasks[task_id].filename,
            )
        else:
            await asyncio.to_thread(
                creator.create_from_tables,
                extraction_result.tables,
                str(output_path),
                source_filename=tasks[task_id].filename,
                ac_number=ac_number,
            )

        update_task_progress(task_id, 95, "Finalizing...")

        # Mark as complete or needs_review based on validation (e.g. OCR quality)
        vr = processor.validation_report
        if vr and (not vr.passed or (vr.confidence < 0.7)):
            tasks[task_id].status = "needs_review"
            tasks[task_id].progress = 100
            tasks[task_id].message = (
                "Conversion completed with validation warnings. Please review the data."
            )
            tasks[task_id].output_file = str(output_path)
            tasks[task_id].validation_issues = {
                "passed": vr.passed,
                "confidence": round(vr.confidence, 3),
                "issues": vr.issues[:50],
                "warnings": getattr(vr, "warnings", [])[:20],
                "suggestions": getattr(vr, "suggestions", [])[:10],
            }
            logger.info(
                f"Task {task_id}: needs_review (passed={vr.passed}, confidence={vr.confidence:.2f}, "
                f"issues={len(vr.issues)})"
            )
        else:
            tasks[task_id].status = "completed"
            tasks[task_id].progress = 100
            tasks[task_id].message = "Conversion completed successfully"
            tasks[task_id].output_file = str(output_path)
            logger.info(f"Conversion completed for task {task_id}: {output_path}")

    except Exception as e:
        logger.error(f"Conversion failed for task {task_id}: {e}", exc_info=True)
        tasks[task_id].status = "failed"
        tasks[task_id].error = str(e)
        tasks[task_id].message = f"Conversion failed: {str(e)}"

    finally:
        # Cleanup uploaded PDF
        await cleanup_file(file_path)


@app.get("/health")
async def health_check():
    """Health check endpoint."""
    return {"status": "healthy", "timestamp": datetime.utcnow().isoformat()}




@app.post("/api/upload", response_model=UploadResponse)
async def upload_pdf(
    file: UploadFile = File(...),
    background_tasks: BackgroundTasks = None,
    force_ocr: bool = Query(False, description="Always use OCR with higher DPI for complex/image election PDFs"),
):
    """
    Upload a PDF file for conversion to Excel.

    - **file**: PDF file to convert (max 50MB)
    - **force_ocr**: If true, always use OCR extraction with higher DPI (for complex/image election PDFs)

    Returns task ID for tracking conversion progress.
    """
    # Validate file
    if not file.filename.lower().endswith(".pdf"):
        raise HTTPException(
            status_code=400,
            detail="Only PDF files are accepted",
        )

    # Read file content
    content = await file.read()

    # Check file size
    if len(content) > MAX_FILE_SIZE:
        raise HTTPException(
            status_code=400,
            detail=f"File too large. Maximum size is {MAX_FILE_SIZE // 1024 // 1024}MB",
        )

    # Validate PDF content
    validation_error = validate_pdf_file(content)
    if validation_error:
        raise HTTPException(status_code=400, detail=validation_error)

    # Generate task ID
    task_id = str(uuid.uuid4())

    # Save file
    file_path = UPLOAD_DIR / f"{task_id}.pdf"
    async with aiofiles.open(file_path, "wb") as f:
        await f.write(content)

    # Create task
    tasks[task_id] = ConversionTask(
        task_id=task_id,
        filename=file.filename,
        status="pending",
        progress=0,
        message="File uploaded, starting conversion...",
    )

    # Start background conversion (force_ocr for election/complex image PDFs)
    background_tasks.add_task(process_conversion, task_id, file_path, False, force_ocr)

    return UploadResponse(
        task_id=task_id,
        filename=file.filename,
        size=len(content),
        message="File uploaded successfully. Conversion started.",
    )


@app.post("/api/booth/upload", response_model=UploadResponse)
async def upload_booth_pdf(
    file: UploadFile = File(...),
    background_tasks: BackgroundTasks = None,
):
    """
    Upload a PDF file for booth-specific conversion to Excel.

    NOTE: Extraction logic has been cleared. Ready for new implementation.

    - **file**: PDF file to convert (max 50MB)

    Returns task ID for tracking conversion progress.
    """
    # Validate file
    if not file.filename.lower().endswith(".pdf"):
        raise HTTPException(
            status_code=400,
            detail="Only PDF files are accepted",
        )

    # Read file content
    content = await file.read()

    # Check file size
    if len(content) > MAX_FILE_SIZE:
        raise HTTPException(
            status_code=400,
            detail=f"File too large. Maximum size is {MAX_FILE_SIZE // 1024 // 1024}MB",
        )

    # Validate PDF content
    validation_error = validate_pdf_file(content)
    if validation_error:
        raise HTTPException(status_code=400, detail=validation_error)

    # Generate task ID
    task_id = str(uuid.uuid4())

    # Save file
    file_path = UPLOAD_DIR / f"{task_id}.pdf"
    async with aiofiles.open(file_path, "wb") as f:
        await f.write(content)

    # Create task
    tasks[task_id] = ConversionTask(
        task_id=task_id,
        filename=file.filename,
        status="pending",
        progress=0,
        message="File uploaded, starting booth conversion...",
    )

    # Start background conversion with constituency processor
    background_tasks.add_task(process_conversion, task_id, file_path, True)

    return UploadResponse(
        task_id=task_id,
        filename=file.filename,
        size=len(content),
        message="File uploaded successfully. Booth conversion started.",
    )


# ---------------------------------------------------------------------------
# Voters PDF upload & conversion
# ---------------------------------------------------------------------------

@app.post("/api/voters/upload", response_model=UploadResponse)
async def upload_voters_pdf(
    file: UploadFile = File(...),
    background_tasks: BackgroundTasks = None,
):
    """
    Upload a voter list PDF for extraction to Excel.

    Extracts voter data (Serial No, Name, Father/Husband Name, House No,
    Age, Gender, Voter ID) with bilingual (Tamil + English) support.
    """
    if not file.filename.lower().endswith(".pdf"):
        raise HTTPException(status_code=400, detail="Only PDF files are accepted")

    content = await file.read()

    if len(content) > MAX_FILE_SIZE:
        raise HTTPException(
            status_code=400,
            detail=f"File too large. Maximum size is {MAX_FILE_SIZE // 1024 // 1024}MB",
        )

    validation_error = validate_pdf_file(content)
    if validation_error:
        raise HTTPException(status_code=400, detail=validation_error)

    task_id = str(uuid.uuid4())

    file_path = UPLOAD_DIR / f"{task_id}.pdf"
    async with aiofiles.open(file_path, "wb") as f:
        await f.write(content)

    tasks[task_id] = ConversionTask(
        task_id=task_id,
        filename=file.filename,
        status="pending",
        progress=0,
        message="File uploaded, starting voter data extraction...",
    )

    background_tasks.add_task(process_voters_conversion, task_id, file_path)

    return UploadResponse(
        task_id=task_id,
        filename=file.filename,
        size=len(content),
        message="File uploaded successfully. Voter extraction started.",
    )


async def process_voters_conversion(task_id: str, file_path: Path):
    """Background task: extract voter data from PDF and create Excel."""
    try:
        tasks[task_id].status = "processing"
        tasks[task_id].message = "Starting voter data extraction..."

        from .voters_pdf_processor import VotersPDFProcessor
        from .voters_excel_creator import VotersExcelCreator

        processor = VotersPDFProcessor(str(file_path))

        def progress_cb(progress: int, message: str):
            scaled = 5 + int(progress * 0.75)
            update_task_progress(task_id, min(scaled, 85), message)

        result = await asyncio.to_thread(processor.extract, progress_cb)

        update_task_progress(task_id, 88, "Creating Excel file...")

        header_info = result["header_info"]
        creator = VotersExcelCreator()
        output_path = OUTPUT_DIR / f"{task_id}.xlsx"

        await asyncio.to_thread(
            creator.create,
            headers=result["headers"],
            rows=result["voters"],
            output_path=str(output_path),
            ac_no=header_info.ac_no,
            part_no=header_info.part_no,
            address=header_info.address,
            total_voters=header_info.total_voters,
            source_filename=tasks[task_id].filename,
        )

        update_task_progress(task_id, 95, "Finalizing...")

        voter_count = len(result["voters"])
        if voter_count == 0:
            tasks[task_id].status = "needs_review"
            tasks[task_id].progress = 100
            tasks[task_id].message = "No voter records found. The PDF format may not be recognized."
            tasks[task_id].output_file = str(output_path)
            tasks[task_id].validation_issues = {
                "passed": False,
                "confidence": 0.0,
                "issues": ["No voter records extracted from PDF"],
                "suggestions": [
                    "Ensure the PDF is an Indian electoral roll (voter list)",
                    "Try a PDF with clearer text or fewer scanned pages",
                ],
            }
        else:
            tasks[task_id].status = "completed"
            tasks[task_id].progress = 100
            tasks[task_id].message = f"Extraction complete: {voter_count} voters from {result['total_pages']} pages"
            tasks[task_id].output_file = str(output_path)
            logger.info(f"Voters extraction done for task {task_id}: {voter_count} voters")

    except Exception as e:
        logger.error(f"Voters extraction failed for task {task_id}: {e}", exc_info=True)
        tasks[task_id].status = "failed"
        tasks[task_id].error = str(e)
        tasks[task_id].message = f"Extraction failed: {str(e)}"

    finally:
        await cleanup_file(file_path)


@app.post("/api/voters/convert-pdf")
async def convert_voters_pdf(
    pdf_file: UploadFile = File(...),
    background_tasks: BackgroundTasks = None,
):
    """
    Async voter list PDF → Excel conversion with progress tracking.

    Returns a job_id immediately. Poll /api/voters/convert-status/{job_id}
    for progress. Download via /api/voters/download/{job_id} when completed.
    """
    if not pdf_file.filename or not pdf_file.filename.lower().endswith(".pdf"):
        raise HTTPException(status_code=400, detail="Only PDF files are accepted")

    content = await pdf_file.read()

    max_size = 50 * 1024 * 1024  # 50 MB
    if len(content) > max_size:
        raise HTTPException(status_code=400, detail="File too large. Maximum size is 50MB")

    validation_error = validate_pdf_file(content)
    if validation_error:
        raise HTTPException(status_code=400, detail=validation_error)

    job_id = str(uuid.uuid4())
    temp_pdf = UPLOAD_DIR / f"{job_id}.pdf"

    async with aiofiles.open(temp_pdf, "wb") as f:
        await f.write(content)

    voter_convert_jobs[job_id] = {
        "status": "processing",
        "progress": {"current_page": 0, "total_pages": 0},
        "download_url": None,
        "output_file": None,
        "filename": pdf_file.filename,
        "error": None,
    }

    background_tasks.add_task(_process_voter_convert_job, job_id, temp_pdf, pdf_file.filename)

    return {"job_id": job_id}


async def _process_voter_convert_job(job_id: str, temp_pdf: Path, original_filename: str):
    """Background task for async voter PDF conversion."""
    output_xlsx = OUTPUT_DIR / f"{job_id}.xlsx"

    try:
        from .voters_pdf_processor import VotersPDFProcessor
        from .voters_excel_creator import VotersExcelCreator

        processor = VotersPDFProcessor(str(temp_pdf))

        def progress_cb(progress: int, message: str):
            job = voter_convert_jobs.get(job_id)
            if not job:
                return
            # Parse page info from message like "Parsed page 3/20 (...)"
            import re as _re
            m = _re.search(r'page\s+(\d+)\s*/\s*(\d+)', message)
            if m:
                job["progress"]["current_page"] = int(m.group(1))
                job["progress"]["total_pages"] = int(m.group(2))
            elif "Opening" in message or "Starting" in message:
                job["progress"]["current_page"] = 0
            # Also try to get total_pages from "Parsing N pages"
            m2 = _re.search(r'Parsing\s+(\d+)\s+pages', message)
            if m2:
                job["progress"]["total_pages"] = int(m2.group(1))

        result = await asyncio.to_thread(processor.extract, progress_cb)

        voter_rows = result["voters"]
        header_info = result["header_info"]

        if not voter_rows:
            voter_convert_jobs[job_id]["status"] = "failed"
            voter_convert_jobs[job_id]["error"] = "No voter records found. Ensure the PDF is an Indian electoral roll."
            return

        # Update progress for Excel creation phase
        job = voter_convert_jobs[job_id]
        job["progress"]["current_page"] = job["progress"]["total_pages"]

        creator = VotersExcelCreator()
        await asyncio.to_thread(
            creator.create,
            headers=result["headers"],
            rows=voter_rows,
            output_path=str(output_xlsx),
            ac_no=header_info.ac_no,
            part_no=header_info.part_no,
            address=header_info.address,
            total_voters=header_info.total_voters,
            source_filename=original_filename,
        )

        ac = header_info.ac_no or "unknown"
        booth = header_info.part_no or "unknown"

        voter_convert_jobs[job_id]["status"] = "completed"
        voter_convert_jobs[job_id]["output_file"] = str(output_xlsx)
        voter_convert_jobs[job_id]["download_url"] = f"/api/voters/download/{job_id}"

        logger.info(
            f"Voter convert job {job_id}: {len(voter_rows)} voters, "
            f"AC={ac}, Booth={booth}, file={original_filename}"
        )

    except Exception as e:
        logger.error(f"Voter convert job {job_id} failed: {e}", exc_info=True)
        voter_convert_jobs[job_id]["status"] = "failed"
        voter_convert_jobs[job_id]["error"] = str(e)
    finally:
        await cleanup_file(temp_pdf)


@app.get("/api/voters/convert-status/{job_id}")
async def get_voter_convert_status(job_id: str):
    """
    Get the status of a voter PDF conversion job.

    Returns status, page-level progress, and download URL when completed.
    """
    job = voter_convert_jobs.get(job_id)
    if not job:
        raise HTTPException(status_code=404, detail="Job not found")

    response = {
        "status": job["status"],
        "progress": job["progress"],
    }

    if job["status"] == "completed":
        response["download_url"] = job["download_url"]
    elif job["status"] == "failed":
        response["error"] = job["error"]

    return response


@app.get("/api/voters/download/{job_id}")
async def download_voter_convert(job_id: str):
    """
    Download the Excel file from a completed voter conversion job.

    Cleans up the output file after download.
    """
    job = voter_convert_jobs.get(job_id)
    if not job:
        raise HTTPException(status_code=404, detail="Job not found")

    if job["status"] != "completed":
        raise HTTPException(status_code=400, detail="Conversion not yet completed")

    output_path = job.get("output_file")
    if not output_path or not Path(output_path).exists():
        raise HTTPException(status_code=404, detail="Output file not found")

    # Build download filename from original
    original = job.get("filename", "voter_list.pdf")
    download_name = original.rsplit(".", 1)[0] + ".xlsx"

    bg = BackgroundTasks()
    bg.add_task(cleanup_file, Path(output_path))
    bg.add_task(lambda: voter_convert_jobs.pop(job_id, None))

    return FileResponse(
        path=output_path,
        filename=download_name,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        background=bg,
    )


# ── Bulk Voter Processing Endpoints ────────────────────────────────────


@app.post("/api/voters/bulk-upload/init")
async def init_bulk_upload():
    """Initialize a bulk voter upload job. Returns a job_id for subsequent calls."""
    job_id = str(uuid.uuid4())
    bulk_voter_jobs[job_id] = {
        "status": "uploading",
        "files": [],
        "progress": {
            "total_pdfs": 0,
            "completed_pdfs": 0,
            "current_file": "",
            "total_voters_so_far": 0,
            "failed_count": 0,
        },
        "output_file": None,
        "download_url": None,
        "error": None,
        "summary": None,
    }
    return {"job_id": job_id}


@app.post("/api/voters/bulk-upload/add/{job_id}")
async def add_bulk_files(
    job_id: str,
    files: list[UploadFile] = File(...),
):
    """Upload a batch of PDFs to an existing bulk job. Call multiple times."""
    job = bulk_voter_jobs.get(job_id)
    if not job:
        raise HTTPException(status_code=404, detail="Bulk job not found")
    if job["status"] != "uploading":
        raise HTTPException(status_code=400, detail="Job is no longer accepting uploads")

    added = 0
    for f in files:
        if not f.filename or not f.filename.lower().endswith(".pdf"):
            continue

        content = await f.read()
        idx = len(job["files"])
        file_path = UPLOAD_DIR / f"{job_id}_bulk_{idx}.pdf"
        async with aiofiles.open(file_path, "wb") as out:
            await out.write(content)

        job["files"].append((idx, str(file_path), f.filename))
        added += 1

    return {"added": added, "total": len(job["files"])}


@app.post("/api/voters/bulk-upload/start/{job_id}")
async def start_bulk_processing(
    job_id: str,
    background_tasks: BackgroundTasks = None,
):
    """Start concurrent processing of all uploaded PDFs for this bulk job."""
    job = bulk_voter_jobs.get(job_id)
    if not job:
        raise HTTPException(status_code=404, detail="Bulk job not found")
    if job["status"] != "uploading":
        raise HTTPException(status_code=400, detail="Job already started or finished")
    if not job["files"]:
        raise HTTPException(status_code=400, detail="No PDF files uploaded yet")

    job["status"] = "processing"
    job["progress"]["total_pdfs"] = len(job["files"])

    background_tasks.add_task(_process_bulk_voter_job, job_id)

    return {"job_id": job_id, "total_files": len(job["files"])}


async def _process_bulk_voter_job(job_id: str):
    """Background task: process all PDFs concurrently and create consolidated Excel."""
    job = bulk_voter_jobs.get(job_id)
    if not job:
        return

    pdf_files = job["files"]
    uploaded_paths = [Path(p) for _, p, _ in pdf_files]

    try:
        from .bulk_voters_processor import BulkVotersProcessor
        from .bulk_voters_excel_creator import BulkVotersExcelCreator

        processor = BulkVotersProcessor()  # uses smart default based on CPU count + BULK_VOTER_WORKERS env

        def on_progress(info: dict):
            j = bulk_voter_jobs.get(job_id)
            if not j:
                return
            j["progress"]["completed_pdfs"] = info["completed"]
            j["progress"]["current_file"] = info["current_file"]
            j["progress"]["total_voters_so_far"] += info["voter_count"]
            if info["had_error"]:
                j["progress"]["failed_count"] += 1

        result = await asyncio.to_thread(processor.process_all, pdf_files, on_progress)

        # Create consolidated Excel (one sheet per booth)
        output_path = OUTPUT_DIR / f"{job_id}_bulk.xlsx"
        creator = BulkVotersExcelCreator()
        await asyncio.to_thread(
            creator.create,
            booth_data=result["booth_data"],
            booth_groups=result["booth_groups"],
            output_path=str(output_path),
            ac_no=result["ac_no"],
            total_pdfs=result["total_pdfs"],
            successful_pdfs=result["successful_pdfs"],
            failed_pdfs=result["failed_pdfs"],
        )

        job["status"] = "completed"
        job["output_file"] = str(output_path)
        job["download_url"] = f"/api/voters/bulk-download/{job_id}"
        job["summary"] = {
            "total_voters": result["total_voters"],
            "successful_pdfs": result["successful_pdfs"],
            "total_pdfs": result["total_pdfs"],
            "failed_pdfs": result["failed_pdfs"],
            "booth_groups": result["booth_groups"],
        }

        logger.info(
            f"Bulk voter job {job_id}: {result['total_voters']} voters from "
            f"{result['successful_pdfs']}/{result['total_pdfs']} PDFs"
        )

    except Exception as e:
        logger.error(f"Bulk voter job {job_id} failed: {e}", exc_info=True)
        job["status"] = "failed"
        job["error"] = str(e)

    finally:
        for p in uploaded_paths:
            await cleanup_file(p)


@app.get("/api/voters/bulk-status/{job_id}")
async def get_bulk_voter_status(job_id: str):
    """Poll status of a bulk voter processing job."""
    job = bulk_voter_jobs.get(job_id)
    if not job:
        raise HTTPException(status_code=404, detail="Bulk job not found")

    response = {
        "status": job["status"],
        "progress": job["progress"],
    }

    if job["status"] == "completed":
        response["download_url"] = job["download_url"]
        response["summary"] = job["summary"]
    elif job["status"] == "failed":
        response["error"] = job["error"]

    return response


@app.get("/api/voters/bulk-download/{job_id}")
async def download_bulk_voters(job_id: str):
    """Download the consolidated Excel from a completed bulk voter job."""
    job = bulk_voter_jobs.get(job_id)
    if not job:
        raise HTTPException(status_code=404, detail="Bulk job not found")

    if job["status"] != "completed":
        raise HTTPException(status_code=400, detail="Processing not yet completed")

    output_path = job.get("output_file")
    if not output_path or not Path(output_path).exists():
        raise HTTPException(status_code=404, detail="Output file not found")

    ac = job.get("summary", {}).get("ac_no", "")
    download_name = f"voters_consolidated_AC{ac}.xlsx" if ac else "voters_consolidated.xlsx"

    bg = BackgroundTasks()
    bg.add_task(cleanup_file, Path(output_path))
    bg.add_task(lambda: bulk_voter_jobs.pop(job_id, None))

    return FileResponse(
        path=output_path,
        filename=download_name,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        background=bg,
    )


# ── Excel Merge ──────────────────────────────────────────────────────────


@app.post("/api/excel-merge")
async def merge_excel_files(files: list[UploadFile] = File(...)):
    """Merge multiple voter Excel files into a single flat Excel."""
    from .excel_merger import ExcelMerger

    if len(files) < 1:
        raise HTTPException(status_code=400, detail="At least one file is required")

    # Validate all files are .xlsx
    for f in files:
        if not f.filename or not f.filename.lower().endswith(".xlsx"):
            raise HTTPException(
                status_code=400,
                detail=f"Invalid file type: {f.filename}. Only .xlsx files are accepted.",
            )

    # Save uploaded files to temp directory
    temp_dir = tempfile.mkdtemp(prefix="excel_merge_")
    saved_paths: list[str] = []

    try:
        for f in files:
            safe_name = re.sub(r"[^\w\-. ]", "_", f.filename or "upload.xlsx")
            file_path = os.path.join(temp_dir, safe_name)
            async with aiofiles.open(file_path, "wb") as out:
                content = await f.read()
                await out.write(content)
            saved_paths.append(file_path)

        # Merge
        output_path = os.path.join(OUTPUT_DIR, f"merged_voters_{uuid.uuid4().hex[:8]}.xlsx")
        merger = ExcelMerger()
        result = merger.merge(saved_paths, output_path)

        # Return the merged file
        download_name = "merged_voters.xlsx"

        bg = BackgroundTasks()
        # Clean up temp uploads
        for p in saved_paths:
            bg.add_task(cleanup_file, Path(p))
        bg.add_task(cleanup_file, Path(temp_dir))

        return FileResponse(
            path=output_path,
            filename=download_name,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            background=bg,
            headers={
                "X-Total-Rows": str(result["total_rows"]),
                "X-Total-Files": str(result["total_files"]),
                "X-Total-Sheets": str(result["total_sheets"]),
            },
        )
    except Exception as e:
        # Clean up on error
        for p in saved_paths:
            try:
                os.unlink(p)
            except OSError:
                pass
        try:
            os.rmdir(temp_dir)
        except OSError:
            pass
        logger.exception("Excel merge failed")
        raise HTTPException(status_code=500, detail=f"Merge failed: {str(e)}")


@app.post("/api/booth/add-booth-name-column", response_model=FullPreviewData)
async def add_booth_name_column(request: AddBoothNameColumnRequest):
    """
    Add a "Booth name" column by extracting text before comma from source column.
    
    Extracts the text before the first comma (",") from each cell in the source column
    and adds it as a new "Booth name" column immediately after the source column.
    """
    if request.task_id not in tasks:
        raise HTTPException(status_code=404, detail="Task not found")

    task = tasks[request.task_id]

    if task.status != "completed":
        raise HTTPException(
            status_code=400,
            detail=f"Task not completed. Current status: {task.status}",
        )

    if not task.output_file or not Path(task.output_file).exists():
        raise HTTPException(status_code=404, detail="Output file not found")

    try:
        import openpyxl

        wb = openpyxl.load_workbook(task.output_file, data_only=True)
        ws = wb.active

        # Find actual data range
        actual_max_row = ws.max_row if ws.max_row else 0
        actual_max_col = ws.max_column if ws.max_column else 0

        # Scan to find actual data boundaries
        last_data_row = actual_max_row
        scan_limit = max(actual_max_row + 100, 1000)
        for row in range(1, min(scan_limit + 1, 2000)):
            for col in range(1, min(actual_max_col + 100, 200)):
                cell = ws.cell(row, col)
                if cell.value is not None and str(cell.value).strip():
                    actual_max_row = max(actual_max_row, row)
                    actual_max_col = max(actual_max_col, col)
                    last_data_row = row
            if row > last_data_row + 50:
                break

        if actual_max_row == 0:
            actual_max_row = ws.max_row if ws.max_row else 100
        if actual_max_col == 0:
            actual_max_col = ws.max_column if ws.max_column else 50

        # Find data start row (skip title rows)
        data_start_row = 1
        document_title = None
        for row in range(1, min(20, actual_max_row + 1)):
            cell_value = ws.cell(row, 1).value
            if cell_value and str(cell_value).strip():
                # Check if this looks like a title row
                if row == 1 and "FORM" in str(cell_value).upper():
                    document_title = str(cell_value).strip()
                    continue
                # Check if this row looks like headers
                non_empty_count = sum(
                    1 for col in range(1, min(actual_max_col + 1, 20))
                    if ws.cell(row, col).value and str(ws.cell(row, col).value).strip()
                )
                if non_empty_count >= 2:
                    data_start_row = row
                    break

        # Find where actual data starts
        first_data_row = data_start_row + 1
        for row in range(data_start_row + 1, min(data_start_row + 10, actual_max_row + 1)):
            cell_val = ws.cell(row, 1).value
            if cell_val is not None:
                if isinstance(cell_val, (int, float)) or (isinstance(cell_val, str) and cell_val.strip().isdigit()):
                    first_data_row = row
                    break

        # Calculate how many header rows we have
        num_header_rows = first_data_row - data_start_row

        # Get headers by merging multi-row headers
        headers = []
        for col in range(1, actual_max_col + 1):
            header_parts = []
            for header_row in range(data_start_row, data_start_row + num_header_rows):
                value = ws.cell(header_row, col).value
                if value is not None and str(value).strip():
                    clean_value = str(value).strip()
                    if clean_value.upper() not in ['PARTY ABBREVIATION', 'NO. OF VALID VOTES CAST IN FAVOUR OF']:
                        header_parts.append(clean_value)

            if header_parts:
                if len(header_parts) == 1:
                    headers.append(header_parts[0])
                else:
                    combined = " - ".join(header_parts[-2:]) if len(header_parts) >= 2 else header_parts[-1]
                    headers.append(combined)
            else:
                headers.append(f"Column {col}")

        if not headers:
            for col in range(1, min(actual_max_col + 1, 50)):
                headers.append(f"Column {col}")

        # Find source column index
        source_col_idx = None
        for idx, header in enumerate(headers):
            if header == request.source_column:
                source_col_idx = idx
                break

        if source_col_idx is None:
            wb.close()
            raise HTTPException(
                status_code=400,
                detail=f"Source column '{request.source_column}' not found in headers",
            )

        # Get ALL data rows
        rows = []
        for row_idx in range(first_data_row, actual_max_row + 1):
            row_data = []
            num_cols_to_read = max(len(headers), actual_max_col)
            for col in range(1, num_cols_to_read + 1):
                value = ws.cell(row_idx, col).value
                if isinstance(value, str):
                    value = sanitize_text(value, single_line=False)
                row_data.append(value)
            while len(row_data) < len(headers):
                row_data.append(None)
            if len(row_data) > len(headers):
                row_data = row_data[:len(headers)]
            rows.append(row_data)

        # Extract booth names from source column
        booth_names = []
        for row in rows:
            source_value = row[source_col_idx] if source_col_idx < len(row) else None
            if source_value is not None:
                booth_name = extract_booth_name(str(source_value))
                booth_names.append(booth_name)
            else:
                booth_names.append("")

        # Insert "Booth name" column after source column
        new_headers = headers[:]
        new_headers.insert(source_col_idx + 1, "Booth name")

        # Update rows with new column
        new_rows = []
        for row_idx, row in enumerate(rows):
            new_row = row[:]
            new_row.insert(source_col_idx + 1, booth_names[row_idx])
            new_rows.append(new_row)

        wb.close()

        return FullPreviewData(
            headers=new_headers,
            rows=new_rows,
            total_rows=len(new_rows),
            total_columns=len(new_headers),
            pages_processed=1,
            document_title=document_title,
        )

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Failed to add booth name column: {e}", exc_info=True)
        raise HTTPException(
            status_code=500,
            detail=f"Failed to add booth name column: {str(e)}",
        )


@app.get("/api/status/{task_id}", response_model=StatusResponse)
async def get_status(task_id: str):
    """Get the status of a conversion task."""
    if task_id not in tasks:
        raise HTTPException(status_code=404, detail="Task not found")

    task = tasks[task_id]
    return StatusResponse(
        task_id=task.task_id,
        status=task.status,
        progress=task.progress,
        message=task.message,
        output_file=task.output_file,
        error=task.error,
        validation_issues=task.validation_issues,
    )


@app.get("/api/progress/{task_id}")
async def progress_stream(task_id: str):
    """
    Server-Sent Events endpoint for real-time progress updates.

    Connect to this endpoint to receive live progress updates during conversion.
    """
    if task_id not in tasks:
        raise HTTPException(status_code=404, detail="Task not found")

    async def event_generator():
        last_progress = -1
        while True:
            if task_id in tasks:
                task = tasks[task_id]

                # Only send update if progress changed
                if task.progress != last_progress:
                    last_progress = task.progress
                    event = ProgressEvent(
                        progress=task.progress,
                        status=task.status,
                        message=task.message,
                    )
                    yield {
                        "event": "progress",
                        "data": json.dumps(event.model_dump()),
                    }

                # Stop streaming when complete, failed, or needs_review
                if task.status in ["completed", "failed", "needs_review"]:
                    break

            await asyncio.sleep(0.3)

    return EventSourceResponse(event_generator())


@app.get("/api/preview/{task_id}", response_model=PreviewData)
async def get_preview(task_id: str):
    """
    Get a preview of the extracted data.

    Returns the first 10 rows of data for preview before download.
    """
    if task_id not in tasks:
        raise HTTPException(status_code=404, detail="Task not found")

    task = tasks[task_id]

    if task.status != "completed":
        raise HTTPException(
            status_code=400,
            detail=f"Task not completed. Current status: {task.status}",
        )

    if not task.output_file or not Path(task.output_file).exists():
        raise HTTPException(status_code=404, detail="Output file not found")

    # Read Excel file for preview
    try:
        import openpyxl

        wb = openpyxl.load_workbook(task.output_file, data_only=True)
        ws = wb.active

        # Find actual data range by scanning for cells with values
        # This is more reliable than max_row/max_column which can be inaccurate
        # Start with max_row/max_column as baseline, then expand if needed
        actual_max_row = ws.max_row if ws.max_row else 0
        actual_max_col = ws.max_column if ws.max_column else 0
        
        # Scan iteratively to find actual data boundaries
        # Check rows beyond max_row (sometimes max_row is inaccurate)
        last_data_row = actual_max_row
        scan_limit = max(actual_max_row + 100, 1000)  # Scan up to 1000 rows or max_row + 100
        for row in range(1, min(scan_limit + 1, 2000)):
            row_has_data = False
            for col in range(1, min(actual_max_col + 100, 200)):
                cell = ws.cell(row, col)
                if cell.value is not None and str(cell.value).strip():
                    actual_max_row = max(actual_max_row, row)
                    actual_max_col = max(actual_max_col, col)
                    row_has_data = True
                    last_data_row = row
            # If we've gone 50 rows without finding new data, stop scanning
            if row > last_data_row + 50:
                break

        # Ensure we have reasonable values
        if actual_max_row == 0:
            actual_max_row = ws.max_row if ws.max_row else 100
        if actual_max_col == 0:
            actual_max_col = ws.max_column if ws.max_column else 50

        # Find data start row (skip title rows)
        data_start_row = 1
        for row in range(1, min(20, actual_max_row + 1)):
            cell_value = ws.cell(row, 1).value
            if cell_value and str(cell_value).strip():
                # Check if this row looks like headers (has multiple non-empty cells)
                non_empty_count = sum(
                    1 for col in range(1, min(actual_max_col + 1, 20))
                    if ws.cell(row, col).value and str(ws.cell(row, col).value).strip()
                )
                if non_empty_count >= 2:  # At least 2 columns with data
                    data_start_row = row
                    break

        # Find where actual data starts (first row with numeric data in first column)
        # This helps us know how many header rows there are
        first_data_row = data_start_row + 1
        for row in range(data_start_row + 1, min(data_start_row + 10, actual_max_row + 1)):
            cell_val = ws.cell(row, 1).value
            # Check if this looks like data (numeric or starts with a number)
            if cell_val is not None:
                if isinstance(cell_val, (int, float)) or (isinstance(cell_val, str) and cell_val.strip().isdigit()):
                    first_data_row = row
                    break

        # Calculate how many header rows we have (between data_start_row and first_data_row)
        num_header_rows = first_data_row - data_start_row

        # Get headers by merging multi-row headers
        # Many election PDFs have 2-3 header rows (main header + candidate name + party)
        # But constituency data typically has only 1 header row
        headers = []
        for col in range(1, actual_max_col + 1):
            # Get values from header rows only (not data rows)
            header_parts = []

            # Only check actual header rows (not data rows)
            for header_row in range(data_start_row, data_start_row + num_header_rows):
                value = ws.cell(header_row, col).value
                if value is not None and str(value).strip():
                    clean_value = str(value).strip()
                    # Skip generic phrases and party abbreviation labels
                    if clean_value.upper() not in ['PARTY ABBREVIATION', 'NO. OF VALID VOTES CAST IN FAVOUR OF']:
                        header_parts.append(clean_value)

            # Combine header parts into a single name
            if header_parts:
                # Use the last non-empty part (usually the party name or most specific info)
                # Or combine if there are multiple meaningful parts
                if len(header_parts) == 1:
                    headers.append(header_parts[0])
                else:
                    # For multi-part headers, prioritize party names (usually in last row)
                    # and candidate names
                    combined = " - ".join(header_parts[-2:]) if len(header_parts) >= 2 else header_parts[-1]
                    headers.append(combined)
            else:
                # Fallback for empty columns
                headers.append(f"Column {col}")

        # Ensure we have at least some headers
        if not headers:
            # Fallback: try to detect columns by scanning first few data rows
            for col in range(1, min(actual_max_col + 1, 50)):
                headers.append(f"Column {col}")

        # Get first 10 data rows (first_data_row was already calculated above)
        rows = []
        preview_row_count = min(10, actual_max_row - first_data_row + 1)
        for row_idx in range(first_data_row, first_data_row + preview_row_count):
            if row_idx > actual_max_row:
                break
            row_data = []
            # Read all columns up to the number of headers we found
            num_cols_to_read = max(len(headers), actual_max_col)
            for col in range(1, num_cols_to_read + 1):
                value = ws.cell(row_idx, col).value
                # Fix reversed/corrupted text in cell values
                if isinstance(value, str):
                    value = sanitize_text(value, single_line=False)
                row_data.append(value)
            # Ensure row_data matches header count
            while len(row_data) < len(headers):
                row_data.append(None)
            # Trim if too long (shouldn't happen, but safety check)
            if len(row_data) > len(headers):
                row_data = row_data[:len(headers)]
            rows.append(row_data)

        # Calculate actual total rows (excluding all header rows)
        total_rows = max(0, actual_max_row - first_data_row + 1)

        return PreviewData(
            headers=headers,
            rows=rows,
            total_rows=total_rows,
            total_columns=len(headers),
            pages_processed=1,  # Simplified for now
        )

    except Exception as e:
        raise HTTPException(
            status_code=500,
            detail=f"Failed to read preview: {str(e)}",
        )


@app.get("/api/preview-full/{task_id}", response_model=FullPreviewData)
async def get_full_preview(task_id: str):
    """
    Get full data for spreadsheet editor.
    """
    if task_id not in tasks:
        raise HTTPException(status_code=404, detail="Task not found")

    task = tasks[task_id]

    if task.status != "completed":
        raise HTTPException(
            status_code=400,
            detail=f"Task not completed. Current status: {task.status}",
        )

    if not task.output_file or not Path(task.output_file).exists():
        raise HTTPException(status_code=404, detail="Output file not found")

    # Read Excel file
    try:
        import openpyxl

        wb = openpyxl.load_workbook(task.output_file, data_only=True)
        ws = wb.active

        # Reuse logic to find data boundaries
        actual_max_row = ws.max_row if ws.max_row else 0
        actual_max_col = ws.max_column if ws.max_column else 0

        last_data_row = actual_max_row
        scan_limit = max(actual_max_row + 100, 1000)
        for row in range(1, min(scan_limit + 1, 2000)):
            for col in range(1, min(actual_max_col + 100, 200)):
                cell = ws.cell(row, col)
                if cell.value is not None and str(cell.value).strip():
                    actual_max_row = max(actual_max_row, row)
                    actual_max_col = max(actual_max_col, col)
                    last_data_row = row
            if row > last_data_row + 50:
                break

        if actual_max_row == 0:
            actual_max_row = ws.max_row if ws.max_row else 100
        if actual_max_col == 0:
            actual_max_col = ws.max_column if ws.max_column else 50

        # Find data start row
        data_start_row = 1
        document_title = None
        for row in range(1, min(20, actual_max_row + 1)):
            cell_value = ws.cell(row, 1).value
            if cell_value and str(cell_value).strip():
                if row == 1 and "FORM" in str(cell_value).upper():
                    document_title = str(cell_value).strip()
                    continue
                non_empty_count = sum(
                    1 for col in range(1, min(actual_max_col + 1, 20))
                    if ws.cell(row, col).value and str(ws.cell(row, col).value).strip()
                )
                if non_empty_count >= 2:
                    data_start_row = row
                    break
        
        first_data_row = data_start_row + 1
        for row in range(data_start_row + 1, min(data_start_row + 10, actual_max_row + 1)):
            cell_val = ws.cell(row, 1).value
            if cell_val is not None:
                if isinstance(cell_val, (int, float)) or (isinstance(cell_val, str) and cell_val.strip().isdigit()):
                    first_data_row = row
                    break

        num_header_rows = first_data_row - data_start_row

        headers = []
        for col in range(1, actual_max_col + 1):
            header_parts = []
            for header_row in range(data_start_row, data_start_row + num_header_rows):
                value = ws.cell(header_row, col).value
                if value is not None and str(value).strip():
                    clean_value = str(value).strip()
                    if clean_value.upper() not in ['PARTY ABBREVIATION', 'NO. OF VALID VOTES CAST IN FAVOUR OF']:
                        header_parts.append(clean_value)
            
            if header_parts:
                if len(header_parts) == 1:
                    headers.append(header_parts[0])
                else:
                    combined = " - ".join(header_parts[-2:]) if len(header_parts) >= 2 else header_parts[-1]
                    headers.append(combined)
            else:
                headers.append(f"Column {col}")

        if not headers:
             for col in range(1, min(actual_max_col + 1, 50)):
                headers.append(f"Column {col}")

        rows = []
        for row_idx in range(first_data_row, actual_max_row + 1):
            row_data = []
            num_cols_to_read = max(len(headers), actual_max_col)
            for col in range(1, num_cols_to_read + 1):
                value = ws.cell(row_idx, col).value
                if isinstance(value, str):
                    value = sanitize_text(value, single_line=False)
                row_data.append(value)
            while len(row_data) < len(headers):
                row_data.append(None)
            if len(row_data) > len(headers):
                row_data = row_data[:len(headers)]
            rows.append(row_data)

        # Calculate pages processed (simplified)
        pages_processed = 1
        
        return FullPreviewData(
            headers=headers,
            rows=rows,
            total_rows=len(rows),
            total_columns=len(headers),
            pages_processed=pages_processed,
            document_title=document_title,
        )

    except Exception as e:
        logger.error(f"Failed to get full preview: {e}", exc_info=True)
        raise HTTPException(
            status_code=500,
            detail=f"Failed to get full preview: {str(e)}",
        )


@app.post("/api/normalize-column", response_model=FullPreviewData)
async def normalize_column(request: NormalizeColumnRequest):
    """
    Normalize a column to Party Names using reference data.
    
    Uses fuzzy matching and reverse text detection to map candidate names
    to their respective parties based on dets.json key-value pairs.
    """
    if request.task_id not in tasks:
        raise HTTPException(status_code=404, detail="Task not found")

    task = tasks[request.task_id]

    if task.status != "completed":
        raise HTTPException(
            status_code=400,
            detail=f"Task not completed. Current status: {task.status}",
        )

    if not task.output_file or not Path(task.output_file).exists():
        raise HTTPException(status_code=404, detail="Output file not found")

    try:
        import openpyxl

        # Initialize normalizer
        normalizer = PartyNormalizer()
        
        wb = openpyxl.load_workbook(task.output_file, data_only=True)
        ws = wb.active

        # Find actual data range (reuse logic from add_booth_name_column)
        actual_max_row = ws.max_row if ws.max_row else 0
        actual_max_col = ws.max_column if ws.max_column else 0

        # Scan to find actual data boundaries
        last_data_row = actual_max_row
        scan_limit = max(actual_max_row + 100, 1000)
        for row in range(1, min(scan_limit + 1, 2000)):
            for col in range(1, min(actual_max_col + 100, 200)):
                cell = ws.cell(row, col)
                if cell.value is not None and str(cell.value).strip():
                    actual_max_row = max(actual_max_row, row)
                    actual_max_col = max(actual_max_col, col)
                    last_data_row = row
            if row > last_data_row + 50:
                break
        
        if actual_max_row == 0:
            actual_max_row = ws.max_row if ws.max_row else 100
        if actual_max_col == 0:
            actual_max_col = ws.max_column if ws.max_column else 50

        # Find data start row
        data_start_row = 1
        document_title = None
        for row in range(1, min(20, actual_max_row + 1)):
            cell_value = ws.cell(row, 1).value
            if cell_value and str(cell_value).strip():
                if row == 1 and "FORM" in str(cell_value).upper():
                    document_title = str(cell_value).strip()
                    continue
                non_empty_count = sum(
                    1 for col in range(1, min(actual_max_col + 1, 20))
                    if ws.cell(row, col).value and str(ws.cell(row, col).value).strip()
                )
                if non_empty_count >= 2:
                    data_start_row = row
                    break
        
        first_data_row = data_start_row + 1
        # Find where numeric data starts to confirm data rows
        for row in range(data_start_row + 1, min(data_start_row + 10, actual_max_row + 1)):
            cell_val = ws.cell(row, 1).value
            if cell_val is not None:
                if isinstance(cell_val, (int, float)) or (isinstance(cell_val, str) and cell_val.strip().isdigit()):
                    first_data_row = row
                    break

        num_header_rows = first_data_row - data_start_row

        # Get headers
        headers = []
        for col in range(1, actual_max_col + 1):
            header_parts = []
            for header_row in range(data_start_row, data_start_row + num_header_rows):
                value = ws.cell(header_row, col).value
                if value is not None and str(value).strip():
                    clean_value = str(value).strip()
                    if clean_value.upper() not in ['PARTY ABBREVIATION', 'NO. OF VALID VOTES CAST IN FAVOUR OF']:
                        header_parts.append(clean_value)
            
            if header_parts:
                if len(header_parts) == 1:
                    headers.append(header_parts[0])
                else:
                    combined = " - ".join(header_parts[-2:]) if len(header_parts) >= 2 else header_parts[-1]
                    headers.append(combined)
            else:
                headers.append(f"Column {col}")
        
        if not headers:
             for col in range(1, min(actual_max_col + 1, 50)):
                headers.append(f"Column {col}")

        # Find target column index
        target_col_idx = None
        for idx, header in enumerate(headers):
            if header == request.column_name:
                target_col_idx = idx + 1 # 1-based index for openpyxl
                break
        
        if target_col_idx is None:
            # Fallback check for simplified match
            for idx, header in enumerate(headers):
                if request.column_name in header:
                    target_col_idx = idx + 1
                    break

        if target_col_idx is None:
             raise HTTPException(
                status_code=400,
                detail=f"Column '{request.column_name}' not found",
            )

        # Find Constituency column index for context (optional)
        constituency_col_idx = None
        for idx, header in enumerate(headers):
            if "CONSTITUENCY" in header.upper():
                constituency_col_idx = idx + 1
                break

        # Normalize values in the column
        normalized_count = 0
        wb_save = openpyxl.load_workbook(task.output_file)
        ws_save = wb_save.active
        
        for row_idx in range(first_data_row, actual_max_row + 1):
            # Read from data_only workbook
            cell = ws.cell(row_idx, target_col_idx)
            original_value = cell.value
            
            if original_value and isinstance(original_value, str):
                # Construct context for collision handling
                context_parts = []
                if document_title:
                    context_parts.append(document_title)
                
                if constituency_col_idx:
                    const_cell = ws.cell(row_idx, constituency_col_idx)
                    const_val = const_cell.value
                    if const_val and str(const_val).strip():
                        context_parts.append(str(const_val).strip())
                
                context = " ".join(context_parts)
                normalized_value = normalizer.normalize_value(original_value, context)
                if normalized_value != original_value:
                    # Write to save workbook
                    ws_save.cell(row_idx, target_col_idx).value = normalized_value
                    normalized_count += 1
        
        logger.info(f"Normalized {normalized_count} values in column {request.column_name}")
        
        # Save updated workbook
        wb_save.save(task.output_file)
        wb_save.close()
        wb.close()

        # Reload for response via get_full_preview logic
        return await get_full_preview(task.task_id)

    except Exception as e:
        logger.error(f"Column normalization error: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/api/normalize-headers", response_model=FullPreviewData)
async def normalize_headers(request: NormalizeColumnRequest):
    """
    Normalize all headers in the first data row (column names) to Party Names.
    Uses 'Constellation Strategy': detected constituency from headers to resolve ambiguity.
    """
    if request.task_id not in tasks:
        raise HTTPException(status_code=404, detail="Task not found")

    task = tasks[request.task_id]

    if task.status != "completed":
        raise HTTPException(status_code=400, detail="Task processing not completed")

    if not os.path.exists(task.output_file):
        raise HTTPException(status_code=404, detail="Output file not found")

    try:
        # Load workbook
        wb = openpyxl.load_workbook(task.output_file)
        ws = wb.active

        # Identify header row
        header_row = None
        
        # Scan first 20 rows
        for row in range(1, 21):
            non_empty_count = sum(1 for col in range(1, 20) if ws.cell(row, col).value)
            if non_empty_count >= 2:
                header_row = row
                break
        
        if not header_row:
             wb.close()
             raise HTTPException(status_code=400, detail="Could not detect header row")

        # Get all headers
        max_col = ws.max_column
        headers = []
        header_cells = []
        for col in range(1, max_col + 1):
            cell = ws.cell(header_row, col)
            val = str(cell.value) if cell.value else ""
            headers.append(val)
            header_cells.append(cell)

        # 1. Detect Constituency from all headers
        normalizer = PartyNormalizer()
        
        # We need to collect valid candidate names to detect constituency
        # normalizer.detect_constituency takes a list of strings
        detected_constituency = normalizer.detect_constituency(headers)
        
        if detected_constituency:
            logger.info(f"Detected constituency context '{detected_constituency}' for task {task.task_id}")
        else:
            logger.warning(f"No constituency context detected for task {task.task_id}")

        # 2. Normalize each header
        normalized_count = 0
        skipped_count = 0
        no_match_count = 0

        for idx, header in enumerate(headers):
            if not header:
                continue

            # Use constituency-filtered normalization if constituency detected
            if detected_constituency:
                # Use strict constituency-based matching
                normalized_value = normalizer.normalize_value(
                    header,
                    constituency=detected_constituency,
                    use_constituency_filter=True  # Enable conservative matching
                )
            else:
                # Fallback to original logic if no constituency detected
                logger.warning(f"No constituency detected, using legacy normalization for '{header}'")
                normalized_value = normalizer.normalize_value(header, use_constituency_filter=False)

            if normalized_value != header and normalized_value != "None":
                # Valid normalization occurred
                # Update cell value
                header_cells[idx].value = normalized_value
                normalized_count += 1
                logger.info(f"✓ Normalized: '{header}' -> '{normalized_value}'")
            elif normalized_value == header:
                # No match found or blacklisted
                no_match_count += 1
                logger.debug(f"○ Kept original: '{header}' (no match or blacklisted)")
            else:
                skipped_count += 1

        logger.info(f"Normalization complete: {normalized_count} normalized, "
                   f"{no_match_count} kept original, {skipped_count} skipped")

        wb.save(task.output_file)
        wb.close()

        # Return updated preview
        return await get_full_preview(task.task_id)

    except Exception as e:
        logger.error(f"Header normalization error: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/api/columns/{task_id}")
async def get_columns(task_id: str):
    """
    Get column names directly from the converted Excel file.
    
    Returns the exact column names as they appear in the Excel file,
    without any processing or merging.
    """
    if task_id not in tasks:
        raise HTTPException(status_code=404, detail="Task not found")

    task = tasks[task_id]

    if task.status != "completed":
        raise HTTPException(
            status_code=400,
            detail=f"Task not completed. Current status: {task.status}",
        )

    if not task.output_file or not Path(task.output_file).exists():
        raise HTTPException(status_code=404, detail="Output file not found")

    # Read Excel file to get column names
    try:
        import openpyxl

        wb = openpyxl.load_workbook(task.output_file, data_only=True)
        ws = wb.active

        # Find actual data range by scanning for cells with values
        actual_max_row = ws.max_row if ws.max_row else 0
        actual_max_col = ws.max_column if ws.max_column else 0
        
        # Scan iteratively to find actual data boundaries
        last_data_row = actual_max_row
        scan_limit = max(actual_max_row + 100, 1000)
        for row in range(1, min(scan_limit + 1, 2000)):
            row_has_data = False
            for col in range(1, min(actual_max_col + 100, 200)):
                cell = ws.cell(row, col)
                if cell.value is not None and str(cell.value).strip():
                    actual_max_row = max(actual_max_row, row)
                    actual_max_col = max(actual_max_col, col)
                    row_has_data = True
                    last_data_row = row
            if row > last_data_row + 50:
                break

        # Ensure we have reasonable values
        if actual_max_row == 0:
            actual_max_row = ws.max_row if ws.max_row else 100
        if actual_max_col == 0:
            actual_max_col = ws.max_column if ws.max_column else 50

        # Find data start row (skip title rows)
        data_start_row = 1
        for row in range(1, min(20, actual_max_row + 1)):
            cell_value = ws.cell(row, 1).value
            if cell_value and str(cell_value).strip():
                # Check if this row looks like headers (has multiple non-empty cells)
                non_empty_count = sum(
                    1 for col in range(1, min(actual_max_col + 1, 20))
                    if ws.cell(row, col).value and str(ws.cell(row, col).value).strip()
                )
                if non_empty_count >= 2:  # At least 2 columns with data
                    data_start_row = row
                    break

        # Find where actual data starts (first row with numeric data in first column)
        first_data_row = data_start_row + 1
        for row in range(data_start_row + 1, min(data_start_row + 10, actual_max_row + 1)):
            cell_val = ws.cell(row, 1).value
            if cell_val is not None:
                if isinstance(cell_val, (int, float)) or (isinstance(cell_val, str) and cell_val.strip().isdigit()):
                    first_data_row = row
                    break

        # Calculate how many header rows we have
        num_header_rows = first_data_row - data_start_row

        # Get headers by reading from the header row(s)
        # Many Excel files have headers in a single row, but some have multi-row headers
        columns = []
        for col in range(1, actual_max_col + 1):
            # Get values from actual header rows only (not data rows)
            header_parts = []

            for header_row in range(data_start_row, data_start_row + num_header_rows):
                value = ws.cell(header_row, col).value
                if value is not None and str(value).strip():
                    clean_value = str(value).strip()
                    # Skip generic phrases
                    if clean_value.upper() not in ['PARTY ABBREVIATION', 'NO. OF VALID VOTES CAST IN FAVOUR OF']:
                        header_parts.append(clean_value)

            # Combine header parts into a single name
            if header_parts:
                # Use the last non-empty part (usually the party name or most specific info)
                # Or combine if there are multiple meaningful parts
                if len(header_parts) == 1:
                    columns.append(header_parts[0])
                else:
                    # For multi-part headers, combine them with " - "
                    combined = " - ".join(header_parts[-2:]) if len(header_parts) >= 2 else header_parts[-1]
                    columns.append(combined)
            else:
                # Fallback for empty columns
                columns.append(f"Column {col}")

        # Ensure we have at least some headers
        if not columns:
            # Fallback: try to detect columns by scanning first few data rows
            for col in range(1, min(actual_max_col + 1, 50)):
                columns.append(f"Column {col}")

        wb.close()

        return {"columns": columns}

    except Exception as e:
        logger.error(f"Failed to read columns from Excel: {str(e)}")
        raise HTTPException(
            status_code=500,
            detail=f"Failed to read columns from Excel file: {str(e)}",
        )


@app.get("/api/download/{task_id}")
async def download_excel(task_id: str):
    """
    Download the generated Excel file.

    Returns the Excel file as a downloadable attachment.
    """
    if task_id not in tasks:
        raise HTTPException(status_code=404, detail="Task not found")

    task = tasks[task_id]

    if task.status != "completed":
        raise HTTPException(
            status_code=400,
            detail=f"Task not completed. Current status: {task.status}",
        )

    if not task.output_file or not Path(task.output_file).exists():
        raise HTTPException(status_code=404, detail="Output file not found")

    # Generate download filename
    original_name = Path(task.filename).stem
    clean_name = clean_excel_filename(original_name)
    download_name = f"{clean_name}.xlsx"

    return FileResponse(
        task.output_file,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename=download_name,
    )


@app.delete("/api/task/{task_id}")
async def delete_task(task_id: str):
    """Delete a task and its associated files."""
    if task_id not in tasks:
        raise HTTPException(status_code=404, detail="Task not found")

    task = tasks[task_id]

    # Cleanup files
    if task.output_file and Path(task.output_file).exists():
        await cleanup_file(Path(task.output_file))

    upload_path = UPLOAD_DIR / f"{task_id}.pdf"
    if upload_path.exists():
        await cleanup_file(upload_path)

    # Remove from storage
    del tasks[task_id]

    return {"message": "Task deleted successfully"}


@app.post("/api/filter-columns", response_model=FilterColumnsResponse)
async def filter_columns(request: FilterColumnsRequest):
    """
    Filter Excel file to include only selected columns.

    Takes a task_id and list of column names, creates a new Excel file
    with only the requested columns in the specified order.

    Args:
        request: FilterColumnsRequest with task_id and columns list

    Returns:
        FilterColumnsResponse with filtered file path and metadata
    """
    # Validate task exists and is completed
    if request.task_id not in tasks:
        raise HTTPException(status_code=404, detail="Task not found")

    task = tasks[request.task_id]

    if task.status != "completed":
        raise HTTPException(
            status_code=400,
            detail=f"Task not completed. Current status: {task.status}",
        )

    if not task.output_file or not Path(task.output_file).exists():
        raise HTTPException(status_code=404, detail="Output file not found")

    # Validate columns list
    if not request.columns or len(request.columns) == 0:
        raise HTTPException(
            status_code=400,
            detail="No columns specified. Please provide at least one column name.",
        )

    try:
        logger.info(f"Filtering columns for task {request.task_id}: {request.columns}")

        # Create filter service
        filter_service = ColumnFilterService()

        # Filter columns
        filtered_file, metadata = await asyncio.to_thread(
            filter_service.filter_columns,
            task.output_file,
            request.columns,
            str(OUTPUT_DIR),
        )

        logger.info(f"Column filtering completed: {filtered_file}")

        # Return response
        return FilterColumnsResponse(
            filtered_file_path=filtered_file,
            original_file=metadata["original_file"],
            selected_columns=metadata["selected_columns"],
            total_columns=metadata["total_columns"],
            total_rows=metadata["total_rows"],
            columns_removed=metadata["columns_removed"],
            timestamp=metadata["timestamp"],
        )

    except ValueError as e:
        # Column validation errors
        raise HTTPException(status_code=400, detail=str(e))
    except Exception as e:
        logger.error(f"Column filtering failed: {e}", exc_info=True)
        raise HTTPException(
            status_code=500,
            detail=f"Column filtering failed: {str(e)}",
        )


@app.post("/api/filter-excel")
async def filter_excel(request: FilterExcelRequest):
    """
    Filter Excel file to include only selected columns.

    Takes a task_id and list of column names.
    Creates a new Excel file with selected columns.

    Args:
        request: FilterExcelRequest with task_id and selected_columns

    Returns:
        FileResponse with the filtered Excel file for direct download
    """
    # Validate task exists and is completed
    if request.task_id not in tasks:
        raise HTTPException(status_code=404, detail="Task not found")

    task = tasks[request.task_id]

    if task.status != "completed":
        raise HTTPException(
            status_code=400,
            detail=f"Task not completed. Current status: {task.status}",
        )

    if not task.output_file or not Path(task.output_file).exists():
        raise HTTPException(status_code=404, detail="Output file not found")

    # Validate columns list
    if not request.selected_columns or len(request.selected_columns) == 0:
        raise HTTPException(
            status_code=400,
            detail="No columns specified. Please provide at least one column name.",
        )

    try:
        logger.info(
            f"Filtering columns for task {request.task_id}: {request.selected_columns}"
        )

        # Create filter service
        filter_service = ColumnFilterService()

        # Filter columns
        filtered_file, metadata = await asyncio.to_thread(
            filter_service.filter_columns,
            task.output_file,
            request.selected_columns,
            str(OUTPUT_DIR),
            request.header_overrides,
            request.task_id,  # Pass task_id to handle alliance mappings
            request.sum_other_columns,
        )

        logger.info(f"Column filtering completed: {filtered_file}")

        # Update task's output file to point to filtered file so editor shows filtered columns
        task.output_file = filtered_file
        logger.info(f"Updated task {request.task_id} output_file to: {filtered_file}")

        # Return the filtered file for download
        return FileResponse(
            filtered_file,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            filename=Path(filtered_file).name,
            headers={
                "Content-Disposition": f'attachment; filename="{Path(filtered_file).name}"'
            },
        )

    except ValueError as e:
        # Column validation errors
        raise HTTPException(status_code=400, detail=str(e))
    except Exception as e:
        logger.error(f"Column filtering failed: {e}", exc_info=True)
        raise HTTPException(
            status_code=500,
            detail=f"Column filtering failed: {str(e)}",
        )


@app.get("/api/download-filtered/{timestamp}")
async def download_filtered_excel(timestamp: str):
    """
    Download a filtered Excel file by timestamp.

    Args:
        timestamp: Timestamp string from the filter operation (format: YYYYMMDD_HHMMSS)

    Returns:
        FileResponse with the filtered Excel file
    """
    # Find the filtered file
    filtered_filename = f"filtered_{timestamp}.xlsx"
    filtered_path = OUTPUT_DIR / filtered_filename

    if not filtered_path.exists():
        raise HTTPException(status_code=404, detail="Filtered file not found")

    return FileResponse(
        filtered_path,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename=filtered_filename,
    )


@app.get("/api/preview-full/{task_id}", response_model=FullPreviewData)
async def get_full_preview(task_id: str):
    """
    Get full preview data with all rows for spreadsheet editor.

    Returns all rows of data (not just first 10) for display in the
    interactive spreadsheet editor.
    """
    if task_id not in tasks:
        raise HTTPException(status_code=404, detail="Task not found")

    task = tasks[task_id]

    if task.status != "completed":
        raise HTTPException(
            status_code=400,
            detail=f"Task not completed. Current status: {task.status}",
        )

    if not task.output_file or not Path(task.output_file).exists():
        raise HTTPException(status_code=404, detail="Output file not found")

    try:
        import openpyxl

        wb = openpyxl.load_workbook(task.output_file, data_only=True)
        ws = wb.active

        # Find actual data range
        actual_max_row = ws.max_row if ws.max_row else 0
        actual_max_col = ws.max_column if ws.max_column else 0

        # Scan to find actual data boundaries
        last_data_row = actual_max_row
        scan_limit = max(actual_max_row + 100, 1000)
        for row in range(1, min(scan_limit + 1, 2000)):
            for col in range(1, min(actual_max_col + 100, 200)):
                cell = ws.cell(row, col)
                if cell.value is not None and str(cell.value).strip():
                    actual_max_row = max(actual_max_row, row)
                    actual_max_col = max(actual_max_col, col)
                    last_data_row = row
            if row > last_data_row + 50:
                break

        if actual_max_row == 0:
            actual_max_row = ws.max_row if ws.max_row else 100
        if actual_max_col == 0:
            actual_max_col = ws.max_column if ws.max_column else 50

        # Find data start row (skip title rows)
        data_start_row = 1
        document_title = None
        for row in range(1, min(20, actual_max_row + 1)):
            cell_value = ws.cell(row, 1).value
            if cell_value and str(cell_value).strip():
                # Check if this looks like a title row
                if row == 1 and "FORM" in str(cell_value).upper():
                    document_title = str(cell_value).strip()
                    continue
                # Check if this row looks like headers
                non_empty_count = sum(
                    1 for col in range(1, min(actual_max_col + 1, 20))
                    if ws.cell(row, col).value and str(ws.cell(row, col).value).strip()
                )
                if non_empty_count >= 2:
                    data_start_row = row
                    break

        # Find where actual data starts (first row with numeric data in first column)
        first_data_row = data_start_row + 1
        for row in range(data_start_row + 1, min(data_start_row + 10, actual_max_row + 1)):
            cell_val = ws.cell(row, 1).value
            if cell_val is not None:
                if isinstance(cell_val, (int, float)) or (isinstance(cell_val, str) and cell_val.strip().isdigit()):
                    first_data_row = row
                    break

        # Calculate how many header rows we have
        num_header_rows = first_data_row - data_start_row

        # Get headers by merging multi-row headers
        headers = []
        for col in range(1, actual_max_col + 1):
            # Get values from actual header rows only (not data rows)
            header_parts = []

            # Only check actual header rows
            for header_row in range(data_start_row, data_start_row + num_header_rows):
                value = ws.cell(header_row, col).value
                if value is not None and str(value).strip():
                    clean_value = str(value).strip()
                    # Skip generic phrases and party abbreviation labels
                    if clean_value.upper() not in ['PARTY ABBREVIATION', 'NO. OF VALID VOTES CAST IN FAVOUR OF']:
                        header_parts.append(clean_value)

            # Combine header parts into a single name
            if header_parts:
                # Use the last non-empty part (usually the party name or most specific info)
                if len(header_parts) == 1:
                    headers.append(header_parts[0])
                else:
                    # For multi-part headers, prioritize party names (usually in last row)
                    combined = " - ".join(header_parts[-2:]) if len(header_parts) >= 2 else header_parts[-1]
                    headers.append(combined)
            else:
                # Fallback for empty columns
                headers.append(f"Column {col}")

        if not headers:
            for col in range(1, min(actual_max_col + 1, 50)):
                headers.append(f"Column {col}")

        # Get ALL data rows (first_data_row was already calculated above)
        rows = []
        for row_idx in range(first_data_row, actual_max_row + 1):
            row_data = []
            num_cols_to_read = max(len(headers), actual_max_col)
            for col in range(1, num_cols_to_read + 1):
                value = ws.cell(row_idx, col).value
                row_data.append(value)
            while len(row_data) < len(headers):
                row_data.append(None)
            if len(row_data) > len(headers):
                row_data = row_data[:len(headers)]
            rows.append(row_data)

        total_rows = len(rows)
        wb.close()

        return FullPreviewData(
            headers=headers,
            rows=rows,
            total_rows=total_rows,
            total_columns=len(headers),
            pages_processed=1,
            document_title=document_title,
        )

    except Exception as e:
        logger.error(f"Failed to read full preview: {e}", exc_info=True)
        raise HTTPException(
            status_code=500,
            detail=f"Failed to read preview: {str(e)}",
        )


@app.post("/api/download-modified")
async def download_modified_excel(request: DownloadModifiedRequest):
    """
    Generate and download Excel file from modified spreadsheet data.

    Accepts the edited data from the spreadsheet editor and creates
    a new Excel file with the modifications.
    """
    if request.task_id not in tasks:
        raise HTTPException(status_code=404, detail="Task not found")

    task = tasks[request.task_id]

    try:
        import openpyxl
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
        from openpyxl.utils import get_column_letter

        # Create new workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Modified Data"

        # Define styles
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=10)
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        # Add document title if provided
        start_row = 1
        if request.document_title:
            ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(request.headers))
            title_cell = ws.cell(row=1, column=1)
            title_cell.value = request.document_title
            title_cell.font = Font(bold=True, size=14)
            title_cell.alignment = Alignment(horizontal="center")
            start_row = 3

        # Write headers
        for col_idx, header in enumerate(request.headers, 1):
            cell = ws.cell(row=start_row, column=col_idx)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
            cell.border = thin_border

        # Write data rows
        for row_idx, row_data in enumerate(request.rows, start_row + 1):
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.value = value
                cell.border = thin_border
                cell.alignment = Alignment(horizontal="center", vertical="center")

        # Auto-adjust column widths
        for col_idx in range(1, len(request.headers) + 1):
            column_letter = get_column_letter(col_idx)
            max_length = 0
            for row in range(start_row, start_row + len(request.rows) + 1):
                cell_value = ws.cell(row=row, column=col_idx).value
                if cell_value:
                    max_length = max(max_length, len(str(cell_value)))
            adjusted_width = min(max_length + 2, 30)
            ws.column_dimensions[column_letter].width = max(adjusted_width, 10)

        # Freeze header row
        ws.freeze_panes = ws.cell(row=start_row + 1, column=1)

        # Generate output filename
        output_filename = f"modified_{task.task_id}_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.xlsx"
        output_path = OUTPUT_DIR / output_filename

        wb.save(str(output_path))
        wb.close()

        logger.info(f"Modified Excel saved: {output_path}")

        # Return file for download
        original_name = Path(task.filename).stem if task.filename else "data"
        clean_name = clean_excel_filename(original_name)
        download_name = f"{clean_name}.xlsx"

        return FileResponse(
            str(output_path),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            filename=download_name,
        )

    except Exception as e:
        logger.error(f"Failed to create modified Excel: {e}", exc_info=True)
        raise HTTPException(
            status_code=500,
            detail=f"Failed to create modified Excel: {str(e)}",
        )


# ============================================================================
# Geocoding Endpoints
# ============================================================================

async def process_geocoding(geocode_task_id: str, addresses: list, region_hint: str):
    """Background task for geocoding addresses."""
    try:
        if geocode_task_id not in geocode_tasks:
            logger.error(f"Geocode task {geocode_task_id} not found in tasks dictionary")
            return
        
        geocode_tasks[geocode_task_id]["status"] = "geocoding"
        geocode_tasks[geocode_task_id]["message"] = "Starting geocoding..."
        logger.info(f"Starting geocoding task {geocode_task_id} with {len(addresses)} addresses")

        service = GeocodingService()

        def progress_callback(current, total, message, success_count, failed_count):
            if geocode_task_id in geocode_tasks:
                geocode_tasks[geocode_task_id].update({
                    "current": current,
                    "total": total,
                    "message": message,
                    "success_count": success_count,
                    "failed_count": failed_count,
                })

        def cancel_check():
            return geocode_tasks.get(geocode_task_id, {}).get("cancelled", False)

        results = await service.geocode_batch(
            addresses,
            region_hint=region_hint if region_hint else "India",
            progress_callback=progress_callback,
            cancel_check=cancel_check,
        )

        # Store results
        if geocode_task_id in geocode_tasks:
            geocode_tasks[geocode_task_id]["results"] = [
                {
                    "row_index": r.row_index,
                    "latitude": r.latitude,
                    "longitude": r.longitude,
                    "status": r.status,
                    "error": r.error,
                }
                for r in results
            ]
            geocode_tasks[geocode_task_id]["status"] = "completed"
            geocode_tasks[geocode_task_id]["message"] = "Geocoding completed"
            geocode_tasks[geocode_task_id]["current"] = len(results)

            success_count = sum(1 for r in results if r.status == "success")
            failed_count = len(results) - success_count
            logger.info(
                f"Geocoding completed for task {geocode_task_id}: "
                f"{success_count} successful, {failed_count} failed out of {len(results)} total"
            )
        else:
            logger.error(f"Geocode task {geocode_task_id} was removed during processing")

    except Exception as e:
        logger.error(f"Geocoding failed for task {geocode_task_id}: {e}", exc_info=True)
        if geocode_task_id in geocode_tasks:
            geocode_tasks[geocode_task_id]["status"] = "failed"
            geocode_tasks[geocode_task_id]["error"] = str(e)
            geocode_tasks[geocode_task_id]["message"] = f"Geocoding failed: {str(e)}"


@app.post("/api/geocode/start", response_model=GeocodeStartResponse)
async def start_geocoding(
    request: GeocodeRequest,
    background_tasks: BackgroundTasks,
):
    """
    Start geocoding addresses from a converted Excel file.

    Reads addresses from the specified column and starts geocoding in the background.
    Use /api/geocode/progress/{geocode_task_id} to track progress via SSE.
    """
    # Validate task exists and is completed
    if request.task_id not in tasks:
        raise HTTPException(status_code=404, detail="Task not found")

    task = tasks[request.task_id]

    if task.status != "completed":
        raise HTTPException(
            status_code=400,
            detail=f"Task not completed. Current status: {task.status}",
        )

    if not task.output_file or not Path(task.output_file).exists():
        raise HTTPException(status_code=404, detail="Output file not found")

    try:
        import openpyxl

        # Read Excel file to get headers and rows
        wb = openpyxl.load_workbook(task.output_file, data_only=True)
        ws = wb.active

        # Find data range (similar logic to get_full_preview)
        actual_max_row = ws.max_row if ws.max_row else 0
        actual_max_col = ws.max_column if ws.max_column else 0

        # Scan to find actual data boundaries
        last_data_row = actual_max_row
        for row in range(1, min(actual_max_row + 100, 2000)):
            for col in range(1, min(actual_max_col + 100, 200)):
                cell = ws.cell(row, col)
                if cell.value is not None and str(cell.value).strip():
                    actual_max_row = max(actual_max_row, row)
                    actual_max_col = max(actual_max_col, col)
                    last_data_row = row
            if row > last_data_row + 50:
                break

        # Find data start row
        data_start_row = 1
        for row in range(1, min(20, actual_max_row + 1)):
            cell_value = ws.cell(row, 1).value
            if cell_value and str(cell_value).strip():
                non_empty_count = sum(
                    1 for col in range(1, min(actual_max_col + 1, 20))
                    if ws.cell(row, col).value and str(ws.cell(row, col).value).strip()
                )
                if non_empty_count >= 2:
                    data_start_row = row
                    break

        # Find first data row
        first_data_row = data_start_row + 1
        for row in range(data_start_row + 1, min(data_start_row + 10, actual_max_row + 1)):
            cell_val = ws.cell(row, 1).value
            if cell_val is not None:
                if isinstance(cell_val, (int, float)) or (isinstance(cell_val, str) and cell_val.strip().isdigit()):
                    first_data_row = row
                    break

        num_header_rows = first_data_row - data_start_row

        # Get headers
        headers = []
        for col in range(1, actual_max_col + 1):
            header_parts = []
            for header_row in range(data_start_row, data_start_row + num_header_rows):
                value = ws.cell(header_row, col).value
                if value is not None and str(value).strip():
                    header_parts.append(str(value).strip())
            if header_parts:
                headers.append(header_parts[-1] if len(header_parts) == 1 else " - ".join(header_parts[-2:]))
            else:
                headers.append(f"Column {col}")

        # Get rows
        rows = []
        for row_idx in range(first_data_row, actual_max_row + 1):
            row_data = []
            for col in range(1, actual_max_col + 1):
                value = ws.cell(row_idx, col).value
                row_data.append(value)
            rows.append(row_data)

        wb.close()

        # Extract addresses from specified column
        try:
            addresses = extract_addresses_from_column(headers, rows, request.address_column)
            logger.info(f"Extracted {len(addresses)} addresses from column '{request.address_column}'")
        except ValueError as e:
            logger.error(f"Failed to extract addresses: {e}")
            raise HTTPException(status_code=400, detail=str(e))
        
        if not addresses:
            raise HTTPException(
                status_code=400,
                detail=f"No addresses found in column '{request.address_column}'. "
                       f"Please verify the column name and ensure it contains address data.",
            )

        # Generate geocode task ID
        geocode_task_id = str(uuid.uuid4())

        # Initialize geocode task
        geocode_tasks[geocode_task_id] = {
            "status": "pending",
            "current": 0,
            "total": len(addresses),
            "message": "Starting geocoding...",
            "success_count": 0,
            "failed_count": 0,
            "results": [],
            "cancelled": False,
            "created_at": datetime.utcnow(),
        }

        # Start background geocoding
        background_tasks.add_task(
            process_geocoding,
            geocode_task_id,
            addresses,
            request.region_hint,
        )

        logger.info(f"Started geocoding task {geocode_task_id} with {len(addresses)} addresses")

        return GeocodeStartResponse(
            geocode_task_id=geocode_task_id,
            total_addresses=len(addresses),
            message=f"Geocoding started for {len(addresses)} addresses",
        )

    except ValueError as e:
        logger.error(f"Value error in geocoding start: {e}")
        raise HTTPException(status_code=400, detail=str(e))
    except HTTPException:
        # Re-raise HTTP exceptions as-is
        raise
    except Exception as e:
        logger.error(f"Failed to start geocoding: {e}", exc_info=True)
        error_detail = str(e)
        # Don't expose internal error details in production
        if "geocoding" not in error_detail.lower():
            error_detail = "Failed to start geocoding. Please try again or contact support if the issue persists."
        raise HTTPException(
            status_code=500,
            detail=error_detail,
        )


@app.get("/api/geocode/progress/{geocode_task_id}")
async def geocode_progress_stream(geocode_task_id: str):
    """
    Server-Sent Events endpoint for real-time geocoding progress updates.

    Connect to this endpoint to receive live progress updates during geocoding.
    """
    if geocode_task_id not in geocode_tasks:
        raise HTTPException(status_code=404, detail="Geocoding task not found")

    async def event_generator():
        last_current = -1
        while True:
            if geocode_task_id in geocode_tasks:
                task = geocode_tasks[geocode_task_id]

                # Only send update if progress changed
                if task["current"] != last_current or task["status"] in ["completed", "failed", "cancelled"]:
                    last_current = task["current"]
                    event = GeocodeProgressEvent(
                        current=task["current"],
                        total=task["total"],
                        status=task["status"],
                        message=task["message"],
                        success_count=task["success_count"],
                        failed_count=task["failed_count"],
                    )
                    yield {
                        "event": "progress",
                        "data": json.dumps(event.model_dump()),
                    }

                # Stop streaming when complete or failed
                if task["status"] in ["completed", "failed", "cancelled"]:
                    break

            await asyncio.sleep(0.3)

    return EventSourceResponse(event_generator())


@app.post("/api/geocode/cancel/{geocode_task_id}")
async def cancel_geocoding(geocode_task_id: str):
    """Cancel an ongoing geocoding operation."""
    if geocode_task_id not in geocode_tasks:
        raise HTTPException(status_code=404, detail="Geocoding task not found")

    geocode_tasks[geocode_task_id]["cancelled"] = True
    geocode_tasks[geocode_task_id]["status"] = "cancelled"
    geocode_tasks[geocode_task_id]["message"] = "Geocoding cancelled by user"

    return {"message": "Geocoding cancelled"}


@app.post("/api/geocode/apply", response_model=GeocodeApplyResponse)
async def apply_geocoding(request: GeocodeApplyRequest):
    """
    Apply geocoding results to spreadsheet data.

    Adds Latitude and Longitude columns to the provided headers and rows
    using the results from a completed geocoding operation.
    """
    if request.geocode_task_id not in geocode_tasks:
        raise HTTPException(status_code=404, detail="Geocoding task not found")

    geocode_task = geocode_tasks[request.geocode_task_id]

    if geocode_task["status"] not in ["completed", "cancelled"]:
        raise HTTPException(
            status_code=400,
            detail=f"Geocoding not completed. Current status: {geocode_task['status']}",
        )

    try:
        results = geocode_task.get("results", [])
        
        if not results:
            logger.warning(f"No geocoding results found for task {request.geocode_task_id}")
            raise HTTPException(
                status_code=400,
                detail="No geocoding results available. The geocoding task may have been cancelled or failed.",
            )

        # Create a mapping of row_index to coordinates
        coords_map = {}
        for r in results:
            row_idx = r.get("row_index")
            if row_idx is not None:
                coords_map[row_idx] = (
                    r.get("latitude"),
                    r.get("longitude"),
                    r.get("status", "error")
                )

        # Check if we have Latitude/Longitude columns already
        has_lat = "Latitude" in request.headers
        has_lng = "Longitude" in request.headers
        
        # Add new headers if they don't exist
        new_headers = list(request.headers)
        if not has_lat:
            new_headers.append("Latitude")
        if not has_lng:
            new_headers.append("Longitude")

        # Add coordinates to each row
        new_rows = []
        successful = 0
        failed = 0
        lat_col_idx = len(request.headers) if not has_lat else request.headers.index("Latitude")
        lng_col_idx = len(request.headers) + (0 if has_lat else 1) if not has_lng else request.headers.index("Longitude")

        for row_idx, row in enumerate(request.rows):
            coords = coords_map.get(row_idx)
            new_row = list(row)
            
            if coords:
                lat, lng, status = coords
                # Update or add coordinates
                if has_lat:
                    new_row[lat_col_idx] = lat
                else:
                    new_row.append(lat)
                
                if has_lng:
                    new_row[lng_col_idx] = lng
                else:
                    new_row.append(lng)
                
                if status == "success":
                    successful += 1
                else:
                    failed += 1
            else:
                # No geocoding result for this row
                if not has_lat:
                    new_row.append(None)
                else:
                    new_row[lat_col_idx] = None
                
                if not has_lng:
                    new_row.append(None)
                else:
                    new_row[lng_col_idx] = None
                
                failed += 1
            
            new_rows.append(new_row)

        logger.info(
            f"Applied geocoding results to {len(new_rows)} rows: "
            f"{successful} successful, {failed} failed"
        )

        return GeocodeApplyResponse(
            headers=new_headers,
            rows=new_rows,
            total_geocoded=len(results),
            successful=successful,
            failed=failed,
        )

    except Exception as e:
        logger.error(f"Failed to apply geocoding: {e}", exc_info=True)
        raise HTTPException(
            status_code=500,
            detail=f"Failed to apply geocoding: {str(e)}",
        )


@app.get("/api/geocode/status/{geocode_task_id}")
async def get_geocode_status(geocode_task_id: str):
    """Get the current status of a geocoding task."""
    if geocode_task_id not in geocode_tasks:
        raise HTTPException(status_code=404, detail="Geocoding task not found")

    task = geocode_tasks[geocode_task_id]
    return {
        "status": task["status"],
        "current": task["current"],
        "total": task["total"],
        "message": task["message"],
        "success_count": task["success_count"],
        "failed_count": task["failed_count"],
    }


# ============================================================================
# Translation Endpoints
# ============================================================================

async def process_translation(
    translate_task_id: str,
    task_id: str,
    input_file: str,
    output_file: str,
    target_lang: str,
):
    """Background task for translating Excel content."""
    try:
        from .excel_translator import ExcelTranslator

        translation_tasks[translate_task_id]["status"] = "translating"
        translation_tasks[translate_task_id]["message"] = "Starting translation..."

        translator = ExcelTranslator()

        def progress_callback(current: int, total: int, message: str):
            translation_tasks[translate_task_id].update({
                "current": current,
                "total": total,
                "message": message,
            })

        # Run synchronous translation in thread pool
        result = await asyncio.to_thread(
            translator.translate_excel_sync,
            input_file,
            output_file,
            target_lang,
            progress_callback,
        )

        translation_tasks[translate_task_id]["status"] = "completed"
        translation_tasks[translate_task_id]["message"] = "Translation completed"
        translation_tasks[translate_task_id]["result"] = result

        logger.info(f"Translation completed for task {translate_task_id}: {result}")

    except TranslationAPIError as e:
        logger.error(f"Translation API error for task {translate_task_id}: {e.message}")
        translation_tasks[translate_task_id]["status"] = "failed"
        translation_tasks[translate_task_id]["error"] = e.message
        translation_tasks[translate_task_id]["message"] = f"API Error: {e.message}"
    except Exception as e:
        logger.error(f"Translation failed for task {translate_task_id}: {e}", exc_info=True)
        translation_tasks[translate_task_id]["status"] = "failed"
        translation_tasks[translate_task_id]["error"] = str(e)
        translation_tasks[translate_task_id]["message"] = f"Translation failed: {str(e)}"


@app.post("/api/translate/start", response_model=TranslateStartResponse)
async def start_translation(
    request: TranslateRequest,
    background_tasks: BackgroundTasks,
):
    """
    Start translating an Excel file to Tamil, Hindi, or English.

    Translates cell content while preserving all Excel formatting.
    Use /api/translate/progress/{translate_task_id} to track progress via SSE.
    """
    # Validate task exists and is completed
    if request.task_id not in tasks:
        raise HTTPException(status_code=404, detail="Task not found")

    task = tasks[request.task_id]

    if task.status != "completed":
        raise HTTPException(
            status_code=400,
            detail=f"Task not completed. Current status: {task.status}",
        )

    if not task.output_file or not Path(task.output_file).exists():
        raise HTTPException(status_code=404, detail="Output file not found")

    # Use target language for caching
    target_lang = request.target_lang
    cached_file = OUTPUT_DIR / f"{request.task_id}_{target_lang}.xlsx"

    # Check if translation already exists (cached)
    if cached_file.exists():
        logger.info(f"Returning cached translation for task {request.task_id} ({target_lang})")
        return TranslateStartResponse(
            translate_task_id=f"cached_{request.task_id}_{target_lang}",
            total_cells=0,
            message=f"Translation already exists (cached)",
        )

    try:
        import openpyxl

        # Count cells to translate for progress tracking
        wb = openpyxl.load_workbook(task.output_file, data_only=True)
        ws = wb.active

        cell_count = 0
        for row in range(1, (ws.max_row or 0) + 1):
            for col in range(1, (ws.max_column or 0) + 1):
                cell = ws.cell(row=row, column=col)
                if cell.value and isinstance(cell.value, str) and not cell.value.startswith("="):
                    cell_count += 1

        wb.close()

        # Generate translate task ID
        translate_task_id = str(uuid.uuid4())

        # Initialize translation task
        translation_tasks[translate_task_id] = {
            "task_id": request.task_id,
            "target_lang": request.target_lang,
            "status": "pending",
            "current": 0,
            "total": cell_count,
            "message": "Starting translation...",
            "cancelled": False,
            "created_at": datetime.utcnow(),
            "output_file": str(cached_file),
        }

        # Start background translation
        background_tasks.add_task(
            process_translation,
            translate_task_id,
            request.task_id,
            task.output_file,
            str(cached_file),
            request.target_lang,
        )

        logger.info(f"Started translation task {translate_task_id} with {cell_count} cells")

        return TranslateStartResponse(
            translate_task_id=translate_task_id,
            total_cells=cell_count,
            message=f"Translation started for {cell_count} cells",
        )

    except Exception as e:
        logger.error(f"Failed to start translation: {e}", exc_info=True)
        raise HTTPException(
            status_code=500,
            detail=f"Failed to start translation: {str(e)}",
        )


@app.get("/api/translate/progress/{translate_task_id}")
async def translate_progress_stream(translate_task_id: str):
    """
    Server-Sent Events endpoint for real-time translation progress updates.

    Connect to this endpoint to receive live progress updates during translation.
    """
    # Handle cached translations
    if translate_task_id.startswith("cached_"):
        async def cached_event():
            event = TranslateProgressEvent(
                current=1,
                total=1,
                status="completed",
                message="Translation loaded from cache",
            )
            yield {
                "event": "progress",
                "data": json.dumps(event.model_dump()),
            }
        return EventSourceResponse(cached_event())

    if translate_task_id not in translation_tasks:
        raise HTTPException(status_code=404, detail="Translation task not found")

    async def event_generator():
        last_current = -1
        while True:
            if translate_task_id in translation_tasks:
                task = translation_tasks[translate_task_id]

                # Only send update if progress changed
                if task["current"] != last_current or task["status"] in ["completed", "failed", "cancelled"]:
                    last_current = task["current"]
                    event = TranslateProgressEvent(
                        current=task["current"],
                        total=task["total"],
                        status=task["status"],
                        message=task["message"],
                    )
                    yield {
                        "event": "progress",
                        "data": json.dumps(event.model_dump()),
                    }

                # Stop streaming when complete or failed
                if task["status"] in ["completed", "failed", "cancelled"]:
                    break

            await asyncio.sleep(0.3)

    return EventSourceResponse(event_generator())


@app.get("/api/translate/status/{task_id}", response_model=TranslateStatusResponse)
async def get_translate_status(task_id: str):
    """
    Check if translated versions exist for a task.

    Returns availability of Tamil, Hindi, and English translations.
    """
    if task_id not in tasks:
        raise HTTPException(status_code=404, detail="Task not found")

    tamil_file = OUTPUT_DIR / f"{task_id}_tamil.xlsx"
    hindi_file = OUTPUT_DIR / f"{task_id}_hindi.xlsx"
    english_file = OUTPUT_DIR / f"{task_id}_english.xlsx"

    return TranslateStatusResponse(
        task_id=task_id,
        has_tamil_version=tamil_file.exists(),
        has_hindi_version=hindi_file.exists(),
        has_english_version=english_file.exists(),
        tamil_file_path=str(tamil_file) if tamil_file.exists() else None,
        hindi_file_path=str(hindi_file) if hindi_file.exists() else None,
        english_file_path=str(english_file) if english_file.exists() else None,
    )


@app.post("/api/translate/cancel/{translate_task_id}")
async def cancel_translation(translate_task_id: str):
    """Cancel an ongoing translation operation."""
    if translate_task_id not in translation_tasks:
        raise HTTPException(status_code=404, detail="Translation task not found")

    translation_tasks[translate_task_id]["cancelled"] = True
    translation_tasks[translate_task_id]["status"] = "cancelled"
    translation_tasks[translate_task_id]["message"] = "Translation cancelled by user"

    return {"message": "Translation cancelled"}


@app.get("/api/download/{task_id}/{language}")
async def download_translated_excel(task_id: str, language: str):
    """
    Download Excel file by language.

    Args:
        task_id: Original conversion task ID
        language: "original", "tamil", "hindi", or "english"

    Returns:
        FileResponse with the Excel file
    """
    if task_id not in tasks:
        raise HTTPException(status_code=404, detail="Task not found")

    task = tasks[task_id]

    if task.status != "completed":
        raise HTTPException(
            status_code=400,
            detail=f"Task not completed. Current status: {task.status}",
        )

    # Determine file path based on language
    if language == "original":
        file_path = Path(task.output_file) if task.output_file else None
    elif language == "tamil":
        file_path = OUTPUT_DIR / f"{task_id}_tamil.xlsx"
    elif language == "hindi":
        file_path = OUTPUT_DIR / f"{task_id}_hindi.xlsx"
    elif language == "english":
        file_path = OUTPUT_DIR / f"{task_id}_english.xlsx"
    else:
        raise HTTPException(
            status_code=400,
            detail=f"Invalid language: {language}. Must be 'original', 'tamil', 'hindi', or 'english'",
        )

    if not file_path or not file_path.exists():
        raise HTTPException(
            status_code=404,
            detail=f"Translation file not found for language: {language}",
        )

    # Generate download filename
    original_name = Path(task.filename).stem
    lang_suffix = f"_{language}" if language != "original" else ""
    download_name = f"{original_name}{lang_suffix}.xlsx"

    return FileResponse(
        str(file_path),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename=download_name,
    )


@app.get("/api/preview-full/{task_id}/{language}")
async def get_translated_full_preview(task_id: str, language: str):
    """
    Get full preview data for a translated Excel file.

    Args:
        task_id: Original conversion task ID
        language: "original", "tamil", "hindi", or "english"

    Returns:
        FullPreviewData with all rows from the translated file
    """
    if task_id not in tasks:
        raise HTTPException(status_code=404, detail="Task not found")

    task = tasks[task_id]

    # Determine file path based on language
    if language == "original":
        file_path = Path(task.output_file) if task.output_file else None
    elif language == "tamil":
        file_path = OUTPUT_DIR / f"{task_id}_tamil.xlsx"
    elif language == "hindi":
        file_path = OUTPUT_DIR / f"{task_id}_hindi.xlsx"
    elif language == "english":
        file_path = OUTPUT_DIR / f"{task_id}_english.xlsx"
    else:
        raise HTTPException(
            status_code=400,
            detail=f"Invalid language: {language}. Must be 'original', 'tamil', 'hindi', or 'english'",
        )

    if not file_path or not file_path.exists():
        raise HTTPException(
            status_code=404,
            detail=f"Translation file not found for language: {language}",
        )

    try:
        import openpyxl

        wb = openpyxl.load_workbook(str(file_path), data_only=True)
        ws = wb.active

        # Find actual data range
        actual_max_row = ws.max_row if ws.max_row else 0
        actual_max_col = ws.max_column if ws.max_column else 0

        # Scan to find actual data boundaries
        last_data_row = actual_max_row
        for row in range(1, min(actual_max_row + 100, 2000)):
            for col in range(1, min(actual_max_col + 100, 200)):
                cell = ws.cell(row, col)
                if cell.value is not None and str(cell.value).strip():
                    actual_max_row = max(actual_max_row, row)
                    actual_max_col = max(actual_max_col, col)
                    last_data_row = row
            if row > last_data_row + 50:
                break

        if actual_max_row == 0:
            actual_max_row = ws.max_row if ws.max_row else 100
        if actual_max_col == 0:
            actual_max_col = ws.max_column if ws.max_column else 50

        # Find data start row
        data_start_row = 1
        document_title = None
        for row in range(1, min(20, actual_max_row + 1)):
            cell_value = ws.cell(row, 1).value
            if cell_value and str(cell_value).strip():
                if row == 1 and "FORM" in str(cell_value).upper():
                    document_title = str(cell_value).strip()
                    continue
                non_empty_count = sum(
                    1 for col in range(1, min(actual_max_col + 1, 20))
                    if ws.cell(row, col).value and str(ws.cell(row, col).value).strip()
                )
                if non_empty_count >= 2:
                    data_start_row = row
                    break

        # Find first data row
        first_data_row = data_start_row + 1
        for row in range(data_start_row + 1, min(data_start_row + 10, actual_max_row + 1)):
            cell_val = ws.cell(row, 1).value
            if cell_val is not None:
                if isinstance(cell_val, (int, float)) or (isinstance(cell_val, str) and cell_val.strip().isdigit()):
                    first_data_row = row
                    break

        num_header_rows = first_data_row - data_start_row

        # Get headers
        headers = []
        for col in range(1, actual_max_col + 1):
            header_parts = []
            for header_row in range(data_start_row, data_start_row + num_header_rows):
                value = ws.cell(header_row, col).value
                if value is not None and str(value).strip():
                    clean_value = str(value).strip()
                    if clean_value.upper() not in ['PARTY ABBREVIATION', 'NO. OF VALID VOTES CAST IN FAVOUR OF']:
                        header_parts.append(clean_value)

            if header_parts:
                if len(header_parts) == 1:
                    headers.append(header_parts[0])
                else:
                    combined = " - ".join(header_parts[-2:]) if len(header_parts) >= 2 else header_parts[-1]
                    headers.append(combined)
            else:
                headers.append(f"Column {col}")

        if not headers:
            for col in range(1, min(actual_max_col + 1, 50)):
                headers.append(f"Column {col}")

        # Get ALL data rows
        rows = []
        for row_idx in range(first_data_row, actual_max_row + 1):
            row_data = []
            num_cols_to_read = max(len(headers), actual_max_col)
            for col in range(1, num_cols_to_read + 1):
                value = ws.cell(row_idx, col).value
                row_data.append(value)
            while len(row_data) < len(headers):
                row_data.append(None)
            if len(row_data) > len(headers):
                row_data = row_data[:len(headers)]
            rows.append(row_data)

        total_rows = len(rows)
        wb.close()

        return FullPreviewData(
            headers=headers,
            rows=rows,
            total_rows=total_rows,
            total_columns=len(headers),
            pages_processed=1,
            document_title=document_title,
        )

    except Exception as e:
        logger.error(f"Failed to read translated preview: {e}", exc_info=True)
        raise HTTPException(
            status_code=500,
            detail=f"Failed to read preview: {str(e)}",
        )


# Startup event for cleanup task
@app.on_event("startup")
async def startup_event():
    """Initialize cleanup background task."""

    logger.info("PDF to Excel Converter API started")

    async def periodic_cleanup():
        from datetime import timedelta

        while True:
            await asyncio.sleep(3600)  # Run every hour

            # Cleanup old tasks (older than 1 hour)
            cutoff = datetime.utcnow() - timedelta(hours=1)
            tasks_to_delete = []

            for task_id, task in tasks.items():
                if task.created_at < cutoff:
                    tasks_to_delete.append(task_id)

            for task_id in tasks_to_delete:
                try:
                    task = tasks[task_id]
                    if task.output_file and Path(task.output_file).exists():
                        await cleanup_file(Path(task.output_file))
                    del tasks[task_id]
                except Exception:
                    pass

            # Cleanup orphaned files
            for file_path in UPLOAD_DIR.glob("*.pdf"):
                try:
                    mtime = datetime.fromtimestamp(file_path.stat().st_mtime)
                    if mtime < cutoff:
                        await cleanup_file(file_path)
                except Exception:
                    pass

            for file_path in OUTPUT_DIR.glob("*.xlsx"):
                try:
                    mtime = datetime.fromtimestamp(file_path.stat().st_mtime)
                    if mtime < cutoff:
                        await cleanup_file(file_path)
                except Exception:
                    pass

    asyncio.create_task(periodic_cleanup())
