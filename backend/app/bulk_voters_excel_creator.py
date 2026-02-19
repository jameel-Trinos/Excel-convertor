"""Consolidated Excel creator for bulk voter processing.

Creates a single Excel workbook with:
- Sheet 1 "Summary": booth-wise overview with hyperlinks to each booth sheet
- One sheet per booth: identical format to single-file VotersExcelCreator
- Final sheet "Errors" (conditional): list of failed PDFs
"""

import logging
import re
from datetime import datetime
from typing import Any, Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)

# Bilingual column headers (same as VotersExcelCreator)
VOTER_COLUMNS = [
    ("Serial No", "வ.எண்"),
    ("Name", "பெயர்"),
    ("Father/Husband Name", "தந்தை/கணவர் பெயர்"),
    ("House No", "வீட்டு எண்"),
    ("Age", "வயது"),
    ("Gender", "பாலினம்"),
    ("Voter ID", "வாக்காளர் அடையாள எண்"),
]

# Styles
_DARK_BLUE = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
_HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
_TAMIL_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=10)
_META_LABEL_FONT = Font(name="Calibri", bold=True, size=11, color="1F4E79")
_META_VALUE_FONT = Font(name="Calibri", size=11)
_DATA_FONT = Font(name="Calibri", size=10)
_THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)
_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
_LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
_COL_WIDTHS = [12, 25, 25, 15, 8, 12, 22]

# Summary styles
_SUMMARY_HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
_SUMMARY_FONT = Font(name="Calibri", size=10)
_LINK_FONT = Font(name="Calibri", size=10, color="0563C1", underline="single")
_LIGHT_FILL = PatternFill(start_color="F2F7FB", end_color="F2F7FB", fill_type="solid")
_ERROR_FILL = PatternFill(start_color="FFF0F0", end_color="FFF0F0", fill_type="solid")

# Excel sheet name max length is 31 chars and cannot contain: \ / * ? : [ ]
_SHEET_NAME_BAD_CHARS = re.compile(r'[\\/*?\[\]:]')


def _safe_sheet_name(ac_no: str, part_no: str, filename: str, index: int) -> str:
    """Create a valid Excel sheet name for a booth using AC number."""
    if ac_no:
        name = f"AC {ac_no}"
        if part_no:
            name += f" - {part_no}"
    elif part_no:
        name = f"Booth {part_no}"
    else:
        # Use filename without extension
        base = filename.rsplit(".", 1)[0] if "." in filename else filename
        name = base[:25]

    # Remove invalid chars
    name = _SHEET_NAME_BAD_CHARS.sub("", name)
    # Truncate to 31 chars (Excel limit)
    name = name[:31].strip()
    if not name:
        name = f"Sheet {index + 1}"
    return name


class BulkVotersExcelCreator:
    """Create a multi-sheet voter list Excel — one sheet per booth."""

    def create(
        self,
        booth_data: list[dict],
        booth_groups: list[dict],
        output_path: str,
        ac_no: str = "",
        total_pdfs: int = 0,
        successful_pdfs: int = 0,
        failed_pdfs: Optional[list[dict]] = None,
    ) -> str:
        """Create Excel with one sheet per booth + Summary.

        Args:
            booth_data: Per-booth extraction results, each containing:
                {part_no, ac_no, address, total_voters, voters, headers, filename}
            booth_groups: Per-booth metadata for summary.
            output_path: Where to save the .xlsx file.
            ac_no: Assembly Constituency number.
            total_pdfs: Total PDFs attempted.
            successful_pdfs: Count of successful PDFs.
            failed_pdfs: List of {filename, error} for failures.

        Returns:
            Path to created file.
        """
        failed_pdfs = failed_pdfs or []
        wb = Workbook()

        # Sheet 1: Summary (rename the default sheet)
        ws_summary = wb.active
        ws_summary.title = "Summary"

        # Track sheet names to handle duplicates
        used_names: set[str] = {"Summary", "Errors"}

        # Create one sheet per booth
        sheet_names: list[str] = []
        for i, bd in enumerate(booth_data):
            name = _safe_sheet_name(bd.get("ac_no", ac_no), bd["part_no"], bd["filename"], i)

            # Handle duplicate sheet names
            original_name = name
            counter = 2
            while name in used_names:
                suffix = f" ({counter})"
                name = original_name[: 31 - len(suffix)] + suffix
                counter += 1
            used_names.add(name)
            sheet_names.append(name)

            ws = wb.create_sheet(name)
            self._write_booth_sheet(
                ws,
                voters=bd["voters"],
                ac_no=bd.get("ac_no", ac_no),
                part_no=bd["part_no"],
                address=bd["address"],
                total_voters=bd["total_voters"],
                filename=bd["filename"],
            )

        # Write Summary sheet (with hyperlinks to each booth sheet)
        self._write_summary_sheet(
            ws_summary, booth_groups, sheet_names, ac_no, failed_pdfs,
            total_pdfs, len(booth_data),
        )

        # Errors sheet (conditional)
        if failed_pdfs:
            ws_errors = wb.create_sheet("Errors")
            self._write_errors_sheet(ws_errors, failed_pdfs)

        wb.save(output_path)
        total_voters = sum(len(bd["voters"]) for bd in booth_data)
        logger.info(
            f"Bulk voter Excel saved: {output_path} "
            f"({total_voters} voters across {len(booth_data)} sheets)"
        )
        return output_path

    def _write_booth_sheet(
        self,
        ws,
        voters: list[list],
        ac_no: str,
        part_no: str,
        address: str,
        total_voters: str,
        filename: str,
    ):
        """Write a single booth sheet — same format as VotersExcelCreator."""
        num_cols = len(VOTER_COLUMNS)
        current_row = 1

        # --- Metadata header ---
        # Title
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=num_cols)
        title_cell = ws.cell(row=1, column=1, value="Voter List / வாக்காளர் பட்டியல்")
        title_cell.font = Font(name="Calibri", bold=True, size=14, color="1F4E79")
        title_cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 30
        current_row = 2

        meta_items = [
            ("AC No / சட்டமன்றத் தொகுதி எண்:", ac_no or "—"),
            ("Booth No / பகுதி எண்:", part_no or "—"),
            ("Address / முகவரி:", address or "—"),
            ("Total Voters / மொத்த வாக்காளர்கள்:", total_voters or "—"),
        ]

        for label, value in meta_items:
            ws.merge_cells(
                start_row=current_row, start_column=1,
                end_row=current_row, end_column=3,
            )
            label_cell = ws.cell(row=current_row, column=1, value=label)
            label_cell.font = _META_LABEL_FONT
            label_cell.alignment = Alignment(
                horizontal="right", vertical="center", wrap_text=True,
            )
            ws.merge_cells(
                start_row=current_row, start_column=4,
                end_row=current_row, end_column=num_cols,
            )
            value_cell = ws.cell(row=current_row, column=4, value=value)
            value_cell.font = _META_VALUE_FONT
            value_cell.alignment = Alignment(horizontal="left", vertical="center")
            ws.row_dimensions[current_row].height = 25
            current_row += 1

        # Blank separator
        current_row += 1

        # --- Bilingual column headers ---
        eng_row = current_row
        for col_idx, (eng, _tam) in enumerate(VOTER_COLUMNS, 1):
            cell = ws.cell(row=eng_row, column=col_idx, value=eng)
            cell.font = _HEADER_FONT
            cell.fill = _DARK_BLUE
            cell.alignment = _CENTER
            cell.border = _THIN_BORDER
        current_row += 1

        tam_row = current_row
        for col_idx, (_eng, tam) in enumerate(VOTER_COLUMNS, 1):
            cell = ws.cell(row=tam_row, column=col_idx, value=tam)
            cell.font = _TAMIL_FONT
            cell.fill = _DARK_BLUE
            cell.alignment = _CENTER
            cell.border = _THIN_BORDER
        current_row += 1

        # --- Data rows ---
        data_start = current_row
        for row_data in voters:
            for col_idx in range(num_cols):
                value = row_data[col_idx] if col_idx < len(row_data) else ""
                cell = ws.cell(row=current_row, column=col_idx + 1, value=value)
                cell.font = _DATA_FONT
                cell.border = _THIN_BORDER
                if col_idx in (0, 4, 5, 6):
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                else:
                    cell.alignment = _LEFT
            current_row += 1

        # --- Formatting ---
        for col_idx, width in enumerate(_COL_WIDTHS, 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = width

        ws.row_dimensions[eng_row].height = 40
        ws.row_dimensions[tam_row].height = 40
        for r in range(data_start, current_row):
            ws.row_dimensions[r].height = 20

        ws.freeze_panes = ws.cell(row=tam_row + 1, column=1)

        # Alternating row colors
        for r in range(data_start, current_row):
            if (r - data_start) % 2 == 1:
                for c in range(1, num_cols + 1):
                    ws.cell(row=r, column=c).fill = _LIGHT_FILL

    def _write_summary_sheet(
        self,
        ws,
        booth_groups: list[dict],
        sheet_names: list[str],
        ac_no: str,
        failed_pdfs: list[dict],
        total_pdfs: int,
        successful_count: int,
    ):
        """Write summary with hyperlinks to each booth sheet."""
        total_voters = sum(bg["voter_count"] for bg in booth_groups)

        # Title
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=6)
        title = ws.cell(
            row=1, column=1,
            value=f"Constituency Summary — AC {ac_no or '—'}",
        )
        title.font = Font(name="Calibri", bold=True, size=14, color="1F4E79")
        title.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 30

        # Stats row
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=6)
        stats = ws.cell(
            row=2, column=1,
            value=(
                f"{total_voters:,} voters  |  {successful_count} booths  |  "
                f"{len(failed_pdfs)} failed  |  "
                f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
            ),
        )
        stats.font = Font(name="Calibri", size=10, color="666666")
        stats.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[2].height = 22

        # Headers
        summary_headers = [
            "Booth No", "Address", "Voter Count", "Source File", "Status", "Go to Sheet",
        ]
        for col_idx, header in enumerate(summary_headers, 1):
            cell = ws.cell(row=4, column=col_idx, value=header)
            cell.font = _SUMMARY_HEADER_FONT
            cell.fill = _DARK_BLUE
            cell.alignment = _CENTER
            cell.border = _THIN_BORDER

        # Data rows — successful booths with hyperlinks
        row = 5
        for i, bg in enumerate(booth_groups):
            ws.cell(row=row, column=1, value=bg["part_no"]).alignment = _CENTER
            ws.cell(row=row, column=2, value=bg["address"]).alignment = _LEFT
            ws.cell(row=row, column=3, value=bg["voter_count"]).alignment = _CENTER
            ws.cell(row=row, column=4, value=bg["filename"]).alignment = _LEFT
            ws.cell(row=row, column=5, value="OK").alignment = _CENTER

            # Hyperlink to the booth's sheet
            if i < len(sheet_names):
                link_cell = ws.cell(row=row, column=6, value="Open →")
                safe_name = sheet_names[i].replace("'", "''")
                link_cell.hyperlink = f"#'{safe_name}'!A1"
                link_cell.font = _LINK_FONT
                link_cell.alignment = _CENTER

            for c in range(1, 7):
                cell = ws.cell(row=row, column=c)
                if cell.font == Font():  # only set font if not already set (link cell)
                    cell.font = _SUMMARY_FONT
                cell.border = _THIN_BORDER
                if (row - 5) % 2 == 1:
                    cell.fill = _LIGHT_FILL
            row += 1

        # Failed PDFs
        for fp in failed_pdfs:
            ws.cell(row=row, column=1, value="—").alignment = _CENTER
            ws.cell(row=row, column=2, value="—").alignment = _LEFT
            ws.cell(row=row, column=3, value=0).alignment = _CENTER
            ws.cell(row=row, column=4, value=fp["filename"]).alignment = _LEFT
            ws.cell(row=row, column=5, value="FAILED").alignment = _CENTER
            ws.cell(row=row, column=6, value=fp.get("error", "")[:50]).alignment = _LEFT
            for c in range(1, 7):
                cell = ws.cell(row=row, column=c)
                cell.font = _SUMMARY_FONT
                cell.border = _THIN_BORDER
                cell.fill = _ERROR_FILL
            row += 1

        # Total row
        row += 1
        ws.cell(row=row, column=1, value="TOTAL").font = Font(
            name="Calibri", bold=True, size=11,
        )
        ws.cell(row=row, column=3, value=total_voters).font = Font(
            name="Calibri", bold=True, size=11,
        )
        ws.cell(
            row=row, column=5,
            value=f"{successful_count} OK / {len(failed_pdfs)} Failed",
        ).font = Font(name="Calibri", bold=True, size=10)
        for c in range(1, 7):
            ws.cell(row=row, column=c).border = _THIN_BORDER
            ws.cell(row=row, column=c).alignment = _CENTER

        # Column widths
        widths = [12, 40, 15, 30, 12, 14]
        for col_idx, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = w

        ws.freeze_panes = ws.cell(row=5, column=1)

    def _write_errors_sheet(self, ws, failed_pdfs: list[dict]):
        """Write the errors sheet listing failed PDFs."""
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=2)
        title = ws.cell(row=1, column=1, value="Failed PDFs")
        title.font = Font(name="Calibri", bold=True, size=14, color="CC0000")
        title.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 30

        for col_idx, header in enumerate(["Filename", "Error"], 1):
            cell = ws.cell(row=3, column=col_idx, value=header)
            cell.font = _SUMMARY_HEADER_FONT
            cell.fill = PatternFill(
                start_color="CC0000", end_color="CC0000", fill_type="solid",
            )
            cell.alignment = _CENTER
            cell.border = _THIN_BORDER

        row = 4
        for fp in failed_pdfs:
            ws.cell(row=row, column=1, value=fp["filename"]).font = _SUMMARY_FONT
            ws.cell(row=row, column=2, value=fp["error"]).font = _SUMMARY_FONT
            for c in range(1, 3):
                ws.cell(row=row, column=c).border = _THIN_BORDER
                ws.cell(row=row, column=c).alignment = _LEFT
            row += 1

        ws.column_dimensions["A"].width = 40
        ws.column_dimensions["B"].width = 60
