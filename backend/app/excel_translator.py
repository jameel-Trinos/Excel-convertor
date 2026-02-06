"""
Excel Translator - Translates Excel content while preserving ALL formatting.

This module handles translating Excel cell values between English and Tamil
while maintaining:
- Column widths
- Row heights
- Merged cells
- Cell styles (fill, font, alignment, border)
- Number formats
- Freeze panes
- Formulas (preserved as-is)
"""

import os
import logging
import asyncio
from pathlib import Path
from typing import Dict, List, Optional, Callable, Any
from copy import copy

import openpyxl
from openpyxl.styles import Font, Fill, Alignment, Border, PatternFill
from openpyxl.cell.cell import Cell

from .translation_service import TranslationService, TranslationAPIError

logger = logging.getLogger(__name__)


class ExcelTranslator:
    """Translate Excel files while preserving all formatting."""

    def __init__(self, translation_service: Optional[TranslationService] = None):
        """
        Initialize ExcelTranslator.

        Args:
            translation_service: TranslationService instance (creates one if not provided)
        """
        self.translator = translation_service or TranslationService()

    def _copy_cell_style(self, source_cell: Cell, target_cell: Cell):
        """
        Copy all styles from source cell to target cell.

        Args:
            source_cell: Cell to copy styles from
            target_cell: Cell to copy styles to
        """
        if source_cell.has_style:
            target_cell.font = copy(source_cell.font)
            target_cell.fill = copy(source_cell.fill)
            target_cell.border = copy(source_cell.border)
            target_cell.alignment = copy(source_cell.alignment)
            target_cell.number_format = source_cell.number_format
            target_cell.protection = copy(source_cell.protection)

    async def translate_excel(
        self,
        input_path: str,
        output_path: str,
        target_lang: str = "tamil",
        progress_callback: Optional[Callable[[int, int, str], None]] = None,
    ) -> Dict[str, Any]:
        """
        Translate Excel file content while preserving formatting.

        Args:
            input_path: Path to source Excel file
            output_path: Path for translated Excel file
            target_lang: Target language (tamil, hindi, english)
            progress_callback: Optional callback(current, total, message) for progress

        Returns:
            Metadata about the translation operation
        """
        logger.info(f"Starting Excel translation: {input_path} -> {output_path}, target_lang={target_lang}")

        if not os.path.exists(input_path):
            raise FileNotFoundError(f"Input file not found: {input_path}")

        # Load workbook with all data and formatting
        wb = openpyxl.load_workbook(input_path)

        # Collect all cells that need translation
        cells_to_translate: List[Dict] = []

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]

            # Iterate through all cells
            for row_idx in range(1, ws.max_row + 1):
                for col_idx in range(1, ws.max_column + 1):
                    cell = ws.cell(row=row_idx, column=col_idx)
                    value = cell.value

                    # Skip empty cells
                    if value is None:
                        continue

                    # Skip non-string values (numbers, dates, etc.)
                    if not isinstance(value, str):
                        continue

                    # Skip formulas (start with =)
                    if value.startswith("="):
                        continue

                    # Check if this text should be translated
                    if self.translator.should_translate(value.strip(), target_lang):
                        cells_to_translate.append({
                            "sheet": sheet_name,
                            "row": row_idx,
                            "col": col_idx,
                            "value": value,
                        })

        total_cells = len(cells_to_translate)
        logger.info(f"Found {total_cells} cells to translate")

        if progress_callback:
            await asyncio.to_thread(progress_callback, 0, total_cells, "Starting translation... (0%)")

        if total_cells == 0:
            # No cells to translate, just save a copy
            wb.save(output_path)
            wb.close()
            return {
                "translated_cells": 0,
                "total_cells": 0,
                "input_file": input_path,
                "output_file": output_path,
                "target_lang": target_lang,
            }

        # Extract unique texts for efficient translation
        unique_texts: Dict[str, None] = {}
        for cell_info in cells_to_translate:
            text = cell_info["value"].strip()
            if text:
                unique_texts[text] = None

        unique_text_list = list(unique_texts.keys())
        logger.info(f"Found {len(unique_text_list)} unique texts to translate")

        if progress_callback:
            await asyncio.to_thread(
                progress_callback, 0, total_cells, f"Translating {len(unique_text_list)} unique texts... (0%)"
            )

        # Batch translate all unique texts
        def translate_progress(current: int, total: int):
            if progress_callback:
                # Map unique text progress to cell progress
                cell_progress = int((current / total) * total_cells * 0.8)  # 80% for translation
                percent = int((current / total) * 80) if total > 0 else 0
                asyncio.create_task(
                    asyncio.to_thread(
                        progress_callback,
                        cell_progress,
                        total_cells,
                        f"Translating text {current}/{total}... ({percent}%)",
                    )
                )

        # Perform translation in a thread to not block
        translated_texts = await asyncio.to_thread(
            lambda: asyncio.get_event_loop().run_until_complete(
                self.translator.translate_batch(unique_text_list, target_lang)
            )
            if asyncio.get_event_loop().is_running()
            else self._sync_translate_batch(unique_text_list, target_lang)
        )

        # Build translation map
        translation_map: Dict[str, str] = {}
        for original, translated in zip(unique_text_list, translated_texts):
            translation_map[original] = translated

        if progress_callback:
            await asyncio.to_thread(
                progress_callback, int(total_cells * 0.8), total_cells, "Applying translations... (80%)"
            )

        # Apply translations to cells
        translated_count = 0
        for i, cell_info in enumerate(cells_to_translate):
            sheet = wb[cell_info["sheet"]]
            cell = sheet.cell(row=cell_info["row"], column=cell_info["col"])
            original_text = cell_info["value"].strip()

            if original_text in translation_map:
                # Get the translated text
                translated_text = translation_map[original_text]

                # Preserve leading/trailing whitespace from original
                leading_space = len(cell_info["value"]) - len(cell_info["value"].lstrip())
                trailing_space = len(cell_info["value"]) - len(cell_info["value"].rstrip())

                if leading_space > 0 or trailing_space > 0:
                    translated_text = (
                        cell_info["value"][:leading_space]
                        + translated_text
                        + cell_info["value"][-trailing_space:]
                        if trailing_space > 0
                        else cell_info["value"][:leading_space] + translated_text
                    )

                # Update cell value (formatting is preserved automatically)
                cell.value = translated_text
                translated_count += 1

            # Reduced callback frequency from every 10 to every 50 cells to reduce overhead
            if progress_callback and i % 50 == 0:
                progress = int(total_cells * 0.8 + (i / total_cells) * total_cells * 0.2)
                percent = int(80 + ((i + 1) / total_cells) * 20)
                await asyncio.to_thread(
                    progress_callback, progress, total_cells, f"Applying translation {i + 1}/{total_cells}... ({percent}%)"
                )

        if progress_callback:
            await asyncio.to_thread(progress_callback, total_cells, total_cells, "Saving file... (100%)")

        # Save the translated workbook
        wb.save(output_path)
        wb.close()

        logger.info(f"Translation complete: {translated_count} cells translated")

        return {
            "translated_cells": translated_count,
            "total_cells": total_cells,
            "unique_texts": len(unique_text_list),
            "input_file": input_path,
            "output_file": output_path,
            "target_lang": target_lang,
        }

    def _sync_translate_batch(self, texts: List[str], target_lang: str) -> List[str]:
        """
        Synchronously translate a batch of texts.

        Args:
            texts: List of texts to translate
            target_lang: Target language (tamil, hindi, english)

        Returns:
            List of translated texts
        """
        results = []
        for text in texts:
            translated = self.translator.translate_text(text, target_lang)
            results.append(translated)
        return results

    def translate_excel_sync(
        self,
        input_path: str,
        output_path: str,
        target_lang: str = "tamil",
        progress_callback: Optional[Callable[[int, int, str], None]] = None,
    ) -> Dict[str, Any]:
        """
        Synchronous version of translate_excel for non-async contexts.

        Args:
            input_path: Path to source Excel file
            output_path: Path for translated Excel file
            target_lang: Target language (tamil, hindi, english)
            progress_callback: Optional callback(current, total, message) for progress

        Returns:
            Metadata about the translation operation
        """
        logger.info(f"Starting sync Excel translation: {input_path} -> {output_path}")

        if not os.path.exists(input_path):
            raise FileNotFoundError(f"Input file not found: {input_path}")

        # Load workbook
        wb = openpyxl.load_workbook(input_path)

        # Collect cells to translate
        cells_to_translate: List[Dict] = []

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]

            for row_idx in range(1, ws.max_row + 1):
                for col_idx in range(1, ws.max_column + 1):
                    cell = ws.cell(row=row_idx, column=col_idx)
                    value = cell.value

                    if value is None or not isinstance(value, str):
                        continue

                    if value.startswith("="):
                        continue

                    if self.translator.should_translate(value.strip(), target_lang):
                        cells_to_translate.append({
                            "sheet": sheet_name,
                            "row": row_idx,
                            "col": col_idx,
                            "value": value,
                        })

        total_cells = len(cells_to_translate)
        logger.info(f"Found {total_cells} cells to translate")

        if progress_callback:
            progress_callback(0, total_cells, "Starting translation... (0%)")

        if total_cells == 0:
            wb.save(output_path)
            wb.close()
            return {
                "translated_cells": 0,
                "total_cells": 0,
                "input_file": input_path,
                "output_file": output_path,
                "target_lang": target_lang,
            }

        # Extract unique texts
        unique_texts: Dict[str, None] = {}
        for cell_info in cells_to_translate:
            text = cell_info["value"].strip()
            if text:
                unique_texts[text] = None

        unique_text_list = list(unique_texts.keys())

        if progress_callback:
            progress_callback(0, total_cells, f"Translating {len(unique_text_list)} unique texts... (0%)")

        # Translate all unique texts in batches for much faster processing
        translation_map: Dict[str, str] = {}

        # Use batch translation instead of one-by-one (much faster)
        def progress_wrapper(current: int, total: int):
            if progress_callback:
                progress = int((current / total) * total_cells * 0.8)
                percent = int((current / total) * 100) if total > 0 else 0
                progress_callback(progress, total_cells, f"Translating text {current}/{total}... ({percent}%)")

        # Use async batch translation if available, otherwise sync batch
        try:
            import asyncio
            loop = asyncio.get_event_loop()
            if loop.is_running():
                # Already in async context, use sync batch method
                translated_texts = self._sync_translate_batch(unique_text_list, target_lang)
            else:
                # Can use async batch
                translated_texts = loop.run_until_complete(
                    self.translator.translate_batch(
                        unique_text_list,
                        target_lang=target_lang,
                        progress_callback=progress_wrapper
                    )
                )
        except:
            # Fallback to sync batch translation
            translated_texts = self._sync_translate_batch(unique_text_list, target_lang)
        
        # Build translation map
        for text, translated in zip(unique_text_list, translated_texts):
            translation_map[text] = translated

        if progress_callback:
            progress_callback(int(total_cells * 0.8), total_cells, "Applying translations... (80%)")

        # Apply translations
        translated_count = 0
        for i, cell_info in enumerate(cells_to_translate):
            sheet = wb[cell_info["sheet"]]
            cell = sheet.cell(row=cell_info["row"], column=cell_info["col"])
            original_text = cell_info["value"].strip()

            if original_text in translation_map:
                translated_text = translation_map[original_text]

                # Preserve whitespace
                leading_space = len(cell_info["value"]) - len(cell_info["value"].lstrip())
                trailing_space = len(cell_info["value"]) - len(cell_info["value"].rstrip())

                if leading_space > 0 or trailing_space > 0:
                    translated_text = (
                        cell_info["value"][:leading_space]
                        + translated_text
                        + (cell_info["value"][-trailing_space:] if trailing_space > 0 else "")
                    )

                cell.value = translated_text
                translated_count += 1

            # Reduced callback frequency from every 10 to every 50 cells to reduce overhead
            if progress_callback and i % 50 == 0:
                progress = int(total_cells * 0.8 + (i / total_cells) * total_cells * 0.2)
                percent = int(80 + ((i + 1) / total_cells) * 20)
                progress_callback(progress, total_cells, f"Applying translation {i + 1}/{total_cells}... ({percent}%)")

        if progress_callback:
            progress_callback(total_cells, total_cells, "Saving file... (100%)")

        # Save
        wb.save(output_path)
        wb.close()

        logger.info(f"Translation complete: {translated_count} cells translated")

        return {
            "translated_cells": translated_count,
            "total_cells": total_cells,
            "unique_texts": len(unique_text_list),
            "input_file": input_path,
            "output_file": output_path,
            "target_lang": target_lang,
        }


def get_translation_cache_path(task_id: str, language: str, output_dir: str) -> Path:
    """
    Get the path for a cached translation file.

    Args:
        task_id: Original task ID
        language: "tamil" or "english"
        output_dir: Output directory path

    Returns:
        Path to the cached translation file
    """
    suffix = f"_{language}"
    return Path(output_dir) / f"{task_id}{suffix}.xlsx"


def check_translation_exists(task_id: str, language: str, output_dir: str) -> bool:
    """
    Check if a translated version exists.

    Args:
        task_id: Original task ID
        language: "tamil" or "english"
        output_dir: Output directory path

    Returns:
        True if translation exists
    """
    cache_path = get_translation_cache_path(task_id, language, output_dir)
    return cache_path.exists()
