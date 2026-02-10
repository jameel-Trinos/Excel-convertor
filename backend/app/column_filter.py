"""Service for filtering Excel columns based on user selection."""

import logging
from datetime import datetime
from pathlib import Path
from typing import List, Optional, Dict

import openpyxl
import pandas as pd
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

logger = logging.getLogger(__name__)


def _normalize_column_name(name: str) -> str:
    """
    Normalize column name for comparison.
    
    Removes extra spaces, converts to uppercase, and handles common variations.
    """
    if not name:
        return ""
    # Remove extra spaces and convert to uppercase
    normalized = " ".join(str(name).upper().split())
    # Remove common punctuation variations
    normalized = normalized.replace(".", "").replace("-", " ").replace("_", " ")
    # Remove extra spaces again
    normalized = " ".join(normalized.split())
    return normalized


def _find_column_match(requested: str, available_columns: List[str]) -> Optional[str]:
    """
    Find the best matching column from available columns.
    
    Returns the actual column name from available_columns if found, None otherwise.
    """
    requested_norm = _normalize_column_name(requested)
    
    # Try exact match first
    for col in available_columns:
        if col == requested:
            return col
    
    # Try case-insensitive match
    for col in available_columns:
        if col.upper().strip() == requested.upper().strip():
            return col
    
    # Try normalized match
    for col in available_columns:
        if _normalize_column_name(col) == requested_norm:
            return col
    
    # Try partial match (contains)
    for col in available_columns:
        col_norm = _normalize_column_name(col)
        if requested_norm in col_norm or col_norm in requested_norm:
            return col
    
    # Try word-based matching (common words)
    requested_words = set(requested_norm.split())
    best_match = None
    best_score = 0
    
    for col in available_columns:
        col_norm = _normalize_column_name(col)
        col_words = set(col_norm.split())
        # Count common words
        common_words = requested_words & col_words
        if len(common_words) > best_score and len(common_words) > 0:
            best_score = len(common_words)
            best_match = col
    
    return best_match


class ColumnFilterService:
    """Service to filter Excel files by selected columns."""

    def __init__(self):
        """Initialize the column filter service."""
        pass

    def filter_columns(
        self,
        input_file: str,
        requested_columns: List[str],
        output_dir: str,
        header_overrides: Optional[Dict[str, str]] = None,
        task_id: Optional[str] = None,
        sum_other_columns: Optional[List[str]] = None,
    ) -> tuple[str, dict]:
        """
        Filter Excel file to include only requested columns.

        Args:
            input_file: Path to the source Excel file
            requested_columns: List of column names to include (in desired order)
            output_dir: Directory to save the filtered Excel file
            header_overrides: Optional mapping of original column name -> desired output header name
            task_id: Optional task ID to check for alliance mappings (for handling merged columns)

        Returns:
            Tuple of (output_file_path, metadata_dict)

        Raises:
            FileNotFoundError: If input file doesn't exist
            ValueError: If requested columns don't exist in the Excel file
        """
        input_path = Path(input_file)
        if not input_path.exists():
            raise FileNotFoundError(f"Input file not found: {input_file}")

        logger.info(f"Loading Excel file: {input_file}")

        # First, use openpyxl to find the correct header row and extract column names
        # (EXACT same logic as preview endpoint to ensure consistency)
        import openpyxl
        wb = openpyxl.load_workbook(input_file, data_only=True)
        ws = wb.active
        
        # Find actual data range by scanning for cells with values
        # This is more reliable than max_row/max_column which can be inaccurate
        actual_max_row = ws.max_row if ws.max_row else 0
        actual_max_col = ws.max_column if ws.max_column else 0
        
        # Scan iteratively to find actual data boundaries
        last_data_row = actual_max_row
        scan_limit = max(actual_max_row + 100, 1000)
        for row in range(1, min(scan_limit + 1, 2000)):
            row_has_data = False
            for col in range(1, min(actual_max_col + 100, 200)):
                cell = ws.cell(row, col)
                if cell.value is not None and str(cell.value).strip():
                    actual_max_row = max(actual_max_row, row)
                    actual_max_col = max(actual_max_col, col)
                    row_has_data = True
                    last_data_row = row
            if row > last_data_row + 50:
                break

        # Ensure we have reasonable values
        if actual_max_row == 0:
            actual_max_row = ws.max_row if ws.max_row else 100
        if actual_max_col == 0:
            actual_max_col = ws.max_column if ws.max_column else 50
        
        # Find data start row (skip title rows) - EXACT same logic as preview
        data_start_row = 1
        for row in range(1, min(20, actual_max_row + 1)):
            cell_value = ws.cell(row, 1).value
            if cell_value and str(cell_value).strip():
                # Check if this row looks like headers (has multiple non-empty cells)
                non_empty_count = sum(
                    1 for col in range(1, min(actual_max_col + 1, 20))
                    if ws.cell(row, col).value and str(ws.cell(row, col).value).strip()
                )
                if non_empty_count >= 2:  # At least 2 columns with data
                    data_start_row = row
                    break
        
        # Find where actual data starts (first row with numeric data in first column)
        # This helps us know how many header rows there are
        first_data_row = data_start_row + 1
        for row in range(data_start_row + 1, min(data_start_row + 10, actual_max_row + 1)):
            cell_val = ws.cell(row, 1).value
            # Check if this looks like data (numeric or starts with a number)
            if cell_val is not None:
                if isinstance(cell_val, (int, float)) or (isinstance(cell_val, str) and cell_val.strip().isdigit()):
                    first_data_row = row
                    break

        # Calculate how many header rows we have (between data_start_row and first_data_row)
        num_header_rows = first_data_row - data_start_row

        # Get headers by merging multi-row headers - EXACT same logic as preview
        # Many election PDFs have 2-3 header rows (main header + candidate name + party)
        headers = []
        for col in range(1, actual_max_col + 1):
            # Get values from header rows only (not data rows)
            header_parts = []

            # Only check actual header rows (not data rows)
            for header_row in range(data_start_row, data_start_row + num_header_rows):
                value = ws.cell(header_row, col).value
                if value is not None and str(value).strip():
                    clean_value = str(value).strip()
                    # Skip generic phrases and party abbreviation labels
                    if clean_value.upper() not in ['PARTY ABBREVIATION', 'NO. OF VALID VOTES CAST IN FAVOUR OF']:
                        header_parts.append(clean_value)

            # Combine header parts into a single name
            if header_parts:
                # Use the last non-empty part (usually the party name or most specific info)
                # Or combine if there are multiple meaningful parts
                if len(header_parts) == 1:
                    headers.append(header_parts[0])
                else:
                    # For multi-part headers, prioritize party names (usually in last row)
                    # and candidate names
                    combined = " - ".join(header_parts[-2:]) if len(header_parts) >= 2 else header_parts[-1]
                    headers.append(combined)
            else:
                # Fallback for empty columns
                headers.append(f"Column {col}")

        # Ensure we have at least some headers
        if not headers:
            # Fallback: try to detect columns by scanning first few data rows
            for col in range(1, min(actual_max_col + 1, 50)):
                headers.append(f"Column {col}")

        wb.close()
        
        # Now load with pandas, skipping all header rows
        # Skip rows before data, then set our custom column names
        skip_rows = list(range(0, first_data_row - 1))  # 0-indexed, skip everything before data
        
        # Read Excel with pandas, ensuring we have the right number of columns
        try:
            # First, read just the first data row to detect number of columns
            df_sample = pd.read_excel(input_file, engine='openpyxl', skiprows=skip_rows, header=None, nrows=1)
            num_cols_detected = len(df_sample.columns)
            
            logger.info(f"Detected {num_cols_detected} columns in Excel file, extracted {len(headers)} headers")
            
            # Adjust headers to match the number of columns pandas detected
            if len(headers) < num_cols_detected:
                # Add default column names for extra columns
                logger.warning(f"Excel has {num_cols_detected} columns but only {len(headers)} headers extracted. Adding default names.")
                for i in range(len(headers), num_cols_detected):
                    headers.append(f"Column {i + 1}")
            elif len(headers) > num_cols_detected:
                # Trim headers if we have more than pandas detected
                logger.warning(f"Extracted {len(headers)} headers but Excel only has {num_cols_detected} columns. Trimming headers.")
                headers = headers[:num_cols_detected]
            
            # Now read the full data with the adjusted headers
            # Use usecols to ensure we only read the columns we have headers for
            df = pd.read_excel(
                input_file, 
                engine='openpyxl', 
                skiprows=skip_rows, 
                header=None, 
                names=headers,
                usecols=range(len(headers))  # Only read the columns we have headers for
            )
            
            # Verify that the column names match what we set
            if list(df.columns) != headers:
                logger.warning(f"Column names mismatch. Expected: {headers}, Got: {list(df.columns)}")
                # Force the column names to match
                df.columns = headers[:len(df.columns)]
                
        except Exception as e:
            logger.warning(f"Error reading with pandas using skiprows: {e}, trying alternative method")
            # Fallback: read without skiprows and manually set headers
            df = pd.read_excel(input_file, engine='openpyxl', header=None)
            # Skip the header rows manually
            if first_data_row > 1:
                df = df.iloc[first_data_row - 1:].reset_index(drop=True)
            # Set column names - ensure we don't exceed the number of columns
            num_cols = len(df.columns)
            if len(headers) <= num_cols:
                df.columns = headers[:num_cols]
            else:
                # If we have more headers than columns, use what we have
                df.columns = headers[:num_cols]
                logger.warning(f"More headers than columns. Using first {num_cols} headers.")

        # Get actual columns from Excel (now using our custom headers)
        available_columns = df.columns.tolist()
        
        # Check if we're getting default pandas column names (numeric or Unnamed patterns)
        # This indicates the headers weren't set correctly
        has_default_names = any(
            str(col).startswith('Unnamed') or 
            (isinstance(col, (int, float)) and str(col).isdigit()) or
            (isinstance(col, str) and ' - ' in str(col) and all(c.strip().isdigit() for c in str(col).split(' - ')))
            for col in available_columns
        )
        
        if has_default_names:
            logger.warning("Detected default pandas column names. Headers may not have been extracted correctly.")
            logger.warning(f"Current column names: {available_columns}")
            # Try to re-read headers directly from the Excel file using openpyxl
            # This is a fallback to ensure we get the correct headers
            wb_fallback = openpyxl.load_workbook(input_file, data_only=True)
            ws_fallback = wb_fallback.active
            
            # Re-extract headers from the header row(s)
            fallback_headers = []
            for col in range(1, min(len(available_columns) + 1, actual_max_col + 1)):
                header_parts = []
                for header_row in range(data_start_row, data_start_row + num_header_rows):
                    value = ws_fallback.cell(header_row, col).value
                    if value is not None and str(value).strip():
                        clean_value = str(value).strip()
                        if clean_value.upper() not in ['PARTY ABBREVIATION', 'NO. OF VALID VOTES CAST IN FAVOUR OF']:
                            header_parts.append(clean_value)
                
                if header_parts:
                    if len(header_parts) == 1:
                        fallback_headers.append(header_parts[0])
                    else:
                        fallback_headers.append(" - ".join(header_parts[-2:]) if len(header_parts) >= 2 else header_parts[-1])
                else:
                    fallback_headers.append(f"Column {col}")
            
            wb_fallback.close()
            
            # Update dataframe column names if we got better headers
            if len(fallback_headers) == len(available_columns):
                df.columns = fallback_headers
                available_columns = fallback_headers
                logger.info(f"Updated column names from Excel file: {available_columns}")
            else:
                logger.warning(f"Fallback header count ({len(fallback_headers)}) doesn't match column count ({len(available_columns)})")
        
        logger.info(f"Available columns ({len(available_columns)}): {available_columns}")
        logger.info(f"Requested columns ({len(requested_columns)}): {requested_columns}")

        # Try to match requested columns with available columns (fuzzy matching)
        # This handles variations in column names
        matched_columns = []
        unmatched_columns = []
        match_mapping = {}  # Track which requested column maps to which actual column
        
        for requested_col in requested_columns:
            matched = _find_column_match(requested_col, available_columns)
            if matched:
                matched_columns.append(matched)
                match_mapping[requested_col] = matched
                if matched != requested_col:
                    logger.info(f"Matched '{requested_col}' to '{matched}'")
            else:
                unmatched_columns.append(requested_col)

        # If we have unmatched columns, raise an error with helpful message
        if unmatched_columns:
            # Try to suggest similar column names using the matching function
            suggestions = {}
            for unmatched in unmatched_columns:
                # Try to find a close match (even if not perfect)
                close_match = _find_column_match(unmatched, available_columns)
                if close_match:
                    suggestions[unmatched] = [close_match]
                else:
                    # Find columns that contain key words from the unmatched column
                    unmatched_words = set(_normalize_column_name(unmatched).split())
                    similar = []
                    for avail_col in available_columns:
                        avail_words = set(_normalize_column_name(avail_col).split())
                        # Count common words
                        common = len(unmatched_words & avail_words)
                        if common > 0:
                            similar.append((avail_col, common))
                    # Sort by number of common words
                    similar.sort(key=lambda x: x[1], reverse=True)
                    if similar:
                        suggestions[unmatched] = [col for col, _ in similar[:3]]  # Top 3 suggestions
            
            error_msg = f"Requested columns not found in Excel file: {unmatched_columns}.\n"
            error_msg += f"Available columns ({len(available_columns)}): {available_columns}\n"
            if suggestions:
                error_msg += "\nSuggestions (try using these exact column names):\n"
                for unmatched, suggested in suggestions.items():
                    error_msg += f"  '{unmatched}' → try: {', '.join(suggested)}\n"
            else:
                error_msg += "\nTip: Make sure the column names match exactly (including spaces and punctuation).\n"
            
            raise ValueError(error_msg)
        
        # Use matched columns (which are the actual column names from Excel)
        requested_columns = matched_columns

        # Filter dataframe to requested columns (preserving order)
        filtered_df = df[requested_columns].copy()

        # Sum unselected party/numeric columns into an "Other" column if requested
        if sum_other_columns:
            # Match the unselected columns against available columns (same fuzzy matching)
            # Use column indices to handle duplicate column names (e.g. multiple "INDEPENDENT")
            other_col_indices = []
            for other_col in sum_other_columns:
                matched = _find_column_match(other_col, available_columns)
                if matched and matched not in requested_columns:
                    # Find ALL indices for this column name (handles duplicates)
                    for idx, col_name in enumerate(available_columns):
                        if col_name == matched and idx not in other_col_indices:
                            other_col_indices.append(idx)
                            break  # one index per requested entry

            if other_col_indices:
                logger.info(f"Summing {len(other_col_indices)} unselected columns into 'Other' (indices: {other_col_indices})")
                # Use iloc to select by position, avoiding duplicate column name issues
                other_values = df.iloc[:, other_col_indices].copy()
                # Convert each column to numeric by position
                for i in range(other_values.shape[1]):
                    other_values.iloc[:, i] = pd.to_numeric(other_values.iloc[:, i], errors='coerce').fillna(0)
                filtered_df["Others Votes"] = other_values.sum(axis=1).astype(int)
            else:
                logger.info("No valid unselected columns found for 'Other' summing")

        # Apply output header overrides (display names) AFTER selecting data by original headers
        header_overrides = header_overrides or {}
        output_headers: List[str] = []
        used: Dict[str, int] = {}

        for col_name in filtered_df.columns.tolist():
            # First check if there's a manual override
            if col_name in header_overrides:
                desired = header_overrides[col_name]
            else:
                # Keep original column name as-is
                desired = col_name
            
            desired = str(desired).strip() if desired is not None else str(col_name).strip()
            if not desired:
                desired = str(col_name).strip() or "Column"

            # Ensure uniqueness in Excel header row
            count = used.get(desired, 0) + 1
            used[desired] = count
            if count > 1:
                desired = f"{desired} ({count})"
            output_headers.append(desired)

        filtered_df.columns = output_headers

        # Generate output filename with timestamp
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_filename = f"filtered_{timestamp}.xlsx"
        output_path = Path(output_dir) / output_filename

        # Ensure output directory exists
        output_path.parent.mkdir(parents=True, exist_ok=True)

        logger.info(f"Writing filtered Excel to: {output_path}")

        # Write to Excel using openpyxl for formatting control
        filtered_df.to_excel(output_path, index=False, engine='openpyxl')

        # Apply professional formatting
        self._apply_formatting(str(output_path), output_headers)

        # Prepare metadata
        metadata = {
            "original_file": input_file,
            "filtered_file": str(output_path),
            "timestamp": timestamp,
            "original_columns": available_columns,
            "selected_columns": requested_columns,
            "output_headers": output_headers,
            "total_columns": len(filtered_df.columns),
            "total_rows": len(filtered_df),
            "columns_removed": len(available_columns) - len(requested_columns),
        }

        logger.info(f"Filtering complete: {len(requested_columns)} columns, {len(filtered_df)} rows")

        return str(output_path), metadata

    def _apply_formatting(self, excel_file: str, column_names: List[str]):
        """
        Apply professional formatting to the filtered Excel file.

        Args:
            excel_file: Path to the Excel file to format
            column_names: List of column names (for column width calculation)
        """
        logger.info(f"Applying formatting to: {excel_file}")

        wb = openpyxl.load_workbook(excel_file)
        ws = wb.active

        # Header row styling
        header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=11)
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        # Border styling
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Apply header formatting
        for col_idx, col_name in enumerate(column_names, start=1):
            cell = ws.cell(row=1, column=col_idx)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
            cell.border = thin_border

            # Set column width based on content
            max_length = len(str(col_name))
            # Check data rows for longer content (sample first 100 rows)
            for row_idx in range(2, min(102, ws.max_row + 1)):
                cell_value = ws.cell(row=row_idx, column=col_idx).value
                if cell_value:
                    max_length = max(max_length, len(str(cell_value)))

            # Set column width (min 15, max 50 characters)
            adjusted_width = min(50, max(15, max_length + 2))
            ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = adjusted_width

        # Set header row height
        ws.row_dimensions[1].height = 70

        # Apply data row formatting
        data_alignment = Alignment(horizontal="left", vertical="center")
        for row_idx in range(2, ws.max_row + 1):
            ws.row_dimensions[row_idx].height = 18
            for col_idx in range(1, len(column_names) + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.border = thin_border
                cell.alignment = data_alignment

        # Freeze header row
        ws.freeze_panes = "A2"

        # Save formatted workbook
        wb.save(excel_file)
        wb.close()

        logger.info("Formatting applied successfully")
