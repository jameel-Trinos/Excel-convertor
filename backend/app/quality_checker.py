"""Quality validation utilities for Excel files."""

import logging
from pathlib import Path
from typing import Dict, List, Optional

from openpyxl import load_workbook

logger = logging.getLogger(__name__)


class ExcelQualityChecker:
    """
    Validates Excel files for quality and formatting issues.

    Checks for:
    - Column count consistency
    - Row count validation
    - Formula errors
    - Proper formatting (widths, heights, borders)
    - Data completeness
    """

    def __init__(self):
        """Initialize the quality checker."""
        self.errors = []
        self.warnings = []

    def check_file(self, file_path: str) -> Dict[str, any]:
        """
        Run all quality checks on an Excel file.

        Args:
            file_path: Path to the Excel file

        Returns:
            Dictionary with check results
        """
        self.errors = []
        self.warnings = []

        if not Path(file_path).exists():
            self.errors.append(f"File not found: {file_path}")
            return self._build_report()

        try:
            workbook = load_workbook(file_path, data_only=False)
            worksheet = workbook.active

            # Run checks
            self._check_dimensions(worksheet)
            self._check_formulas(worksheet)
            self._check_formatting(worksheet)
            self._check_data_completeness(worksheet)

            workbook.close()

        except Exception as e:
            self.errors.append(f"Error loading file: {str(e)}")

        return self._build_report()

    def _check_dimensions(self, worksheet) -> None:
        """Check column widths and row heights."""
        # Check if columns have proper widths (not default 8.43)
        default_width_count = 0
        for col in range(1, worksheet.max_column + 1):
            col_letter = worksheet.cell(1, col).column_letter
            width = worksheet.column_dimensions[col_letter].width

            if width is None or width == 8.43:  # Default Excel width
                default_width_count += 1

        if default_width_count > 0:
            self.warnings.append(
                f"{default_width_count} columns using default width (may look collapsed)"
            )

        # Check if data rows have proper heights (not default 15)
        small_row_count = 0
        for row in range(1, min(worksheet.max_row + 1, 100)):  # Sample first 100 rows
            height = worksheet.row_dimensions[row].height
            if height is None or height < 15:
                small_row_count += 1

        if small_row_count > 10:
            self.warnings.append(
                f"{small_row_count} rows may be too small (could look squished)"
            )

    def _check_formulas(self, worksheet) -> None:
        """Check for formula errors."""
        # Load with data_only=True to check calculated values
        formula_errors = []

        for row in worksheet.iter_rows():
            for cell in row:
                if cell.value and isinstance(cell.value, str):
                    # Check for common Excel error values
                    if cell.value in ["#DIV/0!", "#N/A", "#NAME?", "#NULL!", "#NUM!", "#REF!", "#VALUE!"]:
                        formula_errors.append(f"{cell.coordinate}: {cell.value}")

        if formula_errors:
            self.errors.append(f"Found {len(formula_errors)} formula errors")
            for error in formula_errors[:5]:  # Show first 5
                self.errors.append(f"  - {error}")

    def _check_formatting(self, worksheet) -> None:
        """Check for proper cell formatting."""
        # Check if header row exists and has proper formatting
        if worksheet.max_row < 1:
            self.errors.append("No data found in worksheet")
            return

        # Check first data row for borders
        first_data_row = 2  # Assuming row 1 is header
        if worksheet.max_row >= first_data_row:
            cells_without_borders = 0
            for col in range(1, min(worksheet.max_column + 1, 20)):  # Sample first 20 cols
                cell = worksheet.cell(first_data_row, col)
                if not cell.border or not cell.border.left:
                    cells_without_borders += 1

            if cells_without_borders > 0:
                self.warnings.append(
                    f"{cells_without_borders} cells missing borders (unprofessional appearance)"
                )

    def _check_data_completeness(self, worksheet) -> None:
        """Check for empty rows or missing data."""
        empty_row_count = 0
        total_rows = worksheet.max_row

        for row_idx in range(1, total_rows + 1):
            row_values = [
                worksheet.cell(row_idx, col).value
                for col in range(1, worksheet.max_column + 1)
            ]
            if all(v is None or str(v).strip() == "" for v in row_values):
                empty_row_count += 1

        if empty_row_count > total_rows * 0.5:
            self.warnings.append(
                f"{empty_row_count}/{total_rows} rows are empty (may indicate extraction issues)"
            )

    def _build_report(self) -> Dict[str, any]:
        """Build the final report."""
        return {
            "passed": len(self.errors) == 0,
            "error_count": len(self.errors),
            "warning_count": len(self.warnings),
            "errors": self.errors,
            "warnings": self.warnings,
        }


def check_excel_quality(file_path: str, verbose: bool = True) -> bool:
    """
    Quick function to check Excel file quality.

    Args:
        file_path: Path to Excel file
        verbose: Print detailed report

    Returns:
        True if no errors, False otherwise
    """
    checker = ExcelQualityChecker()
    report = checker.check_file(file_path)

    if verbose:
        print(f"\n{'='*60}")
        print(f"Excel Quality Check: {file_path}")
        print(f"{'='*60}")

        if report["passed"]:
            print("✓ PASSED - No critical errors found")
        else:
            print("✗ FAILED - Critical errors found")

        if report["errors"]:
            print(f"\nErrors ({report['error_count']}):")
            for error in report["errors"]:
                print(f"  ✗ {error}")

        if report["warnings"]:
            print(f"\nWarnings ({report['warning_count']}):")
            for warning in report["warnings"]:
                print(f"  ⚠ {warning}")

        if not report["errors"] and not report["warnings"]:
            print("\n✓ Perfect! No issues found.")

        print(f"{'='*60}\n")

    return report["passed"]
