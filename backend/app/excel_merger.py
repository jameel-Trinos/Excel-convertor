"""Excel merger for voter list files.

Merges multiple system-generated voter Excel files into a single flat Excel
with a standardised column schema. Reads metadata (AC No, Booth No) from
each sheet's header section and maps voter data rows to the target columns.
"""

import logging
import os
import re
import tempfile
from typing import Any, Optional

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)

# ── Target column schema ────────────────────────────────────────────────
MERGED_COLUMNS = [
    "Serial No",
    "voter_name",
    "voter_name_tamil",
    "relation_name",
    "relation_name_tamil",
    "relation_type",
    "epic_number",
    "gender",
    "age",
    "house_number",
    "section_name",
    "section_name_tamil",
    "part_number",
    "ac_number",
    "district_code",
]

# Column widths for the merged output
_COL_WIDTHS = [10, 22, 22, 22, 22, 14, 18, 10, 8, 14, 18, 18, 14, 14, 14]

# ── Styles ───────────────────────────────────────────────────────────────
_DARK_BLUE = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
_LIGHT_FILL = PatternFill(start_color="F2F7FB", end_color="F2F7FB", fill_type="solid")
_HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
_DATA_FONT = Font(name="Calibri", size=10)
_THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)
_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
_LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)

# ── Regex helpers for metadata extraction ────────────────────────────────
_AC_NO_RE = re.compile(
    r"(?:AC\s*No|சட்டமன்றத்\s*தொகுதி\s*எண்)[^:]*:\s*(\d+)", re.IGNORECASE
)
_BOOTH_NO_RE = re.compile(
    r"(?:Booth\s*No|பகுதி\s*எண்)[^:]*:\s*(\d+)", re.IGNORECASE
)

# Sheets to skip when reading bulk voter Excel files
_SKIP_SHEETS = {"summary", "errors"}

# ── Section name mapping by booth/part number (AC 149 — Ariyalur) ────────
_SECTION_MAP: list[tuple[set[int], str, str]] = [
    (
        set(range(1, 44)) | set(range(70, 84)) | set(range(107, 110)),
        "Ariyalur North",
        "அரியலூர் வடக்கு",
    ),
    (
        set(range(84, 107)) | set(range(110, 141)),
        "Ariyalur South",
        "அரியலூர் தெற்கு",
    ),
    (
        set(range(44, 70)),
        "Ariyalur Town",
        "அரியலூர் நகரம்",
    ),
    (
        set(range(141, 165)) | set(range(168, 186)) | set(range(240, 266)),
        "Thirumanur West",
        "திருமானூர் மேற்கு",
    ),
    (
        set(range(165, 168)) | set(range(186, 240)),
        "Thirumanur East",
        "திருமானூர் கிழக்கு",
    ),
    (
        set(range(266, 278)),
        "Jayankondam ALU",
        "ஜெயங்கொண்டம் ஏ.எல்.யு",
    ),
    (
        set(range(278, 331)),
        "T Palur West",
        "டி பழூர் மேற்கு",
    ),
]


def _resolve_section(part_no: str) -> tuple[str, str]:
    """Return (section_name, section_name_tamil) for a given booth number."""
    try:
        n = int(part_no)
    except (ValueError, TypeError):
        return ("", "")
    for nums, eng, tam in _SECTION_MAP:
        if n in nums:
            return (eng, tam)
    return ("", "")


class ExcelMerger:
    """Merge multiple voter Excel files into one flat Excel."""

    def merge(self, file_paths: list[str], output_path: str) -> dict:
        """Read all voter Excel files, merge into one flat sheet.

        Args:
            file_paths: Paths to .xlsx files to merge.
            output_path: Destination path for the merged .xlsx.

        Returns:
            dict with keys: output_path, total_rows, total_files, total_sheets.
        """
        all_rows: list[list[Any]] = []
        total_sheets = 0

        for fpath in file_paths:
            try:
                sheet_rows, sheets_read = self._read_file(fpath)
                all_rows.extend(sheet_rows)
                total_sheets += sheets_read
                logger.info(
                    "Read %d rows from %d sheets in %s",
                    len(sheet_rows), sheets_read, os.path.basename(fpath),
                )
            except Exception:
                logger.exception("Failed to read %s", fpath)
                raise

        # Sort by ac_number (col 13), then part_number (col 12), then original order
        def sort_key(row: list) -> tuple:
            ac = row[13] or ""  # ac_number
            part = row[12] or ""  # part_number
            # Numeric sort where possible
            try:
                ac_n = int(ac)
            except (ValueError, TypeError):
                ac_n = 0
            try:
                part_n = int(part)
            except (ValueError, TypeError):
                part_n = 0
            return (ac_n, part_n)

        all_rows.sort(key=sort_key)

        # Assign sequential serial numbers
        for idx, row in enumerate(all_rows, 1):
            row[0] = idx

        # Write output
        self._write_merged_excel(all_rows, output_path)

        return {
            "output_path": output_path,
            "total_rows": len(all_rows),
            "total_files": len(file_paths),
            "total_sheets": total_sheets,
        }

    # ── Reading ──────────────────────────────────────────────────────────

    def _read_file(self, fpath: str) -> tuple[list[list[Any]], int]:
        """Read a single Excel file and return (rows, sheets_read)."""
        wb = load_workbook(fpath, read_only=True, data_only=True)
        all_rows: list[list[Any]] = []
        sheets_read = 0

        for sheet_name in wb.sheetnames:
            if sheet_name.lower().strip() in _SKIP_SHEETS:
                continue

            ws = wb[sheet_name]
            ac_no, part_no = self._extract_metadata(ws)
            data_rows = self._extract_data_rows(ws, ac_no, part_no)

            if data_rows:
                all_rows.extend(data_rows)
                sheets_read += 1
                logger.debug(
                    "Sheet '%s': ac=%s part=%s rows=%d",
                    sheet_name, ac_no, part_no, len(data_rows),
                )

        wb.close()
        return all_rows, sheets_read

    def _extract_metadata(self, ws) -> tuple[str, str]:
        """Extract AC No and Booth/Part No from the metadata section (rows 1-6)."""
        ac_no = ""
        part_no = ""

        for row in ws.iter_rows(min_row=1, max_row=8, max_col=7, values_only=False):
            for cell in row:
                val = cell.value
                if val is None:
                    continue
                text = str(val).strip()
                if not text:
                    continue

                # Check for AC No
                if not ac_no:
                    m = _AC_NO_RE.search(text)
                    if m:
                        ac_no = m.group(1)

                # Check for Booth/Part No
                if not part_no:
                    m = _BOOTH_NO_RE.search(text)
                    if m:
                        part_no = m.group(1)

            if ac_no and part_no:
                break

        # Fallback: check if metadata is in label+value pattern
        # (label in col A merged, value in col D)
        if not ac_no or not part_no:
            for row_idx in range(1, 9):
                label_cell = ws.cell(row=row_idx, column=1)
                value_cell = ws.cell(row=row_idx, column=4)
                label = str(label_cell.value or "").strip()
                value = str(value_cell.value or "").strip()

                if not ac_no and ("AC No" in label or "தொகுதி எண்" in label):
                    # Value might be just the number
                    digits = re.search(r"(\d+)", value)
                    if digits:
                        ac_no = digits.group(1)

                if not part_no and (
                    "Booth No" in label
                    or "பகுதி எண்" in label
                    or "Part No" in label
                ):
                    digits = re.search(r"(\d+)", value)
                    if digits:
                        part_no = digits.group(1)

        # Fallback: try to parse from sheet title (e.g. "AC 149 - 51")
        if not ac_no or not part_no:
            title = ws.title or ""
            m = re.match(r"AC\s*(\d+)\s*-\s*(\d+)", title)
            if m:
                if not ac_no:
                    ac_no = m.group(1)
                if not part_no:
                    part_no = m.group(2)

        return ac_no, part_no

    def _extract_data_rows(
        self, ws, ac_no: str, part_no: str
    ) -> list[list[Any]]:
        """Find and read voter data rows from a sheet."""
        section_name, section_tamil = _resolve_section(part_no)

        # Find the header row by scanning for "Serial No" in column A
        header_row = None
        for row_idx in range(1, 15):
            cell_val = ws.cell(row=row_idx, column=1).value
            if cell_val and "serial" in str(cell_val).lower().strip():
                header_row = row_idx
                break

        if header_row is None:
            logger.debug("No header row found in sheet '%s'", ws.title)
            return []

        # Data starts 2 rows after English header (skip Tamil header row)
        data_start = header_row + 2

        rows: list[list[Any]] = []
        for row in ws.iter_rows(min_row=data_start, max_col=7, values_only=True):
            # Skip empty rows
            if all(v is None or str(v).strip() == "" for v in row):
                continue

            # Pad row to 7 columns
            vals = list(row) + [None] * (7 - len(row))

            serial_no = str(vals[0] or "").strip()
            name = str(vals[1] or "").strip()
            relation = str(vals[2] or "").strip()
            house_no = str(vals[3] or "").strip()
            age = str(vals[4] or "").strip()
            gender = str(vals[5] or "").strip()
            voter_id = str(vals[6] or "").strip()

            # Skip rows that look like headers or metadata
            if serial_no.lower() in ("serial no", "வ.எண்", ""):
                # Allow empty serial if there's other data
                if not name and not voter_id:
                    continue

            # Build target row (serial placeholder = 0, will be renumbered)
            merged_row = [
                0,                  # Serial No (placeholder)
                name,               # voter_name
                name,               # voter_name_tamil (same Tamil text)
                relation,           # relation_name
                relation,           # relation_name_tamil (same Tamil text)
                "F",                # relation_type (default Father)
                voter_id,           # epic_number
                gender,             # gender
                age,                # age
                house_no,           # house_number
                section_name,       # section_name
                section_tamil,      # section_name_tamil
                part_no,            # part_number
                ac_no,              # ac_number
                "",                 # district_code (blank)
            ]
            rows.append(merged_row)

        return rows

    # ── Writing ──────────────────────────────────────────────────────────

    def _write_merged_excel(self, rows: list[list[Any]], output_path: str):
        """Write the merged data to a professionally formatted Excel file."""
        wb = Workbook()
        ws = wb.active
        ws.title = "Merged Voters"

        # Header row
        for col_idx, header in enumerate(MERGED_COLUMNS, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = _HEADER_FONT
            cell.fill = _DARK_BLUE
            cell.alignment = _CENTER
            cell.border = _THIN_BORDER

        # Data rows
        for row_idx, row_data in enumerate(rows, 2):
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.font = _DATA_FONT
                cell.border = _THIN_BORDER
                # Center-align: Serial No, relation_type, gender, age,
                # part_number, ac_number, district_code
                if col_idx in (1, 6, 8, 9, 13, 14, 15):
                    cell.alignment = _CENTER
                else:
                    cell.alignment = _LEFT

            # Alternating row colors
            if row_idx % 2 == 0:
                for col_idx in range(1, len(MERGED_COLUMNS) + 1):
                    ws.cell(row=row_idx, column=col_idx).fill = _LIGHT_FILL

        # Column widths
        for col_idx, width in enumerate(_COL_WIDTHS, 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = width

        # Row heights
        ws.row_dimensions[1].height = 30
        for r in range(2, len(rows) + 2):
            ws.row_dimensions[r].height = 20

        # Freeze header row
        ws.freeze_panes = ws.cell(row=2, column=1)

        wb.save(output_path)
        logger.info(
            "Merged Excel saved: %s (%d rows, %d columns)",
            output_path, len(rows), len(MERGED_COLUMNS),
        )
