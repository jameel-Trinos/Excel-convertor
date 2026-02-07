"""Simplified Excel file creation with basic formatting."""

import logging
from datetime import datetime
from typing import List, Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from .models import TableData
from .utils import sanitize_text

logger = logging.getLogger(__name__)


class ExcelCreator:
    """
    Create Excel files from extracted table data with simple, professional formatting.

    Features:
    - Simple header formatting (dark blue with white text)
    - Borders on all cells
    - Center alignment
    - Auto-fit column widths
    - SUM formulas for numeric columns
    """

    def __init__(self):
        """Initialize the Excel creator."""
        pass

    def create_from_tables(
        self,
        tables: List[TableData],
        output_path: str,
        source_filename: str = "Untitled",
        page_texts: Optional[List[str]] = None,
        ac_number: Optional[str] = None,
    ) -> str:
        """
        Create an Excel file from extracted table data.

        Args:
            tables: List of TableData objects (typically one merged table)
            output_path: Path where Excel file will be saved
            source_filename: Original PDF filename
            page_texts: Not used in simplified version
            ac_number: Optional Assembly Constituency number to add as first column

        Returns:
            Path to the created Excel file
        """
        logger.info(f"Creating Excel file with {len(tables)} tables")

        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "Data"

        if not tables or all(t.is_empty for t in tables):
            worksheet.cell(row=1, column=1).value = "No data extracted"
            workbook.save(output_path)
            return output_path

        # Use the first (and typically only) table
        table = tables[0]
        headers = table.headers
        rows = table.rows

        if not headers:
            worksheet.cell(row=1, column=1).value = "No headers found"
            workbook.save(output_path)
            return output_path

        # Add AC NO as first column (always add, even if ac_number is None)
        headers = ["AC NO."] + headers
        if ac_number:
            logger.info(f"Added AC NO column with value: {ac_number}")
        else:
            logger.info("Added AC NO column (AC number not found in PDF)")

        current_row = 1

        # Fix headers before writing to Excel
        from .header_fixer import HeaderFixer
        headers = HeaderFixer.fix_header_list(headers)
        
        # Write headers
        header_row = current_row
        for col_idx, header in enumerate(headers, 1):
            cell = worksheet.cell(row=header_row, column=col_idx)
            # Headers are already fixed, just sanitize for RTL characters
            cell.value = sanitize_text(header) if header else f"Column {col_idx}"

        # Format header row
        self._format_header_row(worksheet, header_row, len(headers))
        current_row += 1

        # Write data rows with final validation pass
        data_start_row = current_row
        from .utils import _is_likely_reversed
        corrected_cells = 0
        
        for row_idx, row_data in enumerate(rows):
            # Write AC NO in first column (always present, may be empty if not found)
            cell = worksheet.cell(row=data_start_row + row_idx, column=1)
            cell.value = ac_number if ac_number else ""
            
            # Write remaining data columns (shifted by 1 since AC NO column is always first)
            start_col = 2
            for col_idx, value in enumerate(row_data, 0):
                excel_col = start_col + col_idx
                if excel_col <= len(headers):  # Ensure we don't exceed header count
                    cell = worksheet.cell(row=data_start_row + row_idx, column=excel_col)
                    
                    # Final validation: Double-check for reversed text before writing
                    if isinstance(value, str):
                        sanitized = sanitize_text(value)
                        # Re-check for reversed text after sanitization
                        if _is_likely_reversed(sanitized):
                            # Try reversing and re-sanitizing
                            reversed_text = sanitized[::-1]
                            re_sanitized = sanitize_text(reversed_text)
                            if not _is_likely_reversed(re_sanitized):
                                cell.value = re_sanitized
                                corrected_cells += 1
                                logger.debug(
                                    f"Final validation fix: Row {data_start_row + row_idx}, Col {excel_col}: "
                                    f"'{value[:50]}...' -> '{re_sanitized[:50]}...'"
                                )
                            else:
                                # Keep original sanitized if reversed is still reversed
                                cell.value = sanitized
                        else:
                            cell.value = sanitized
                    else:
                        cell.value = value
        
        if corrected_cells > 0:
            logger.info(f"Final validation: Corrected {corrected_cells} cells with reversed/corrupted text before writing to Excel")

        data_end_row = data_start_row + len(rows) - 1

        # Format data cells
        self._format_data_cells(
            worksheet,
            start_row=data_start_row,
            end_row=data_end_row,
            start_col=1,
            end_col=len(headers),
        )

        # Add TOTAL row with SUM formulas
        if rows:
            total_row = data_end_row + 1
            self._add_total_row(
                worksheet,
                total_row=total_row,
                header_row=header_row,
                data_start_row=data_start_row,
                data_end_row=data_end_row,
                num_columns=len(headers),
                has_ac_column=True,  # AC NO column is always present as first column
            )

        # Auto-fit column widths
        self._auto_fit_columns(worksheet, len(headers))

        # Freeze header row
        worksheet.freeze_panes = worksheet.cell(row=header_row + 1, column=1)

        # Save workbook
        workbook.save(output_path)
        logger.info(f"Excel file created: {output_path}")

        return output_path

    def _format_header_row(self, worksheet, row_number: int, num_columns: int):
        """Format the header row with dark blue background and white text."""
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=10)
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        for col in range(1, num_columns + 1):
            cell = worksheet.cell(row=row_number, column=col)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
            cell.border = thin_border

    def _format_data_cells(
        self, worksheet, start_row: int, end_row: int, start_col: int, end_col: int
    ):
        """Format data cells with borders and center alignment."""
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )
        center_alignment = Alignment(horizontal="center", vertical="center")

        for row in range(start_row, end_row + 1):
            for col in range(start_col, end_col + 1):
                cell = worksheet.cell(row=row, column=col)
                cell.border = thin_border
                cell.alignment = center_alignment

    def _add_total_row(
        self,
        worksheet,
        total_row: int,
        header_row: int,
        data_start_row: int,
        data_end_row: int,
        num_columns: int,
        has_ac_column: bool = False,
    ):
        """
        Add a TOTAL row with SUM formulas for numeric columns.
        
        Args:
            worksheet: The worksheet to modify
            total_row: Row number for the total row
            header_row: Row number of the header row
            data_start_row: First row of data
            data_end_row: Last row of data
            num_columns: Total number of columns
            has_ac_column: Whether AC NO column exists (should be skipped in totals)
        """
        # First column gets "TOTAL" label
        worksheet.cell(row=total_row, column=1).value = "TOTAL"

        # Start from column 2 (skip AC NO column if it exists, or skip TOTAL label column)
        start_col = 2 if has_ac_column else 2
        
        # Check each column for numeric content and add SUM formulas
        # Skip the first column (AC NO or TOTAL label)
        for col in range(start_col, num_columns + 1):
            if self._is_numeric_column(worksheet, col, data_start_row, data_end_row):
                col_letter = get_column_letter(col)
                formula = f"=SUM({col_letter}{data_start_row}:{col_letter}{data_end_row})"
                cell = worksheet.cell(row=total_row, column=col)
                cell.value = formula
                cell.number_format = "#,##0"

        # Format total row
        self._format_data_cells(worksheet, total_row, total_row, 1, num_columns)
        # Make TOTAL label bold
        worksheet.cell(row=total_row, column=1).font = Font(bold=True)

    def _is_numeric_column(
        self, worksheet, column: int, start_row: int, end_row: int, threshold: float = 0.5
    ) -> bool:
        """
        Check if a column contains predominantly numeric values.

        Args:
            worksheet: The worksheet to check
            column: Column number (1-indexed)
            start_row: First row to check
            end_row: Last row to check
            threshold: Minimum ratio of numeric values (0.0 to 1.0)

        Returns:
            True if the column is predominantly numeric
        """
        numeric_count = 0
        total_count = 0

        for row in range(start_row, end_row + 1):
            value = worksheet.cell(row=row, column=column).value
            if value is not None and str(value).strip():
                total_count += 1
                if self._is_numeric_value(value):
                    numeric_count += 1

        if total_count == 0:
            return False

        return (numeric_count / total_count) >= threshold

    def _is_numeric_value(self, value) -> bool:
        """Check if a value is numeric."""
        if isinstance(value, (int, float)):
            return True

        if isinstance(value, str):
            # Remove commas and check if it's a number
            cleaned = value.replace(",", "").replace(" ", "").strip()
            try:
                float(cleaned)
                return True
            except (ValueError, TypeError):
                return False

        return False

    def _auto_fit_columns(self, worksheet, num_columns: int):
        """Auto-fit column widths based on content."""
        for col in range(1, num_columns + 1):
            column_letter = get_column_letter(col)
            max_length = 0

            # Check header
            header_value = worksheet.cell(row=1, column=col).value
            if header_value:
                max_length = max(max_length, len(str(header_value)))

            # Check data rows (sample first 20 rows for performance)
            for row in range(2, min(22, worksheet.max_row + 1)):
                cell_value = worksheet.cell(row=row, column=col).value
                if cell_value:
                    max_length = max(max_length, len(str(cell_value)))

            # Set width with some padding (min 10, max 50)
            adjusted_width = min(max(max_length + 2, 10), 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width
