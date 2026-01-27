"""Deterministic Excel creator that preserves exact schema from PDF."""

import logging
import re
from datetime import datetime
from typing import List, Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from .utils import sanitize_text

logger = logging.getLogger(__name__)


class DeterministicExcelCreator:
    """
    Create Excel files with exact schema preservation.

    Rules:
    - Preserves exact column names from PDF
    - No AI processing or column merging
    - Professional formatting only
    - Adds validation formulas
    """

    # Color constants
    HEADER_BG_COLOR = "1F4E79"  # Dark blue for headers
    HEADER_FONT_COLOR = "FFFFFF"  # White text
    ALT_ROW_COLOR = "D6DCE5"  # Light blue/gray for alternating rows

    def __init__(self):
        """Initialize creator."""
        pass

    def create_excel(
        self,
        headers: List[str],
        data_rows: List[List[any]],
        output_path: str,
        title: str = "Election Results",
        source_filename: str = "Untitled",
        title_lines: Optional[List[str]] = None,
    ) -> str:
        """
        Create Excel file from headers and data rows.

        Args:
            headers: Column headers (exact from PDF)
            data_rows: Data rows
            output_path: Output file path
            title: Document title
            source_filename: Original PDF filename
            title_lines: Optional list of title lines from PDF (e.g., FORM 20, Election type, Constituency, Electors)

        Returns:
            Path to created Excel file
        """
        logger.info(f"Creating Excel with {len(headers)} columns, {len(data_rows)} rows")

        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "Election Results"

        current_row = 1

        # Add title section - use extracted title lines if available
        if title_lines and len(title_lines) > 0:
            current_row = self._add_multi_row_title_section(
                worksheet,
                title_lines=title_lines,
                num_columns=len(headers),
                start_row=current_row,
            )
        else:
            self._add_title_section(
                worksheet,
                title=title,
                subtitle=f"Source: {source_filename} | Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
                num_columns=len(headers),
                start_row=current_row,
            )
            current_row += 3  # Title takes 3 rows

        # Add headers
        header_row = current_row
        for col_idx, header in enumerate(headers, 1):
            cell = worksheet.cell(row=header_row, column=col_idx)
            # Sanitize header to remove RTL/bidirectional control characters
            cell.value = sanitize_text(header) if header else f"Column {col_idx}"

        # Format header row
        self._format_header_row(worksheet, header_row, len(headers))
        current_row += 1

        # Add data rows
        data_start_row = current_row
        for row_idx, row_data in enumerate(data_rows):
            for col_idx, value in enumerate(row_data, 1):
                cell = worksheet.cell(row=data_start_row + row_idx, column=col_idx)
                # Sanitize and convert to appropriate type
                if isinstance(value, str):
                    sanitized = sanitize_text(value)
                    if sanitized.isdigit():
                        cell.value = int(sanitized)
                    else:
                        cell.value = sanitized
                else:
                    cell.value = value

        data_end_row = data_start_row + len(data_rows) - 1

        # Format data cells
        self._format_data_cells(
            worksheet,
            data_start_row,
            data_end_row,
            len(headers),
        )

        # Apply number formatting
        self._apply_number_formatting(
            worksheet,
            data_start_row,
            data_end_row,
            len(headers),
        )

        # Add total row
        total_row = data_end_row + 1
        self._add_total_row(
            worksheet,
            header_row,
            data_start_row,
            data_end_row,
            total_row,
            len(headers),
            headers,
        )

        # Set column widths
        self._set_column_widths(worksheet, len(headers))

        # Set row heights
        self._set_row_heights(worksheet, header_row, data_start_row, data_end_row)

        # Freeze header row
        worksheet.freeze_panes = worksheet.cell(row=header_row + 1, column=1)

        # Save workbook
        workbook.save(output_path)
        logger.info(f"Excel file created: {output_path}")

        return output_path

    def _add_multi_row_title_section(
        self,
        worksheet,
        title_lines: List[str],
        num_columns: int,
        start_row: int,
    ) -> int:
        """
        Add multi-row title section matching the PDF format.

        Args:
            worksheet: The worksheet to modify
            title_lines: List of title lines (e.g., FORM 20, Election type, Constituency, Electors)
            num_columns: Number of columns to merge across
            start_row: Starting row

        Returns:
            Next available row after title section
        """
        current_row = start_row

        for i, line in enumerate(title_lines):
            if not line or not line.strip():
                continue

            cell = worksheet.cell(row=current_row, column=1)
            cell.value = sanitize_text(line.strip())

            # First line is main title - bigger and bolder
            if i == 0:
                cell.font = Font(size=14, bold=True, color="000000")
            else:
                cell.font = Font(size=11, bold=False, color="000000")

            cell.alignment = Alignment(horizontal="center", vertical="center")

            # Merge cells across all columns
            worksheet.merge_cells(
                start_row=current_row,
                start_column=1,
                end_row=current_row,
                end_column=num_columns,
            )

            worksheet.row_dimensions[current_row].height = 20
            current_row += 1

        # Add empty row for spacing
        worksheet.row_dimensions[current_row].height = 10
        current_row += 1

        return current_row

    def _add_title_section(
        self,
        worksheet,
        title: str,
        subtitle: str,
        num_columns: int,
        start_row: int,
    ):
        """Add title section to worksheet."""
        # Main title
        title_cell = worksheet.cell(row=start_row, column=1)
        title_cell.value = title
        title_cell.font = Font(size=16, bold=True, color="1F4E78")
        title_cell.alignment = Alignment(horizontal="center", vertical="center")
        worksheet.merge_cells(
            start_row=start_row,
            start_column=1,
            end_row=start_row,
            end_column=num_columns,
        )

        # Subtitle
        subtitle_cell = worksheet.cell(row=start_row + 1, column=1)
        subtitle_cell.value = subtitle
        subtitle_cell.font = Font(size=10, italic=True, color="7F7F7F")
        subtitle_cell.alignment = Alignment(horizontal="center", vertical="center")
        worksheet.merge_cells(
            start_row=start_row + 1,
            start_column=1,
            end_row=start_row + 1,
            end_column=num_columns,
        )

        # Empty row
        worksheet.row_dimensions[start_row + 2].height = 10

    def _format_header_row(self, worksheet, row_number: int, num_columns: int):
        """Format header row with dark blue background and white text."""
        header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True,
        )
        border = Border(
            left=Side(style="thin", color="000000"),
            right=Side(style="thin", color="000000"),
            top=Side(style="thin", color="000000"),
            bottom=Side(style="thin", color="000000"),
        )

        for col in range(1, num_columns + 1):
            cell = worksheet.cell(row=row_number, column=col)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
            cell.border = border

        # Set header row height
        worksheet.row_dimensions[row_number].height = 70

    def _format_data_cells(
        self,
        worksheet,
        start_row: int,
        end_row: int,
        num_columns: int,
    ):
        """Format data cells with borders, center alignment, and alternating row colors."""
        border = Border(
            left=Side(style="thin", color="D3D3D3"),
            right=Side(style="thin", color="D3D3D3"),
            top=Side(style="thin", color="D3D3D3"),
            bottom=Side(style="thin", color="D3D3D3"),
        )
        alignment = Alignment(horizontal="center", vertical="center")

        # Alternating row fill colors
        alt_fill = PatternFill(
            start_color=self.ALT_ROW_COLOR,
            end_color=self.ALT_ROW_COLOR,
            fill_type="solid",
        )

        for row in range(start_row, end_row + 1):
            # Calculate row index relative to start for alternating colors
            row_index = row - start_row
            use_alt_color = (row_index % 2) == 1  # Odd rows get alternating color

            for col in range(1, num_columns + 1):
                cell = worksheet.cell(row=row, column=col)
                cell.border = border
                cell.alignment = alignment

                # Apply alternating row color
                if use_alt_color:
                    cell.fill = alt_fill

    def _apply_number_formatting(
        self,
        worksheet,
        start_row: int,
        end_row: int,
        num_columns: int,
    ):
        """Apply number formatting with thousand separators."""
        for row in range(start_row, end_row + 1):
            for col in range(1, num_columns + 1):
                cell = worksheet.cell(row=row, column=col)
                if isinstance(cell.value, (int, float)):
                    cell.number_format = "#,##0"

    def _add_total_row(
        self,
        worksheet,
        header_row: int,
        data_start_row: int,
        data_end_row: int,
        total_row: int,
        num_columns: int,
        headers: List[str],
    ):
        """Add TOTAL row with SUM formulas."""
        # First column gets "TOTAL" label
        total_label_cell = worksheet.cell(row=total_row, column=1)
        total_label_cell.value = "TOTAL"
        total_label_cell.font = Font(bold=True, size=11)
        total_label_cell.fill = PatternFill(
            start_color="E7E6E6", end_color="E7E6E6", fill_type="solid"
        )

        # Add formulas for numeric columns
        for col in range(2, num_columns + 1):
            if self._is_numeric_column(worksheet, col, data_start_row, data_end_row):
                col_letter = get_column_letter(col)
                formula = f"=SUM({col_letter}{data_start_row}:{col_letter}{data_end_row})"
                cell = worksheet.cell(row=total_row, column=col)
                cell.value = formula
                cell.number_format = "#,##0"
                cell.font = Font(bold=True, size=11)
                cell.fill = PatternFill(
                    start_color="E7E6E6", end_color="E7E6E6", fill_type="solid"
                )

        # Format total row
        border = Border(
            left=Side(style="medium", color="000000"),
            right=Side(style="medium", color="000000"),
            top=Side(style="medium", color="000000"),
            bottom=Side(style="medium", color="000000"),
        )
        for col in range(1, num_columns + 1):
            cell = worksheet.cell(row=total_row, column=col)
            cell.border = border
            cell.alignment = Alignment(horizontal="center", vertical="center")

    def _is_numeric_column(
        self,
        worksheet,
        column: int,
        start_row: int,
        end_row: int,
    ) -> bool:
        """Check if column contains predominantly numeric values."""
        numeric_count = 0
        total_count = 0

        for row in range(start_row, min(start_row + 20, end_row + 1)):
            value = worksheet.cell(row=row, column=column).value
            if value is not None:
                total_count += 1
                if isinstance(value, (int, float)):
                    numeric_count += 1

        if total_count == 0:
            return False

        return (numeric_count / total_count) >= 0.5

    def _set_column_widths(self, worksheet, num_columns: int):
        """Set fixed column widths."""
        for col in range(1, num_columns + 1):
            col_letter = get_column_letter(col)
            # First few columns wider for names/stations
            if col <= 3:
                worksheet.column_dimensions[col_letter].width = 20
            else:
                worksheet.column_dimensions[col_letter].width = 16

    def _set_row_heights(
        self,
        worksheet,
        header_row: int,
        data_start_row: int,
        data_end_row: int,
    ):
        """Set row heights."""
        # Header row already set in _format_header_row
        # Set data rows
        for row in range(data_start_row, data_end_row + 1):
            worksheet.row_dimensions[row].height = 18
