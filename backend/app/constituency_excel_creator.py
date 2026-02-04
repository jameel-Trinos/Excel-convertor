"""Specialized Excel creator for constituency data with perfect cell alignment validation."""

import logging
from typing import List, Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from .models import TableData
from .utils import sanitize_text

logger = logging.getLogger(__name__)


class ConstituencyExcelCreator:
    """
    Create Excel files from constituency table data with perfect cell alignment.
    
    Features:
    - Master headers placed in first row
    - Strict validation that data rows match header count exactly
    - Perfect cell-to-column mapping
    - Professional formatting
    - Auto-fit column widths
    """

    def __init__(self):
        """Initialize the constituency Excel creator."""
        pass

    def create_from_tables(
        self,
        tables: List[TableData],
        output_path: str,
        source_filename: str = "Untitled",
        page_texts: List[str] = None,
    ) -> str:
        """
        Create an Excel file from constituency table data with perfect alignment.
        
        Args:
            tables: List of TableData objects (typically one merged table with master headers)
            output_path: Path where Excel file will be saved
            source_filename: Original PDF filename
            page_texts: Not used
            
        Returns:
            Path to the created Excel file
        """
        logger.info(f"Creating constituency Excel file with {len(tables)} tables")

        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "Constituency Data"

        if not tables or all(t.is_empty for t in tables):
            worksheet.cell(row=1, column=1).value = "No data extracted"
            workbook.save(output_path)
            return output_path

        # Use the first (and typically only) table
        table = tables[0]
        headers = table.headers
        rows = table.rows

        if not headers:
            worksheet.cell(row=1, column=1).value = "No headers found"
            workbook.save(output_path)
            return output_path

        current_row = 1

        # Add title rows with merged cells and colored backgrounds
        title_row = current_row
        worksheet.cell(row=title_row, column=1).value = "Constituency Data"
        worksheet.merge_cells(start_row=title_row, start_column=1, end_row=title_row, end_column=len(headers))
        self._format_title_row(worksheet, title_row, len(headers))
        current_row += 1

        # Write master headers in header row
        header_row = current_row
        for col_idx, header in enumerate(headers, 1):
            cell = worksheet.cell(row=header_row, column=col_idx)
            cell.value = sanitize_text(header) if header else f"Column {col_idx}"

        # Format header row
        self._format_header_row(worksheet, header_row, len(headers))
        current_row += 1

        # Write data rows with strict validation
        data_start_row = current_row
        validated_rows = 0
        invalid_rows = 0
        
        for row_idx, row_data in enumerate(rows):
            # Validate row length matches header count
            validated_row = self._validate_and_align_row(row_data, len(headers), row_idx + 1)
            
            if validated_row is None:
                invalid_rows += 1
                logger.warning(
                    f"Row {row_idx + 1}: Invalid alignment - expected {len(headers)} columns, "
                    f"got {len(row_data)}"
                )
                continue
            
            # Write validated row
            for col_idx, value in enumerate(validated_row, 1):
                if col_idx <= len(headers):  # Safety check
                    cell = worksheet.cell(row=data_start_row + validated_rows, column=col_idx)
                    cell.value = sanitize_text(value) if isinstance(value, str) else value
            
            validated_rows += 1

        data_end_row = data_start_row + validated_rows - 1

        if invalid_rows > 0:
            logger.warning(f"Skipped {invalid_rows} rows due to alignment issues")

        if validated_rows == 0:
            worksheet.cell(row=data_start_row, column=1).value = "No valid data rows found"
            workbook.save(output_path)
            return output_path

        # Format data cells with alternating row colors
        self._format_data_cells(
            worksheet,
            start_row=data_start_row,
            end_row=data_end_row,
            start_col=1,
            end_col=len(headers),
        )

        # Add TOTAL row with SUM formulas
        if validated_rows > 0:
            total_row = data_end_row + 1
            self._add_total_row(
                worksheet,
                total_row=total_row,
                header_row=header_row,
                data_start_row=data_start_row,
                data_end_row=data_end_row,
                num_columns=len(headers),
            )

        # Auto-fit column widths
        self._auto_fit_columns(worksheet, len(headers))

        # Freeze header rows (title + header)
        worksheet.freeze_panes = worksheet.cell(row=header_row + 1, column=1)

        # Save workbook
        workbook.save(output_path)
        logger.info(
            f"Constituency Excel file created: {output_path} "
            f"({validated_rows} rows, {len(headers)} columns)"
        )

        return output_path

    def _validate_and_align_row(
        self, row_data: List, expected_col_count: int, row_number: int
    ) -> Optional[List[str]]:
        """
        Validate and align a data row to match expected column count.
        
        Args:
            row_data: Original row data
            expected_col_count: Expected number of columns (from headers)
            row_number: Row number for logging
            
        Returns:
            Aligned row with exact column count, or None if validation fails
        """
        # Convert all values to strings and clean
        cleaned_row = []
        for value in row_data:
            if value is None:
                cleaned_row.append("")
            elif isinstance(value, str):
                cleaned_row.append(value.strip())
            else:
                cleaned_row.append(str(value).strip())

        # Align row length to match headers
        if len(cleaned_row) < expected_col_count:
            # Pad with empty strings
            cleaned_row.extend([""] * (expected_col_count - len(cleaned_row)))
        elif len(cleaned_row) > expected_col_count:
            # Truncate to match
            cleaned_row = cleaned_row[:expected_col_count]

        # Final validation
        if len(cleaned_row) != expected_col_count:
            return None

        return cleaned_row

    def _format_header_row(self, worksheet, row_number: int, num_columns: int):
        """Format the header row with dark blue background and white text."""
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=10)
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        for col in range(1, num_columns + 1):
            cell = worksheet.cell(row=row_number, column=col)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
            cell.border = thin_border

    def _format_title_row(self, worksheet, row_number: int, num_columns: int):
        """Format the title row with merged cells and colored background."""
        # Light blue background
        title_fill = PatternFill(start_color="D6E3F0", end_color="D6E3F0", fill_type="solid")
        title_font = Font(bold=True, size=14, color="000000")
        title_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        cell = worksheet.cell(row=row_number, column=1)
        cell.fill = title_fill
        cell.font = title_font
        cell.alignment = title_alignment
        cell.border = thin_border

    def _format_data_cells(
        self, worksheet, start_row: int, end_row: int, start_col: int, end_col: int
    ):
        """Format data cells with borders, center alignment, and alternating row colors."""
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )
        # Center alignment with text wrapping
        center_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        # Alternating row colors
        light_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

        for row in range(start_row, end_row + 1):
            # Alternate row colors
            row_fill = light_fill if (row - start_row) % 2 == 1 else white_fill
            
            for col in range(start_col, end_col + 1):
                cell = worksheet.cell(row=row, column=col)
                cell.border = thin_border
                cell.alignment = center_alignment
                cell.fill = row_fill

    def _add_total_row(
        self,
        worksheet,
        total_row: int,
        header_row: int,
        data_start_row: int,
        data_end_row: int,
        num_columns: int,
    ):
        """Add a TOTAL row with SUM formulas for numeric columns."""
        # First column gets "TOTAL" label
        worksheet.cell(row=total_row, column=1).value = "TOTAL"

        # Check each column for numeric content and add SUM formulas
        for col in range(2, num_columns + 1):
            if self._is_numeric_column(worksheet, col, data_start_row, data_end_row):
                col_letter = get_column_letter(col)
                formula = f"=SUM({col_letter}{data_start_row}:{col_letter}{data_end_row})"
                cell = worksheet.cell(row=total_row, column=col)
                cell.value = formula
                cell.number_format = "#,##0"

        # Format total row
        self._format_data_cells(worksheet, total_row, total_row, 1, num_columns)
        # Make TOTAL label bold
        worksheet.cell(row=total_row, column=1).font = Font(bold=True)

    def _is_numeric_column(
        self, worksheet, column: int, start_row: int, end_row: int, threshold: float = 0.5
    ) -> bool:
        """
        Check if a column contains predominantly numeric values.
        
        Args:
            worksheet: The worksheet to check
            column: Column number (1-indexed)
            start_row: First row to check
            end_row: Last row to check
            threshold: Minimum ratio of numeric values (0.0 to 1.0)
            
        Returns:
            True if the column is predominantly numeric
        """
        numeric_count = 0
        total_count = 0

        for row in range(start_row, end_row + 1):
            value = worksheet.cell(row=row, column=column).value
            if value is not None and str(value).strip():
                total_count += 1
                if self._is_numeric_value(value):
                    numeric_count += 1

        if total_count == 0:
            return False

        return (numeric_count / total_count) >= threshold

    def _is_numeric_value(self, value) -> bool:
        """Check if a value is numeric."""
        if isinstance(value, (int, float)):
            return True

        if isinstance(value, str):
            # Remove commas and check if it's a number
            cleaned = value.replace(",", "").replace(" ", "").strip()
            try:
                float(cleaned)
                return True
            except (ValueError, TypeError):
                return False

        return False

    def _auto_fit_columns(self, worksheet, num_columns: int):
        """Auto-fit column widths based on content."""
        for col in range(1, num_columns + 1):
            column_letter = get_column_letter(col)
            max_length = 0

            # Check header
            header_value = worksheet.cell(row=1, column=col).value
            if header_value:
                max_length = max(max_length, len(str(header_value)))

            # Check data rows (sample first 20 rows for performance)
            for row in range(2, min(22, worksheet.max_row + 1)):
                cell_value = worksheet.cell(row=row, column=col).value
                if cell_value:
                    max_length = max(max_length, len(str(cell_value)))

            # Set width with some padding (min 10, max 50)
            adjusted_width = min(max(max_length + 2, 10), 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width

