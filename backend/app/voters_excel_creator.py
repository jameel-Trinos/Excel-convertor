"""Excel creator for voter list data.

Generates a professionally formatted Excel workbook with:
- Header section: AC No, Booth No, Address, Total Voters
- Bilingual column headers (English + Tamil)
- Professional formatting matching the project's Excel standards
"""

import logging
from typing import Any, Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)

# Bilingual column headers: (English, Tamil)
VOTER_COLUMNS = [
    ("Serial No", "வ.எண்"),
    ("Name", "பெயர்"),
    ("Father/Husband Name", "தந்தை/கணவர் பெயர்"),
    ("House No", "வீட்டு எண்"),
    ("Age", "வயது"),
    ("Gender", "பாலினம்"),
    ("Voter ID", "வாக்காளர் அடையாள எண்"),
    ("Street Name", "தெரு பெயர்"),
]

# Styles
_DARK_BLUE = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
_LIGHT_BLUE = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
_WHITE_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
_HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
_TAMIL_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=10)
_META_LABEL_FONT = Font(name="Calibri", bold=True, size=11, color="1F4E79")
_META_VALUE_FONT = Font(name="Calibri", size=11)
_DATA_FONT = Font(name="Calibri", size=10)
_THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)
_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
_LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)

# Column widths
_COL_WIDTHS = [12, 25, 25, 15, 8, 12, 22, 35]


class VotersExcelCreator:
    """Create professionally formatted voter list Excel files."""

    def create(
        self,
        headers: list[str],
        rows: list[list[Any]],
        output_path: str,
        ac_no: str = "",
        part_no: str = "",
        address: str = "",
        total_voters: str = "",
        source_filename: str = "",
        expected_total: str = "",
        extracted_total: str = "",
    ) -> str:
        """
        Create voter list Excel file.

        Args:
            headers: Column header names
            rows: Voter data rows
            output_path: Where to save the .xlsx file
            ac_no: Assembly Constituency number
            part_no: Part/Booth number
            address: Booth address
            total_voters: Total voter count
            source_filename: Original PDF filename
            expected_total: Expected total from PDF header (மொத்தம்)
            extracted_total: Actual extracted count

        Returns:
            Path to created file.
        """
        wb = Workbook()
        ws = wb.active
        ws.title = "Voter List"

        current_row = 1

        # --- Header metadata section ---
        current_row = self._write_metadata(
            ws, current_row, ac_no, part_no, address, total_voters, source_filename,
            expected_total=expected_total, extracted_total=extracted_total,
        )

        # Blank row separator
        current_row += 1

        # --- Bilingual column headers ---
        # Row 1: English headers
        eng_row = current_row
        for col_idx, (eng, _tam) in enumerate(VOTER_COLUMNS, 1):
            cell = ws.cell(row=eng_row, column=col_idx, value=eng)
            cell.font = _HEADER_FONT
            cell.fill = _DARK_BLUE
            cell.alignment = _CENTER
            cell.border = _THIN_BORDER

        current_row += 1

        # Row 2: Tamil headers
        tam_row = current_row
        for col_idx, (_eng, tam) in enumerate(VOTER_COLUMNS, 1):
            cell = ws.cell(row=tam_row, column=col_idx, value=tam)
            cell.font = _TAMIL_FONT
            cell.fill = _DARK_BLUE
            cell.alignment = _CENTER
            cell.border = _THIN_BORDER

        current_row += 1

        # --- Data rows ---
        data_start = current_row
        for row_data in rows:
            for col_idx in range(len(VOTER_COLUMNS)):
                value = row_data[col_idx] if col_idx < len(row_data) else ""
                cell = ws.cell(row=current_row, column=col_idx + 1, value=value)
                cell.font = _DATA_FONT
                cell.border = _THIN_BORDER
                # Center align Serial No, Age, Gender, Voter ID
                if col_idx in (0, 4, 5, 6):
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                else:
                    cell.alignment = _LEFT
            current_row += 1

        # --- Formatting ---
        # Column widths
        for col_idx, width in enumerate(_COL_WIDTHS, 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = width

        # Row heights — tall enough to show bilingual headers clearly
        ws.row_dimensions[eng_row].height = 40
        ws.row_dimensions[tam_row].height = 40
        for r in range(data_start, current_row):
            ws.row_dimensions[r].height = 20

        # Freeze panes below the Tamil header row
        ws.freeze_panes = ws.cell(row=tam_row + 1, column=1)

        # Alternating row colors for readability
        light_fill = PatternFill(start_color="F2F7FB", end_color="F2F7FB", fill_type="solid")
        for r in range(data_start, current_row):
            if (r - data_start) % 2 == 1:
                for c in range(1, len(VOTER_COLUMNS) + 1):
                    ws.cell(row=r, column=c).fill = light_fill

        wb.save(output_path)
        logger.info(f"Voter Excel saved: {output_path} ({len(rows)} rows)")
        return output_path

    def _write_metadata(
        self,
        ws,
        start_row: int,
        ac_no: str,
        part_no: str,
        address: str,
        total_voters: str,
        source_filename: str,
        expected_total: str = "",
        extracted_total: str = "",
    ) -> int:
        """Write metadata header section. Returns next row number."""
        row = start_row

        # Title row (merged)
        ws.merge_cells(
            start_row=row, start_column=1,
            end_row=row, end_column=len(VOTER_COLUMNS),
        )
        title_cell = ws.cell(row=row, column=1, value="Voter List / வாக்காளர் பட்டியல்")
        title_cell.font = Font(name="Calibri", bold=True, size=14, color="1F4E79")
        title_cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[row].height = 30
        row += 1

        # Metadata rows: label in columns A-C (merged), value in columns D-G (merged)
        # Build total voters display with reconciliation info
        total_display = total_voters or "—"
        if expected_total and extracted_total and expected_total != extracted_total:
            total_display = f"{extracted_total} (Expected: {expected_total})"
        elif expected_total and extracted_total and expected_total == extracted_total:
            total_display = f"{extracted_total} (Matched)"

        meta_items = [
            ("AC No / சட்டமன்றத் தொகுதி எண்:", ac_no or "—"),
            ("Booth No / பகுதி எண்:", part_no or "—"),
            ("Address / முகவரி:", address or "—"),
            ("Total Voters / மொத்த வாக்காளர்கள்:", total_display),
        ]

        for label, value in meta_items:
            # Merge label across columns 1-3 for enough width
            ws.merge_cells(
                start_row=row, start_column=1,
                end_row=row, end_column=3,
            )
            label_cell = ws.cell(row=row, column=1, value=label)
            label_cell.font = _META_LABEL_FONT
            label_cell.alignment = Alignment(
                horizontal="right", vertical="center", wrap_text=True
            )

            # Merge value across remaining columns (4-7)
            ws.merge_cells(
                start_row=row, start_column=4,
                end_row=row, end_column=len(VOTER_COLUMNS),
            )
            value_cell = ws.cell(row=row, column=4, value=value)
            value_cell.font = _META_VALUE_FONT
            value_cell.alignment = Alignment(horizontal="left", vertical="center")
            ws.row_dimensions[row].height = 25
            row += 1

        return row
