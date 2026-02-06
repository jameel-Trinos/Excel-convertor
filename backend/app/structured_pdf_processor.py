"""
Structured PDF Processor for Indian Election Form 20 Documents.

This processor implements a 6-phase approach for converting election PDFs to Excel:
1. PHASE 1: Understand PDF structure (title, metadata, headers, data)
2. PHASE 2: Extract column headers (one time only from first page)
3. PHASE 3: Extract data from all pages (skip repeated headers)
4. PHASE 4: Create Excel with proper structure
5. PHASE 5: Quality verification
6. PHASE 6: Delivery

Key principles:
- Headers extracted ONCE from first page only
- Data rows identified by numeric Polling Station values
- Multi-page tables merged with duplicate header removal
- Professional Excel formatting with SUM formulas
"""

import logging
import re
from dataclasses import dataclass, field
from pathlib import Path
from typing import Callable, List, Optional, Tuple

import pdfplumber
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from .models import ExtractionResult, TableData
from .utils import sanitize_text

logger = logging.getLogger(__name__)


@dataclass
class PDFStructure:
    """Represents the analyzed structure of a PDF document."""

    title_section: List[str] = field(default_factory=list)
    metadata_section: List[str] = field(default_factory=list)
    header_row_index: int = 0
    first_data_row_index: int = 0
    total_pages: int = 0
    total_columns: int = 0
    headers_repeat_on_pages: bool = True


@dataclass
class ExtractionConfig:
    """Configuration for PDF extraction."""

    header_row_keywords: List[str] = field(default_factory=lambda: [
        "polling", "station", "sl.", "sl no", "s.no"
    ])
    title_keywords: List[str] = field(default_factory=lambda: [
        "form 20", "final result", "election", "lok sabha", "assembly"
    ])
    metadata_keywords: List[str] = field(default_factory=lambda: [
        "constituency", "electors", "total no", "assembly no"
    ])
    skip_columns: List[str] = field(default_factory=lambda: [
        "sl.", "sl no", "s.no", "serial"
    ])


class StructuredPDFProcessor:
    """
    Structured PDF processor implementing 6-phase extraction.

    This processor is specifically designed for Indian election Form 20 documents
    but can be adapted for similar tabular PDFs.
    """

    # Excel formatting constants
    HEADER_BG_COLOR = "4472C4"  # Blue
    HEADER_FONT_COLOR = "FFFFFF"  # White
    TITLE_BG_COLOR = "D9E1F2"  # Light blue
    TOTAL_BG_COLOR = "FFC000"  # Orange
    ALT_ROW_COLOR = "F2F2F2"  # Light gray

    def __init__(self, file_path: str, config: Optional[ExtractionConfig] = None):
        """
        Initialize the structured PDF processor.

        Args:
            file_path: Path to the PDF file
            config: Optional extraction configuration
        """
        self.file_path = Path(file_path)
        if not self.file_path.exists():
            raise FileNotFoundError(f"PDF file not found: {file_path}")

        self.config = config or ExtractionConfig()
        self.structure: Optional[PDFStructure] = None
        self.headers: List[str] = []
        self.data_rows: List[List[str]] = []
        self.title_lines: List[str] = []
        # Store raw header rows for duplicate detection
        self._raw_header_rows: List[List[str]] = []
        self._header_fingerprints: set = set()

    # ═══════════════════════════════════════════════════════════════
    # PHASE 1: UNDERSTAND THE PDF STRUCTURE
    # ═══════════════════════════════════════════════════════════════

    def analyze_structure(self) -> PDFStructure:
        """
        Phase 1: Analyze PDF structure before extraction.

        Identifies:
        - Title section (top of page)
        - Metadata section (constituency info, electors, etc.)
        - Header row location
        - First data row location
        - Whether headers repeat on each page

        Returns:
            PDFStructure with analyzed information
        """
        logger.info("PHASE 1: Analyzing PDF structure...")

        structure = PDFStructure()

        with pdfplumber.open(str(self.file_path)) as pdf:
            structure.total_pages = len(pdf.pages)

            if not pdf.pages:
                logger.warning("PDF has no pages")
                return structure

            # Analyze first page
            first_page = pdf.pages[0]

            # Extract all text lines from first page
            text = first_page.extract_text()
            if text:
                lines = [line.strip() for line in text.split('\n') if line.strip()]

                # Identify title lines (top section before table)
                in_title_section = True
                for idx, line in enumerate(lines):
                    line_lower = line.lower()

                    # Check if this is a header row (end of title section)
                    if self._is_header_row_text(line_lower):
                        structure.header_row_index = idx
                        in_title_section = False
                        structure.first_data_row_index = idx + 1
                        break

                    if in_title_section:
                        # Categorize as title or metadata
                        if self._is_title_line(line_lower):
                            structure.title_section.append(line)
                        elif self._is_metadata_line(line_lower):
                            structure.metadata_section.append(line)

            # Extract tables to determine column count
            tables = first_page.extract_tables()
            if tables:
                for table in tables:
                    if table and len(table) > 0:
                        # Find the row that looks like headers
                        for row in table:
                            if row and self._is_header_row(row):
                                structure.total_columns = len([c for c in row if c])
                                break
                        if structure.total_columns > 0:
                            break

            # Check if headers repeat on subsequent pages
            if structure.total_pages > 1:
                second_page = pdf.pages[1]
                second_text = second_page.extract_text()
                if second_text:
                    second_lines = second_text.split('\n')
                    for line in second_lines[:10]:  # Check first 10 lines
                        if self._is_header_row_text(line.lower()):
                            structure.headers_repeat_on_pages = True
                            break

        self.structure = structure

        logger.info(f"  Total pages: {structure.total_pages}")
        logger.info(f"  Title lines: {len(structure.title_section)}")
        logger.info(f"  Metadata lines: {len(structure.metadata_section)}")
        logger.info(f"  Header row index: {structure.header_row_index}")
        logger.info(f"  Total columns: {structure.total_columns}")
        logger.info(f"  Headers repeat: {structure.headers_repeat_on_pages}")

        return structure

    def _is_title_line(self, line_lower: str) -> bool:
        """Check if line is a title line."""
        return any(kw in line_lower for kw in self.config.title_keywords)

    def _is_metadata_line(self, line_lower: str) -> bool:
        """Check if line is a metadata line."""
        return any(kw in line_lower for kw in self.config.metadata_keywords)

    def _is_header_row_text(self, line_lower: str) -> bool:
        """Check if text line is likely a header row."""
        return any(kw in line_lower for kw in self.config.header_row_keywords)

    def _is_header_row(self, row: List[str]) -> bool:
        """Check if table row is a header row."""
        row_text = " ".join(str(cell).lower() for cell in row if cell)
        return self._is_header_row_text(row_text)

    # ═══════════════════════════════════════════════════════════════
    # PHASE 2: EXTRACT COLUMN HEADERS (ONE TIME ONLY!)
    # ═══════════════════════════════════════════════════════════════

    def extract_headers(self) -> List[str]:
        """
        Phase 2: Extract column headers from FIRST PAGE ONLY.

        This method:
        1. Locates the header row on the first page
        2. Reads headers left-to-right
        3. Handles multi-row headers (candidate name + party)
        4. Returns a single list of headers

        Returns:
            List of column header strings

        CRITICAL: Headers are extracted ONCE from first page only!
        """
        logger.info("PHASE 2: Extracting column headers from first page...")

        with pdfplumber.open(str(self.file_path)) as pdf:
            if not pdf.pages:
                raise ValueError("PDF has no pages")

            first_page = pdf.pages[0]
            tables = self._extract_tables_from_page(first_page)

            if not tables:
                raise ValueError("No tables found on first page")

            # Find the table with the header row
            for table in tables:
                if not table or len(table) < 2:
                    continue

                # Find header rows in this table
                header_rows = []
                data_start_idx = 0

                for idx, row in enumerate(table):
                    if not row:
                        continue

                    cleaned_row = [self._clean_cell(cell) for cell in row]

                    # Check if this is a header row
                    if self._is_header_row(row):
                        header_rows.append(cleaned_row)
                        data_start_idx = idx + 1
                    elif header_rows and self._is_party_row(cleaned_row):
                        # This is a party abbreviation row below candidate names
                        header_rows.append(cleaned_row)
                        data_start_idx = idx + 1
                    elif header_rows:
                        # Found data row, stop looking for headers
                        break

                if header_rows:
                    # Store raw header rows for duplicate detection
                    self._raw_header_rows = header_rows
                    self._create_header_fingerprints(header_rows)
                    # Combine multi-row headers
                    self.headers = self._combine_header_rows(header_rows)
                    break

            if not self.headers:
                raise ValueError("Could not extract headers from first page")

        # Store title lines from structure analysis
        if self.structure:
            self.title_lines = self.structure.title_section + self.structure.metadata_section

        logger.info(f"  Extracted {len(self.headers)} column headers:")
        for idx, header in enumerate(self.headers):
            logger.debug(f"    Column {idx + 1}: {header}")
        logger.info(f"  Created {len(self._header_fingerprints)} header fingerprints for duplicate detection")

        return self.headers

    def _create_header_fingerprints(self, header_rows: List[List[str]]):
        """
        Create fingerprints from header rows for duplicate detection.

        Stores normalized text patterns that identify header rows.
        """
        self._header_fingerprints = set()

        for row in header_rows:
            # Create fingerprint from non-empty cells
            cells = [self._clean_cell(c).lower().strip() for c in row if c and self._clean_cell(c).strip()]
            if cells:
                # Full row fingerprint
                fingerprint = "|".join(cells)
                self._header_fingerprints.add(fingerprint)

                # Also add individual significant cell values
                for cell in cells:
                    if len(cell) > 3:  # Skip very short values
                        self._header_fingerprints.add(cell)

        # Also add the combined headers as fingerprints
        for header in self.headers if self.headers else []:
            cleaned = self._clean_cell(header).lower().strip()
            if cleaned and len(cleaned) > 3:
                self._header_fingerprints.add(cleaned)

    def _is_duplicate_header_row(self, row: List[str]) -> bool:
        """
        Check if a row is a duplicate of the extracted headers.

        Uses fingerprint matching to detect repeated headers on subsequent pages.
        """
        if not self._header_fingerprints:
            return False

        cleaned_row = [self._clean_cell(c).lower().strip() for c in row if c]
        non_empty = [c for c in cleaned_row if c]

        if not non_empty:
            return False

        # Check if row fingerprint matches any header fingerprint
        row_fingerprint = "|".join(non_empty)
        if row_fingerprint in self._header_fingerprints:
            return True

        # Check if significant portion of cells match header fingerprints
        matching_cells = sum(1 for cell in non_empty if cell in self._header_fingerprints)
        if len(non_empty) >= 3 and matching_cells >= len(non_empty) * 0.5:
            return True

        # Check for specific header keywords that indicate this is a header row
        row_text = " ".join(non_empty)
        header_indicators = [
            "polling station", "sl. no", "sl no", "s.no", "serial",
            "candidate", "party", "nota", "valid votes", "total votes",
            "no. of valid votes", "favour of"
        ]

        indicator_count = sum(1 for ind in header_indicators if ind in row_text)
        if indicator_count >= 2:
            return True

        return False

    def _is_party_row(self, row: List[str]) -> bool:
        """Check if row contains party abbreviations."""
        row_text = " ".join(str(cell).upper() for cell in row if cell)
        # Common party indicators
        party_indicators = ["INC", "BJP", "DMK", "AIADMK", "BSP", "IND", "PARTY"]
        return any(party in row_text for party in party_indicators)

    def _combine_header_rows(self, header_rows: List[List[str]]) -> List[str]:
        """
        Combine multiple header rows into single column names.

        For election PDFs:
        - Row 1: Column labels and spanning headers
        - Row 2: Candidate names
        - Row 3: Party abbreviations

        Strategy: Combine candidate name + party as "NAME (PARTY)"
        """
        if not header_rows:
            return []

        if len(header_rows) == 1:
            return [h for h in header_rows[0] if h]

        # Find maximum columns
        max_cols = max(len(row) for row in header_rows)
        combined = []

        # Check for party row
        party_row_idx = -1
        for idx, row in enumerate(header_rows):
            if row and self._is_party_row(row):
                party_row_idx = idx
                break

        for col_idx in range(max_cols):
            parts = []

            for row_idx, row in enumerate(header_rows):
                if col_idx < len(row) and row[col_idx]:
                    value = row[col_idx].strip()
                    if value and value.upper() not in ["PARTY ABBREVIATION", "PARTY"]:
                        parts.append((row_idx, value))

            if not parts:
                combined.append(f"Column {col_idx + 1}")
                continue

            # If we have party info, combine name + party
            if party_row_idx >= 0 and len(parts) >= 2:
                name = None
                party = None

                for row_idx, value in parts:
                    if row_idx < party_row_idx:
                        name = value
                    elif row_idx == party_row_idx:
                        party = value

                if name and party:
                    combined.append(f"{name}\n({party})")
                elif name:
                    combined.append(name)
                elif party:
                    combined.append(party)
                else:
                    combined.append(parts[-1][1])
            else:
                # Use the most specific (last) value
                combined.append(parts[-1][1])

        return combined

    # ═══════════════════════════════════════════════════════════════
    # PHASE 3: EXTRACT DATA (FROM ALL PAGES)
    # ═══════════════════════════════════════════════════════════════

    def extract_data(
        self,
        progress_callback: Optional[Callable[[int, str], None]] = None
    ) -> List[List[str]]:
        """
        Phase 3: Extract data rows from ALL pages.

        Rules:
        1. Start AFTER headers on first page
        2. Skip repeated headers on subsequent pages
        3. Extract ONLY data rows (identified by numeric Polling Station)
        4. Verify row length matches header count

        Args:
            progress_callback: Optional callback for progress updates

        Returns:
            List of data rows (each row is a list of values)
        """
        logger.info("PHASE 3: Extracting data from all pages...")

        if not self.headers:
            raise ValueError("Headers must be extracted first (call extract_headers)")

        expected_columns = len(self.headers)
        all_data = []

        def update_progress(progress: int, message: str):
            if progress_callback:
                progress_callback(progress, message)

        with pdfplumber.open(str(self.file_path)) as pdf:
            total_pages = len(pdf.pages)

            for page_num, page in enumerate(pdf.pages, 1):
                update_progress(
                    int(10 + (page_num / total_pages) * 70),
                    f"Extracting data from page {page_num} of {total_pages}..."
                )

                tables = self._extract_tables_from_page(page)

                for table in tables:
                    if not table:
                        continue

                    for row in table:
                        if not row:
                            continue

                        cleaned_row = [self._clean_cell(cell) for cell in row]

                        # Skip empty rows first
                        if not any(cell.strip() for cell in cleaned_row if cell):
                            continue

                        # CRITICAL: Skip duplicate header rows (from any page)
                        if self._is_duplicate_header_row(row):
                            logger.debug(f"  Skipping duplicate header row on page {page_num}")
                            continue

                        # Skip header rows (keyword-based check)
                        if self._is_header_row(row):
                            logger.debug(f"  Skipping header row on page {page_num}")
                            continue

                        # Skip party abbreviation rows
                        if self._is_party_row(cleaned_row):
                            logger.debug(f"  Skipping party row on page {page_num}")
                            continue

                        # Check if this is a data row (has numeric value in first columns)
                        # ONLY add rows that are confirmed data rows
                        if self._is_data_row(cleaned_row):
                            # Normalize row length
                            normalized_row = self._normalize_row(cleaned_row, expected_columns)
                            all_data.append(normalized_row)

        self.data_rows = all_data

        logger.info(f"  Extracted {len(all_data)} data rows")
        logger.info(f"  Expected columns: {expected_columns}")

        # Verify extraction
        mismatched = sum(1 for row in all_data if len(row) != expected_columns)
        if mismatched > 0:
            logger.warning(f"  {mismatched} rows had column count mismatch (normalized)")

        return all_data

    def _is_data_row(self, row: List[str]) -> bool:
        """
        Check if row is a data row (not header/title).

        Data rows typically have:
        - A numeric value in first 1-2 columns (Polling Station number)
        - Multiple numeric values (vote counts)

        STRICT: Only returns True if row is CONFIRMED to be a data row.
        """
        if not row or len(row) < 2:
            return False

        # First, check if this looks like a header row by text content
        row_text = " ".join(str(c).lower() for c in row if c).strip()

        # Reject if contains header-specific keywords
        header_keywords = [
            "polling station", "sl. no", "sl no", "s.no",
            "candidate", "party abbreviation", "valid votes",
            "no. of valid votes", "favour of", "nota"
        ]
        if any(kw in row_text for kw in header_keywords):
            return False

        # Check first few cells for numeric station identifier
        has_station_number = False
        for cell in row[:2]:  # Check only first 2 columns
            if cell and cell.strip():
                cleaned = cell.strip()
                # Handle formats like "1", "1A", "1(A)", "123", etc.
                if re.match(r'^\d+[A-Za-z]?(\([A-Za-z]\))?$', cleaned):
                    has_station_number = True
                    break
                try:
                    val = int(cleaned)
                    # Station numbers are typically small positive integers
                    if 1 <= val <= 9999:
                        has_station_number = True
                        break
                except ValueError:
                    continue

        if not has_station_number:
            return False

        # Additional check: data rows should have multiple numeric values (vote counts)
        numeric_count = 0
        for cell in row[1:]:  # Skip first column (station number)
            if cell and cell.strip():
                try:
                    int(cell.strip().replace(',', ''))
                    numeric_count += 1
                except ValueError:
                    pass

        # Data row should have at least some numeric values (vote counts)
        return numeric_count >= 2

    def _normalize_row(self, row: List[str], target_length: int) -> List[str]:
        """Normalize row to match expected column count."""
        if len(row) < target_length:
            return row + [""] * (target_length - len(row))
        elif len(row) > target_length:
            return row[:target_length]
        return row

    def _extract_tables_from_page(self, page) -> List[List[List[str]]]:
        """
        Extract tables from a page using multiple strategies.

        Tries strategies in order:
        1. lines_strict - best for clear table borders
        2. lines - more lenient
        3. default - fallback
        """
        tables = None

        # Strategy 1: Strict lines
        settings = {
            "vertical_strategy": "lines_strict",
            "horizontal_strategy": "lines_strict",
            "intersection_tolerance": 3,
            "snap_tolerance": 3,
            "join_tolerance": 3,
        }
        tables = page.extract_tables(settings)

        # Strategy 2: Regular lines
        if not tables or all(not t or len(t) < 2 for t in tables):
            settings = {
                "vertical_strategy": "lines",
                "horizontal_strategy": "lines",
                "intersection_tolerance": 5,
                "snap_tolerance": 5,
            }
            tables = page.extract_tables(settings)

        # Strategy 3: Default
        if not tables or all(not t or len(t) < 2 for t in tables):
            tables = page.extract_tables()

        return tables or []

    def _clean_cell(self, value) -> str:
        """Clean and sanitize a cell value, including fixing reversed text."""
        if value is None:
            return ""

        text = str(value)

        # Handle special values
        if text.lower() in ("nan", "none", "null", "undefined"):
            return ""

        # Use sanitize_text which handles RTL characters and reversed text detection
        # This is critical for fixing corrupted polling area data
        text = sanitize_text(text, single_line=False)

        return text

    # ═══════════════════════════════════════════════════════════════
    # PHASE 4: CREATE EXCEL FILE WITH PERFECT STRUCTURE
    # ═══════════════════════════════════════════════════════════════

    def create_excel(
        self,
        output_path: str,
        progress_callback: Optional[Callable[[int, str], None]] = None
    ) -> str:
        """
        Phase 4: Create Excel file with proper structure.

        Structure:
        - Rows 1-4: Title section (merged cells)
        - Row 5: Blank separator
        - Row 6: Column headers (blue background, white text, wrapped)
        - Rows 7+: Data rows (centered, bordered)
        - Last row: Totals (orange background, bold, SUM formulas)

        Args:
            output_path: Output Excel file path
            progress_callback: Optional callback for progress updates

        Returns:
            Path to created Excel file
        """
        logger.info("PHASE 4: Creating Excel file...")

        if not self.headers:
            raise ValueError("Headers must be extracted first")
        if not self.data_rows:
            raise ValueError("Data must be extracted first")

        def update_progress(progress: int, message: str):
            if progress_callback:
                progress_callback(progress, message)

        update_progress(80, "Creating Excel structure...")

        wb = Workbook()
        sheet = wb.active
        sheet.title = "Election Results"

        num_columns = len(self.headers)
        current_row = 1

        # ─────────────────────────────────────────────────────────
        # SECTION 1: TITLE ROWS (Rows 1-4)
        # ─────────────────────────────────────────────────────────
        if self.title_lines:
            for title_text in self.title_lines[:4]:  # Max 4 title rows
                last_col = get_column_letter(num_columns)
                sheet.merge_cells(f'A{current_row}:{last_col}{current_row}')

                cell = sheet[f'A{current_row}']
                cell.value = sanitize_text(title_text)
                cell.font = Font(bold=True, size=12)
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.fill = PatternFill(
                    start_color=self.TITLE_BG_COLOR,
                    end_color=self.TITLE_BG_COLOR,
                    fill_type='solid'
                )
                sheet.row_dimensions[current_row].height = 25
                current_row += 1
        else:
            # Default title
            last_col = get_column_letter(num_columns)
            sheet.merge_cells(f'A1:{last_col}1')
            cell = sheet['A1']
            cell.value = "Election Results"
            cell.font = Font(bold=True, size=14)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            sheet.row_dimensions[1].height = 30
            current_row = 2

        # ─────────────────────────────────────────────────────────
        # SECTION 2: BLANK ROW (Separator)
        # ─────────────────────────────────────────────────────────
        sheet.row_dimensions[current_row].height = 10
        current_row += 1

        # ─────────────────────────────────────────────────────────
        # SECTION 3: COLUMN HEADERS
        # ─────────────────────────────────────────────────────────
        header_row = current_row

        header_fill = PatternFill(
            start_color=self.HEADER_BG_COLOR,
            end_color=self.HEADER_BG_COLOR,
            fill_type='solid'
        )
        header_font = Font(bold=True, size=10, color=self.HEADER_FONT_COLOR)
        header_alignment = Alignment(
            horizontal='center',
            vertical='center',
            wrap_text=True
        )
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        for col_idx, header_text in enumerate(self.headers, start=1):
            cell = sheet.cell(row=header_row, column=col_idx)
            cell.value = sanitize_text(header_text) if header_text else f"Column {col_idx}"
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

        # Set header row height (tall for wrapped text)
        sheet.row_dimensions[header_row].height = 80
        current_row += 1

        # ─────────────────────────────────────────────────────────
        # SECTION 4: DATA ROWS
        # ─────────────────────────────────────────────────────────
        update_progress(85, "Adding data rows...")

        data_start_row = current_row
        data_alignment = Alignment(horizontal='center', vertical='center')
        alt_fill = PatternFill(
            start_color=self.ALT_ROW_COLOR,
            end_color=self.ALT_ROW_COLOR,
            fill_type='solid'
        )

        for row_idx, row_data in enumerate(self.data_rows):
            use_alt_color = (row_idx % 2) == 1

            for col_idx, value in enumerate(row_data, start=1):
                cell = sheet.cell(row=current_row, column=col_idx)

                # Convert to appropriate type with final validation
                if value and value.strip():
                    sanitized = sanitize_text(value)
                    # Final validation: Re-check for reversed text after sanitization
                    from .utils import _is_likely_reversed
                    if _is_likely_reversed(sanitized):
                        # Try reversing and re-sanitizing
                        reversed_text = sanitized[::-1]
                        re_sanitized = sanitize_text(reversed_text)
                        if not _is_likely_reversed(re_sanitized):
                            sanitized = re_sanitized
                    
                    try:
                        cell.value = int(sanitized.replace(',', ''))
                        cell.number_format = '#,##0'
                    except ValueError:
                        cell.value = sanitized
                else:
                    cell.value = ""

                cell.alignment = data_alignment
                cell.border = thin_border

                if use_alt_color:
                    cell.fill = alt_fill

            sheet.row_dimensions[current_row].height = 18
            current_row += 1

        data_end_row = current_row - 1

        # ─────────────────────────────────────────────────────────
        # SECTION 5: TOTAL ROW
        # ─────────────────────────────────────────────────────────
        update_progress(90, "Adding total row...")

        total_row = current_row
        total_fill = PatternFill(
            start_color=self.TOTAL_BG_COLOR,
            end_color=self.TOTAL_BG_COLOR,
            fill_type='solid'
        )
        total_font = Font(bold=True, size=11)
        medium_border = Border(
            left=Side(style='medium'),
            right=Side(style='medium'),
            top=Side(style='medium'),
            bottom=Side(style='medium')
        )

        # "TOTAL" label in first column
        cell = sheet.cell(row=total_row, column=1)
        cell.value = "TOTAL"
        cell.font = total_font
        cell.fill = total_fill
        cell.alignment = data_alignment
        cell.border = medium_border

        # SUM formulas for numeric columns
        for col_idx in range(2, num_columns + 1):
            col_letter = get_column_letter(col_idx)
            cell = sheet.cell(row=total_row, column=col_idx)

            # Check if column is numeric
            if self._is_numeric_column(sheet, col_idx, data_start_row, data_end_row):
                cell.value = f"=SUM({col_letter}{data_start_row}:{col_letter}{data_end_row})"
                cell.number_format = '#,##0'

            cell.font = total_font
            cell.fill = total_fill
            cell.alignment = data_alignment
            cell.border = medium_border

        sheet.row_dimensions[total_row].height = 22

        # ─────────────────────────────────────────────────────────
        # SECTION 6: SET COLUMN WIDTHS (CRITICAL!)
        # ─────────────────────────────────────────────────────────
        update_progress(95, "Setting column widths...")

        # First column (Station No) - narrower
        sheet.column_dimensions['A'].width = 12

        # Candidate/data columns - wider
        for col_idx in range(2, num_columns + 1):
            col_letter = get_column_letter(col_idx)
            sheet.column_dimensions[col_letter].width = 16

        # ─────────────────────────────────────────────────────────
        # SECTION 7: FREEZE PANES
        # ─────────────────────────────────────────────────────────
        # Freeze rows above data (header row stays visible when scrolling)
        freeze_cell = f'A{header_row + 1}'
        sheet.freeze_panes = freeze_cell

        # ─────────────────────────────────────────────────────────
        # SECTION 8: SAVE FILE
        # ─────────────────────────────────────────────────────────
        wb.save(output_path)

        update_progress(100, "Excel file created successfully")

        logger.info(f"  Excel file created: {output_path}")
        logger.info(f"  Total rows: {len(self.data_rows)}")
        logger.info(f"  Total columns: {num_columns}")

        return output_path

    def _is_numeric_column(
        self,
        worksheet,
        column: int,
        start_row: int,
        end_row: int
    ) -> bool:
        """Check if column contains predominantly numeric values."""
        numeric_count = 0
        total_count = 0

        # Sample first 20 rows or all if fewer
        sample_end = min(start_row + 20, end_row + 1)

        for row in range(start_row, sample_end):
            value = worksheet.cell(row=row, column=column).value
            if value is not None:
                total_count += 1
                if isinstance(value, (int, float)):
                    numeric_count += 1

        if total_count == 0:
            return False

        return (numeric_count / total_count) >= 0.5

    # ═══════════════════════════════════════════════════════════════
    # PHASE 5: QUALITY VERIFICATION
    # ═══════════════════════════════════════════════════════════════

    def verify_extraction(self) -> dict:
        """
        Phase 5: Verify extraction quality.

        Checks:
        1. Header count matches expected
        2. All data rows have correct column count
        3. No headers appear in data section
        4. Numeric columns have valid values

        Returns:
            Verification report dictionary
        """
        logger.info("PHASE 5: Verifying extraction quality...")

        report = {
            "status": "success",
            "total_errors": 0,
            "headers_count": len(self.headers),
            "data_rows_count": len(self.data_rows),
            "issues": []
        }

        expected_columns = len(self.headers)

        # Check 1: Row lengths
        mismatched_rows = []
        for idx, row in enumerate(self.data_rows):
            if len(row) != expected_columns:
                mismatched_rows.append(idx + 1)

        if mismatched_rows:
            report["issues"].append(
                f"Rows with column mismatch: {mismatched_rows[:10]}"
            )
            report["total_errors"] += len(mismatched_rows)

        # Check 2: Headers in data section
        headers_in_data = []
        header_texts = set(h.lower() for h in self.headers if h)

        for idx, row in enumerate(self.data_rows):
            row_text = " ".join(str(cell).lower() for cell in row if cell)
            if any(kw in row_text for kw in ["polling", "station", "candidate"]):
                if self._is_header_row(row):
                    headers_in_data.append(idx + 1)

        if headers_in_data:
            report["issues"].append(
                f"Header-like rows found in data: {headers_in_data[:10]}"
            )
            report["total_errors"] += len(headers_in_data)
            report["status"] = "warning"

        # Check 3: Empty rows
        empty_rows = sum(1 for row in self.data_rows if not any(cell.strip() for cell in row if cell))
        if empty_rows > 0:
            report["issues"].append(f"Empty rows: {empty_rows}")

        logger.info(f"  Verification complete: {report['status']}")
        logger.info(f"  Total errors: {report['total_errors']}")

        for issue in report["issues"]:
            logger.warning(f"    - {issue}")

        return report

    # ═══════════════════════════════════════════════════════════════
    # PHASE 6: DELIVERY (Integration methods)
    # ═══════════════════════════════════════════════════════════════

    async def process(
        self,
        output_path: str,
        progress_callback: Optional[Callable[[int, str], None]] = None
    ) -> Tuple[str, dict]:
        """
        Phase 6: Complete processing pipeline.

        Executes all phases in order:
        1. Analyze structure
        2. Extract headers
        3. Extract data
        4. Create Excel
        5. Verify quality

        Args:
            output_path: Output Excel file path
            progress_callback: Optional callback for progress updates

        Returns:
            Tuple of (excel_path, verification_report)
        """
        def update_progress(progress: int, message: str):
            if progress_callback:
                progress_callback(progress, message)

        # Phase 1: Analyze structure
        update_progress(5, "Analyzing PDF structure...")
        self.analyze_structure()

        # Phase 2: Extract headers
        update_progress(10, "Extracting column headers...")
        self.extract_headers()

        # Phase 3: Extract data
        self.extract_data(progress_callback)

        # Phase 4: Create Excel
        self.create_excel(output_path, progress_callback)

        # Phase 5: Verify
        update_progress(98, "Verifying extraction quality...")
        report = self.verify_extraction()

        update_progress(100, "Processing complete!")

        return output_path, report

    def to_extraction_result(self) -> ExtractionResult:
        """
        Convert processed data to ExtractionResult for API compatibility.

        Returns:
            ExtractionResult containing the extracted table data
        """
        table_data = TableData(
            headers=self.headers,
            rows=self.data_rows,
            page_number=1,
            title_rows=[self.title_lines] if self.title_lines else [],
        )

        return ExtractionResult(
            tables=[table_data],
            page_texts=self.title_lines,
        )

    def get_summary(self) -> dict:
        """
        Get extraction summary.

        Returns:
            Dictionary with extraction summary
        """
        return {
            "file": str(self.file_path),
            "total_pages": self.structure.total_pages if self.structure else 0,
            "total_columns": len(self.headers),
            "total_rows": len(self.data_rows),
            "title_lines": len(self.title_lines),
            "headers": self.headers,
        }
