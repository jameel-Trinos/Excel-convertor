"""
Convert Polling Stations PDF to Excel Format

This script extracts polling station data from a structured PDF and creates
a professionally formatted Excel file with proper styling.
"""

import re
import sys
from pathlib import Path
from typing import List, Dict, Optional, Tuple

import pdfplumber
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


class PollingStationExtractor:
    """Extract polling station data from PDF using table extraction."""
    
    def __init__(self):
        """Initialize the extractor."""
        # Pattern to match sub-areas: [number] -text
        self.area_pattern = re.compile(r'\[(\d+)\]\s*-\s*([^,\[]+)')
        # Patterns to filter out headers/footers
        self.header_patterns = [
            re.compile(r'^List of Polling Stations', re.IGNORECASE),
            re.compile(r'^Page Number\s*:\s*\d+\s+of\s+\d+', re.IGNORECASE),
            re.compile(r'^\s*Sl\.No\s*\|\s*Polling station No\.', re.IGNORECASE),
            re.compile(r'^---', re.IGNORECASE),
        ]
    
    def extract_tables_from_pdf(self, pdf_path: str) -> List[List]:
        """
        Extract tables from all pages of the PDF.
        
        Args:
            pdf_path: Path to the PDF file
            
        Returns:
            List of table rows (each row is a list of cells)
        """
        all_rows = []
        first_page_headers = None
        
        try:
            with pdfplumber.open(pdf_path) as pdf:
                print(f"📄 PDF has {len(pdf.pages)} pages")
                
                for page_num, page in enumerate(pdf.pages, 1):
                    print(f"   Processing page {page_num}...")
                    
                    # Try multiple extraction strategies
                    tables = None
                    
                    # Strategy 1: lines_strict (best for tables with clear borders)
                    table_settings = {
                        "vertical_strategy": "lines_strict",
                        "horizontal_strategy": "lines_strict",
                        "intersection_tolerance": 3,
                        "snap_tolerance": 3,
                        "join_tolerance": 3,
                    }
                    tables = page.extract_tables(table_settings)
                    
                    # Strategy 2: lines (more lenient)
                    if not tables or all(not t or len(t) < 2 for t in tables):
                        table_settings = {
                            "vertical_strategy": "lines",
                            "horizontal_strategy": "lines",
                            "intersection_tolerance": 5,
                            "snap_tolerance": 5,
                            "join_tolerance": 5,
                        }
                        tables = page.extract_tables(table_settings)
                    
                    # Strategy 3: default extraction
                    if not tables or all(not t or len(t) < 2 for t in tables):
                        tables = page.extract_tables()
                    
                    # Process tables from this page
                    for table in tables:
                        if not table or len(table) < 2:
                            continue
                        
                        # First page: extract headers and data
                        if page_num == 1:
                            # Find header row (usually first row with text headers)
                            header_row_idx = 0
                            for idx, row in enumerate(table):
                                if row and len(row) >= 3:
                                    first_cell = str(row[0]).strip().upper()
                                    if any(keyword in first_cell for keyword in ['SL.', 'SERIAL', 'NO.', 'NUMBER']):
                                        header_row_idx = idx
                                        break
                            
                            # Extract headers
                            if header_row_idx < len(table):
                                headers = [self._clean_cell(cell) for cell in table[header_row_idx]]
                                first_page_headers = headers
                                print(f"      Found headers: {len(headers)} columns")
                            
                            # Extract data rows (skip header row)
                            for row in table[header_row_idx + 1:]:
                                cleaned_row = [self._clean_cell(cell) for cell in row]
                                if cleaned_row and any(cell.strip() for cell in cleaned_row):
                                    all_rows.append(cleaned_row)
                        else:
                            # Subsequent pages: extract data rows only
                            for row in table:
                                cleaned_row = [self._clean_cell(cell) for cell in row]
                                # Skip header-like rows
                                if cleaned_row and len(cleaned_row) >= 3:
                                    first_cell = str(cleaned_row[0]).strip()
                                    # If first cell is numeric, it's a data row
                                    if first_cell.isdigit():
                                        all_rows.append(cleaned_row)
                                    # Skip rows that look like headers
                                    elif any(keyword in first_cell.upper() for keyword in ['SL.', 'SERIAL', 'NO.']):
                                        continue
                                    elif cleaned_row and any(cell.strip() for cell in cleaned_row):
                                        all_rows.append(cleaned_row)
                    
                    print(f"      Extracted {len(all_rows)} total rows so far")
        
        except Exception as e:
            print(f"❌ Error reading PDF: {e}")
            raise
        
        return all_rows
    
    def _clean_cell(self, value) -> str:
        """Clean and sanitize a cell value."""
        if value is None:
            return ""
        
        text = str(value)
        # Remove null bytes and control characters
        text = text.replace("\x00", "")
        text = "".join(char for char in text if ord(char) >= 32 or char in "\n\t")
        # Replace newlines with spaces
        text = text.replace("\n", " ").replace("\r", " ")
        # Normalize whitespace
        text = " ".join(text.split())
        return text.strip()
    
    def parse_table_row(self, row: List[str]) -> Optional[Dict]:
        """
        Parse a table row into an entry dictionary.
        
        Args:
            row: List of cell values from table extraction
            
        Returns:
            Dictionary with entry data or None if not a valid entry
        """
        if not row or len(row) < 3:
            return None
        
        # Expected columns: Sl.No, Polling station No., Location, Polling Areas, Polling Station Type
        # Handle variable number of columns
        try:
            # Try to extract serial number (first column)
            serial_str = str(row[0]).strip()
            if not serial_str or not serial_str.isdigit():
                return None
            
            serial_num = int(serial_str)
            
            # Extract polling station number (second column)
            id_str = str(row[1]).strip() if len(row) > 1 else ""
            if not id_str or not id_str.isdigit():
                # If second column is not a number, try to extract from first column
                # Handle "11", "22" format for entries 1-9
                if len(serial_str) == 2 and serial_str[0] == serial_str[1]:
                    serial_num = int(serial_str[0])
                    id_num = int(serial_str[0])
                else:
                    id_num = serial_num  # Use serial as ID if not found
            else:
                id_num = int(id_str)
            
            # Extract location (third column or combined)
            location = ""
            if len(row) > 2:
                location = str(row[2]).strip()
            
            # Extract polling areas (fourth column or later)
            areas_text = ""
            if len(row) > 3:
                areas_text = str(row[3]).strip()
            elif len(row) > 2:
                # Areas might be in the same column as location
                full_text = str(row[2]).strip()
                # Check if it contains [1] marker
                if '[1]' in full_text:
                    # Split location and areas
                    parts = full_text.split('[1]', 1)
                    location = parts[0].strip()
                    areas_text = '[1]' + parts[1] if len(parts) > 1 else ""
            
            # Extract polling station type (last column or default)
            station_type = "ALL VOTERS"
            if len(row) > 4:
                type_str = str(row[4]).strip().upper()
                if "ALL VOTERS" in type_str:
                    station_type = "ALL VOTERS"
            
            # Parse areas
            areas = []
            if areas_text:
                for area_match in self.area_pattern.finditer(areas_text):
                    area_num = area_match.group(1)
                    area_text = area_match.group(2).strip()
                    # Clean up area text
                    area_text = re.sub(r',\s*$', '', area_text)
                    area_text = ' '.join(area_text.split())
                    areas.append(f"[{area_num}] {area_text}")
            
            # Clean location
            location = location.strip()
            location = re.sub(r'\s+', ' ', location)
            location = re.sub(r',\s*$', '', location)
            
            return {
                'serial_num': serial_num,
                'id_num': id_num,
                'location': location,
                'areas': areas,
                'areas_text': ' | '.join(areas) if areas else '',
                'station_type': station_type
            }
        
        except (ValueError, IndexError) as e:
            # Skip invalid rows
            return None
    
    def extract_entries(self, pdf_path: str) -> List[Dict]:
        """
        Extract all entries from the PDF using table extraction.
        
        Args:
            pdf_path: Path to the PDF file
            
        Returns:
            List of entry dictionaries
        """
        print("📖 Extracting tables from PDF...")
        table_rows = self.extract_tables_from_pdf(pdf_path)
        
        if not table_rows:
            raise ValueError("No table data extracted from PDF")
        
        print("🔍 Parsing entries...")
        entries = []
        
        for row in table_rows:
            entry = self.parse_table_row(row)
            if entry:
                entries.append(entry)
        
        # Sort by serial number
        entries.sort(key=lambda x: x['serial_num'])
        
        print(f"✅ Extracted {len(entries)} entries")
        return entries


class PollingStationExcelCreator:
    """Create formatted Excel file from polling station data."""
    
    def __init__(self):
        """Initialize the Excel creator."""
        # Header styling
        self.header_fill = PatternFill(
            start_color="366092",  # Dark blue
            end_color="366092",
            fill_type="solid"
        )
        self.header_font = Font(bold=True, color="FFFFFF", size=11)
        self.header_alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True
        )
        
        # Data styling
        self.data_font = Font(size=10)
        self.data_alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True
        )
        
        # Border
        self.thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin")
        )
    
    def create_excel(
        self,
        entries: List[Dict],
        output_path: str,
        title: str = "List of Polling Stations"
    ) -> str:
        """
        Create formatted Excel file from entries.
        
        Args:
            entries: List of entry dictionaries
            output_path: Path to save Excel file
            title: Title for the Excel file
            
        Returns:
            Path to created Excel file
        """
        print(f"📊 Creating Excel file: {output_path}")
        
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "Polling Stations"
        
        current_row = 1
        
        # Add title section
        current_row = self._add_title(worksheet, title, current_row)
        
        # Define headers
        headers = [
            "Sl.No",
            "Polling Station No.",
            "Location and Name of Building",
            "Polling Areas",
            "Polling Station Type"
        ]
        
        # Write headers
        header_row = current_row
        for col_idx, header in enumerate(headers, 1):
            cell = worksheet.cell(row=header_row, column=col_idx)
            cell.value = header
            cell.fill = self.header_fill
            cell.font = self.header_font
            cell.alignment = self.header_alignment
            cell.border = self.thin_border
        
        # Set header row height
        worksheet.row_dimensions[header_row].height = 50
        
        current_row += 1
        
        # Write data rows
        data_start_row = current_row
        for entry in entries:
            worksheet.cell(row=current_row, column=1).value = entry['serial_num']
            worksheet.cell(row=current_row, column=2).value = entry['id_num']
            worksheet.cell(row=current_row, column=3).value = entry['location']
            worksheet.cell(row=current_row, column=4).value = entry['areas_text']
            worksheet.cell(row=current_row, column=5).value = entry['station_type']
            
            # Format row
            for col in range(1, len(headers) + 1):
                cell = worksheet.cell(row=current_row, column=col)
                cell.font = self.data_font
                cell.alignment = self.data_alignment
                cell.border = self.thin_border
            
            # Set row height for better readability
            worksheet.row_dimensions[current_row].height = 30
            
            # Alternate row colors for readability
            if current_row % 2 == 0:
                fill = PatternFill(
                    start_color="F2F2F2",
                    end_color="F2F2F2",
                    fill_type="solid"
                )
                for col in range(1, len(headers) + 1):
                    worksheet.cell(row=current_row, column=col).fill = fill
            
            current_row += 1
        
        data_end_row = current_row - 1
        
        # Set column widths
        column_widths = {
            1: 10,   # Sl.No
            2: 15,   # Polling Station No.
            3: 60,   # Location
            4: 80,   # Polling Areas
            5: 20    # Polling Station Type
        }
        
        for col, width in column_widths.items():
            col_letter = get_column_letter(col)
            worksheet.column_dimensions[col_letter].width = width
        
        # Freeze header row
        worksheet.freeze_panes = worksheet.cell(row=header_row + 1, column=1)
        
        # Save workbook
        workbook.save(output_path)
        print(f"✅ Excel file created successfully: {output_path}")
        
        return output_path
    
    def _add_title(self, worksheet, title: str, start_row: int) -> int:
        """
        Add title section to worksheet.
        
        Args:
            worksheet: Worksheet to modify
            title: Title text
            start_row: Starting row
            
        Returns:
            Next available row
        """
        # Merge cells for title
        num_columns = 5
        worksheet.merge_cells(
            start_row=start_row,
            start_column=1,
            end_row=start_row,
            end_column=num_columns
        )
        
        # Format title cell
        title_cell = worksheet.cell(row=start_row, column=1)
        title_cell.value = title
        title_cell.font = Font(bold=True, size=14)
        title_cell.alignment = Alignment(
            horizontal="center",
            vertical="center"
        )
        title_cell.fill = PatternFill(
            start_color="D9E1F2",
            end_color="D9E1F2",
            fill_type="solid"
        )
        
        # Set title row height
        worksheet.row_dimensions[start_row].height = 30
        
        return start_row + 2  # Add spacing row


def main():
    """Main function to convert PDF to Excel."""
    # Get PDF path from command line or use default
    if len(sys.argv) > 1:
        pdf_path = sys.argv[1]
    else:
        # Look for AC001.pdf in current directory
        pdf_path = "AC001.pdf"
        if not Path(pdf_path).exists():
            # Try in parent directory
            pdf_path = Path(__file__).parent / "AC001.pdf"
            if not pdf_path.exists():
                print("❌ Error: PDF file not found")
                print("Usage: python convert_polling_stations.py <pdf_path>")
                sys.exit(1)
    
    pdf_path = str(pdf_path)
    
    if not Path(pdf_path).exists():
        print(f"❌ Error: PDF file not found: {pdf_path}")
        sys.exit(1)
    
    # Create output directory if it doesn't exist
    output_dir = Path("backend/outputs")
    output_dir.mkdir(parents=True, exist_ok=True)
    
    # Generate output filename
    pdf_name = Path(pdf_path).stem
    output_path = output_dir / f"{pdf_name}_converted.xlsx"
    
    print("=" * 80)
    print("POLLING STATIONS PDF TO EXCEL CONVERTER")
    print("=" * 80)
    print(f"Input PDF: {pdf_path}")
    print(f"Output Excel: {output_path}")
    print()
    
    try:
        # Extract entries
        extractor = PollingStationExtractor()
        entries = extractor.extract_entries(pdf_path)
        
        if not entries:
            print("❌ Error: No entries extracted from PDF")
            sys.exit(1)
        
        # Create Excel file
        creator = PollingStationExcelCreator()
        creator.create_excel(entries, str(output_path))
        
        print()
        print("=" * 80)
        print(f"✅ SUCCESS! Excel file created: {output_path}")
        print(f"   Total entries: {len(entries)}")
        print("=" * 80)
        
    except Exception as e:
        print(f"❌ Error: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)


if __name__ == "__main__":
    main()

